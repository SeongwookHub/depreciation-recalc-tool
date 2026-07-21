# -*- coding: utf-8 -*-
"""
고정자산 감가상각비 재계산 스크립트
- sample_asset_ledger.xlsx (또는 회사 고정자산대장)를 읽어 각 자산의 당기(기준일 회계연도)
  감가상각비를 월할상각(정액법) / 국세청 상각률표(정률법) 기준으로 재계산한다.
- 아래 3가지 특수 케이스를 처리한다.
  1) 당기 중 처분(매각) 자산: 처분일까지만 상각
  2) 내용연수 재추정 자산: 재추정 시점의 장부가액을 새 내용연수로 재상각
  3) 내용연수 종료 자산: 더 이상 상각하지 않음
- 회사반영금액과 비교하여 차이나는 자산만 별도 시트에 정리하고 recalc_result.xlsx 로 저장한다.
- 취득원가/잔존가치/내용연수가 계산 불가능한 값(내용연수 0/음수, 잔존가치≥취득원가,
  취득원가 0 이하)인 자산은 재계산에서 제외하고 "데이터오류" 시트에 사유와 함께 모아
  보여준다(다른 정상 자산의 계산에는 영향을 주지 않는다).
"""
import argparse
import datetime as dt
import difflib
import os
import sys
from decimal import Decimal, ROUND_HALF_UP
from typing import Optional

import pandas as pd
import yaml
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

try:
    import anthropic
except ImportError:
    anthropic = None

# Windows 콘솔(PowerShell/cmd)이 시스템 코드페이지(cp949 등)로 표준출력을 열어
# 한글 안내 메시지가 깨지는 것을 막기 위해 표준출력 인코딩을 UTF-8로 고정한다.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

# 재계산결과 시트의 상각률/재계산 수치가 참조하는 엑셀 수식용 참조 시트 이름.
# (도구 내부 구조에 묶여 있는 값이라 config.yaml로 빼지 않는다 — 이름을 바꾸려면
# _inject_recalc_formulas/_write_info_sheet의 수식 문자열도 같이 고쳐야 한다.)
RATE_TABLE_SHEET = "상각률표(별표4)"
INFO_SHEET = "기준정보"

# ---------------------------------------------------------------------------
# 설정 파일(config.yaml) 로딩
# 회사·감사 건마다 달라질 수 있는 값(입출력 경로, 기준일, 중요성 기준액,
# 다기간 비교 연도, AI 모델, 컬럼명 매핑)은 코드가 아니라 config.yaml에서
# 읽어온다. 설정 파일이 없거나 특정 키가 빠져 있어도 에러 없이 아래 기본값으로
# 자연스럽게 대체된다 — "설정 파일은 있으면 좋고 없어도 동작하는" 방식.
# ---------------------------------------------------------------------------
DEFAULT_CONFIG_PATH = "config.yaml"

# COLUMN_MAP의 코드 내장 기본값. config.yaml의 column_map은 이 위에 부분적으로
# 덮어씌워진다(병합) — config에서 빠뜨린 항목은 여기 기본값이 그대로 쓰인다.
_DEFAULT_COLUMN_MAP = {
    "자산명": "자산명",
    "자산분류": "자산분류(유형자산/무형자산)",
    "취득일": "취득일",
    "취득원가": "취득원가",
    "잔존가치": "잔존가치",
    "내용연수": "내용연수(년)",
    "상각방법": "상각방법(정액법/정률법)",
    "회사반영상각비": "회사반영_당기감가상각비",
    "처분일": "처분일",
    "재추정일": "내용연수재추정일",
    "재추정내용연수": "재추정후내용연수(년)",
    "재추정후상각방법": "재추정후상각방법(정액법/정률법)",
    "총예정생산량": "총예정생산량",
    "당기실제생산량": "당기실제생산량",
    "전기말누적생산량": "전기말누적생산량",
    "자본적지출일": "자본적지출일",
    "자본적지출액": "자본적지출액",
    "상각중단시작일": "상각중단시작일",
    "상각중단종료일": "상각중단종료일",
    "회사반영누계상각액": "회사반영_전기말감가상각누계액",
}


def load_config(path: str) -> dict:
    """설정 파일(YAML)을 읽어 dict로 반환한다. 파일이 없거나, 비어 있거나,
    형식이 잘못돼 파싱에 실패하면 안내 메시지만 출력하고 빈 dict를 반환한다
    (호출부는 이 dict를 .get(key, 기존기본값) 형태로만 쓰므로, 빈 dict는
    "설정 파일 없음 = 코드 내장 기본값 전부 사용"과 동일하게 자연스럽게
    동작한다 — 이 함수가 예외를 던져 프로그램을 죽이는 일은 없다).
    """
    if not path or not os.path.exists(path):
        if path and path != DEFAULT_CONFIG_PATH:
            print(f"[!!] 설정 파일을 찾을 수 없어 내장 기본값을 사용합니다: {path}")
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = yaml.safe_load(f)
        return data if isinstance(data, dict) else {}
    except Exception as e:
        print(f"[!!] 설정 파일을 읽는 중 오류가 발생해 내장 기본값을 사용합니다: {path} ({e})")
        return {}


def _cfg_get(config: dict, key: str, default, caster=None):
    """config[key]가 없거나 None이면 default를, 있으면 caster(config[key])를
    반환한다. caster 변환이 실패하면(값 타입이 잘못된 경우 등) 안내를 출력하고
    default로 폴백한다 — 설정 파일에 오탈자/잘못된 값이 있어도 죽지 않는다.
    """
    value = config.get(key)
    if value is None:
        return default
    if caster is None:
        return value
    try:
        return caster(value)
    except (TypeError, ValueError):
        print(f"[!!] 설정값 '{key}'가 올바르지 않아 기본값을 사용합니다: {value!r}")
        return default


def _parse_ref_date(value) -> dt.date:
    """ref_date 설정값을 date로 정규화한다. PyYAML은 따옴표 없는
    2025-12-31 같은 값을 자동으로 datetime.date로 파싱하고, 따옴표를 붙이면
    문자열로 남긴다 — 사용자가 어느 쪽으로 적어도 정상 동작하도록 두 경우와
    datetime.datetime(자정 타임스탬프로 파싱되는 경우)까지 모두 받아들인다.
    """
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    return dt.date.fromisoformat(str(value))


def resolve_settings(config: dict) -> dict:
    """config.yaml에서 읽은 dict(없으면 빈 dict)를 코드 내장 기본값과 합쳐,
    실제로 사용할 설정값 dict를 만든다. 모든 항목이 개별적으로 폴백되므로
    config.yaml에 일부 키만 있어도(또는 아예 없어도) 안전하게 동작한다.
    """
    ref_date = _cfg_get(config, "ref_date", dt.date(2025, 12, 31), _parse_ref_date)
    fy_year = ref_date.year
    cfg_column_map = config.get("column_map")
    column_map = {
        **_DEFAULT_COLUMN_MAP,
        **(cfg_column_map if isinstance(cfg_column_map, dict) else {}),
    }
    return {
        "IN_PATH": _cfg_get(config, "input_path", "sample_asset_ledger.xlsx", str),
        "OUT_PATH": _cfg_get(config, "output_path", "recalc_result.xlsx", str),
        "REF_DATE": ref_date,
        "FY_YEAR": fy_year,
        "MATERIALITY_THRESHOLD": _cfg_get(config, "materiality_threshold", 1_000_000, int),
        "OVERALL_MATERIALITY": _cfg_get(config, "overall_materiality", 50_000_000, int),
        "PERFORMANCE_MATERIALITY": _cfg_get(config, "performance_materiality", 10_000_000, int),
        "ANTHROPIC_MODEL": _cfg_get(config, "anthropic_model", "claude-sonnet-4-6", str),
        "COMPARISON_YEARS": _cfg_get(config, "comparison_years", [fy_year], list),
        "YOY_ANOMALY_THRESHOLD_PCT": _cfg_get(config, "yoy_anomaly_threshold_pct", 20.0, float),
        "COLUMN_MAP": column_map,
    }


_settings = resolve_settings(load_config(DEFAULT_CONFIG_PATH))

# 입출력 파일 경로. `python recalc.py --input 파일.xlsx --output 결과.xlsx`처럼
# 인자로 넘기면 config.yaml의 값보다 우선한다(아래 __main__ 블록 참고).
IN_PATH = _settings["IN_PATH"]
OUT_PATH = _settings["OUT_PATH"]

# 기준일: 이 날짜가 속한 회계연도(1/1~12/31)의 당기 감가상각비를 재계산한다.
REF_DATE = _settings["REF_DATE"]
FY_YEAR = _settings["FY_YEAR"]

# 중요성 기준 금액(=AMPT, Audit Misstatement Posting Threshold, 원): 회사반영 대비
# 재계산 차이의 절대값이 이 금액 미만이면 "경미한 차이", 이상이면 "유의한 차이"로
# 구분한다(재계산결과 시트의 실제 판정 로직에 쓰이는 값은 이 상수뿐이다).
MATERIALITY_THRESHOLD = _settings["MATERIALITY_THRESHOLD"]

# 아래 두 값은 판정 로직에는 쓰이지 않고, 기준정보 시트에 참고용으로만 표시된다
# (전체 재무제표 수준 중요성과 수행중요성 — 감사 실무에서 통상 함께 보는 지표).
OVERALL_MATERIALITY = _settings["OVERALL_MATERIALITY"]  # 중요성
PERFORMANCE_MATERIALITY = _settings["PERFORMANCE_MATERIALITY"]  # 수행중요성(PM)

# "유의한 차이" 자산의 AI 추정원인 코멘트 생성에 사용할 모델.
ANTHROPIC_MODEL = _settings["ANTHROPIC_MODEL"]

# 다기간(연도별) 비교 대상 회계연도 목록. 기본값처럼 [FY_YEAR] 하나만 두면
# 기존과 완전히 동일하게 단일 연도만 계산한다. 여러 연도 추이를 보려면
# config.yaml의 comparison_years에 예: [2023, 2024, 2025] 처럼 지정한다.
COMPARISON_YEARS = _settings["COMPARISON_YEARS"]

# 다기간 비교에서 전년대비 이 비율(%) 이상 증감하면 "경고"로 표시한다.
YOY_ANOMALY_THRESHOLD_PCT = _settings["YOY_ANOMALY_THRESHOLD_PCT"]

# ---------------------------------------------------------------------------
# 컬럼명 매핑
# 회사마다 고정자산대장의 컬럼명이 다를 수 있으므로, config.yaml의 column_map만
# 바꿔 지정하면 스크립트 본문은 수정할 필요가 없다.
# 예: 회사 파일에서 취득원가 컬럼이 "취득가액"이라는 이름이면
#     column_map의 "취득원가": "취득가액" 으로 바꾸면 된다.
#
# 처분일 / 내용연수재추정일 / 재추정후내용연수 는 선택 항목이다.
# 회사 파일에 해당 컬럼 자체가 없으면(매핑된 이름이 파일에 없으면) 모든 자산에
# 처분/재추정이 없는 것으로 간주하고 계속 진행한다.
# ---------------------------------------------------------------------------
COLUMN_MAP = _settings["COLUMN_MAP"]

REQUIRED_KEYS = [
    "자산명", "자산분류", "취득일", "취득원가", "잔존가치",
    "내용연수", "상각방법", "회사반영상각비",
]
OPTIONAL_KEYS = ["처분일", "재추정일", "재추정내용연수", "재추정후상각방법",
                 "총예정생산량", "당기실제생산량", "전기말누적생산량",
                 "자본적지출일", "자본적지출액", "상각중단시작일", "상각중단종료일",
                 "회사반영누계상각액"]

# ---------------------------------------------------------------------------
# 컬럼명 동의어 사전
# COLUMN_MAP에 지정한 이름이 파일에 없을 때, 여기 등록된 이름 중 하나가 있으면
# 자동으로 그 컬럼을 쓴다(회사마다 "자산명" 대신 "자산코드"/"품목명"을 쓰는 식의
# 흔한 표기 차이를 코드 수정 없이 흡수하기 위함). 정확 일치가 항상 우선이고,
# 여기 없는 이름은 여전히 못 찾는 게 정상이다 — 예측 불가능한 추측을 자동으로
# 반영하지 않기 위해 목록은 의도적으로 보수적으로 유지한다. 특히 "장부가액"은
# 취득원가와 다른 개념(순장부가액)일 수 있어 일부러 어느 필드의 동의어에도
# 넣지 않는다 — 이런 경우는 COLUMN_MAP을 사람이 직접 판단해서 고쳐야 한다.
# ---------------------------------------------------------------------------
COLUMN_SYNONYMS = {
    "자산명": ["자산명", "자산코드", "자산코드/명", "품목명", "품목", "자산 이름"],
    "자산분류": ["자산분류", "자산구분", "구분", "자산분류(유형자산/무형자산)", "계정과목"],
    "취득일": ["취득일", "취득일자", "취득년월일"],
    "취득원가": ["취득원가", "취득가액", "취득금액"],
    "잔존가치": ["잔존가치", "잔존가액", "잔가"],
    "내용연수": ["내용연수", "내용년수", "내용연수(년)", "내용년수(년)"],
    "상각방법": ["상각방법", "감가상각방법", "상각방법(정액법/정률법)"],
    "회사반영상각비": ["회사반영_당기감가상각비", "당기상각비", "당기감가상각비", "회사계상상각비"],
    "처분일": ["처분일", "처분일자", "매각일"],
    "재추정일": ["내용연수재추정일", "재추정일", "재추정일자"],
    "재추정내용연수": ["재추정후내용연수(년)", "재추정내용연수", "재추정후내용연수"],
    "재추정후상각방법": ["재추정후상각방법(정액법/정률법)", "재추정후상각방법", "변경후상각방법"],
    "총예정생산량": ["총예정생산량", "총생산예정량", "총추정생산량"],
    "당기실제생산량": ["당기실제생산량", "당기생산량", "당기실제생산"],
    "전기말누적생산량": ["전기말누적생산량", "전기말누계생산량"],
    "자본적지출일": ["자본적지출일", "자본적지출일자"],
    "자본적지출액": ["자본적지출액", "자본적지출금액"],
    "상각중단시작일": ["상각중단시작일", "감가상각중단시작일", "상각중단개시일"],
    "상각중단종료일": ["상각중단종료일", "감가상각중단종료일"],
    "회사반영누계상각액": ["회사반영_전기말감가상각누계액", "전기말감가상각누계액", "기초감가상각누계액"],
}

# 상각방법 값 표기 차이("정액"/"정률"처럼 줄인 표기, 영문 약자 등)를 정규화하는
# 사전. 공백 제거 + 대소문자 무시 후 이 표에서 찾는다(normalize_method 참고).
# 여기 없는 값은 정규화하지 않고 원본 그대로 반환한다(validate_asset_inputs가
# "인식 불가"로 잡아 데이터오류 시트로 격리하지, 배치 전체를 죽이지 않는다).
METHOD_ALIASES = {
    "정액법": "정액법", "정액": "정액법", "sl": "정액법", "straightline": "정액법",
    "정률법": "정률법", "정률": "정률법", "db": "정률법", "decliningbalance": "정률법",
    "이중체감법": "이중체감법", "이중체감": "이중체감법", "ddb": "이중체감법",
    "doubledeclining": "이중체감법", "이중체감잔액법": "이중체감법",
    "연수합계법": "연수합계법", "연수합계": "연수합계법", "syd": "연수합계법",
    "sumofyearsdigits": "연수합계법",
    "생산량비례법": "생산량비례법", "생산량비례": "생산량비례법", "uop": "생산량비례법",
    "unitsofproduction": "생산량비례법",
}

KNOWN_METHODS = ("정액법", "정률법", "이중체감법", "연수합계법", "생산량비례법")

# 법인세법 시행규칙 [별표4] 감가상각자산의 상각률표(내용연수 2~60년, 정률법 할푼리를
# 소수로 변환한 값). 표에 없는 내용연수(1년 이하, 60년 초과)만 잔존가치 5% 가정
# (r = 1-0.05**(1/n))으로 근사한다.
RATE_TABLE = {
    2: 0.777, 3: 0.632, 4: 0.528, 5: 0.451, 6: 0.394,
    7: 0.349, 8: 0.313, 9: 0.284, 10: 0.259,
    11: 0.239, 12: 0.221, 13: 0.206, 14: 0.193, 15: 0.182,
    16: 0.171, 17: 0.162, 18: 0.154, 19: 0.146, 20: 0.140,
    21: 0.133, 22: 0.128, 23: 0.123, 24: 0.118, 25: 0.113,
    26: 0.109, 27: 0.106, 28: 0.102, 29: 0.099, 30: 0.096,
    31: 0.093, 32: 0.090, 33: 0.087, 34: 0.085, 35: 0.083,
    36: 0.080, 37: 0.078, 38: 0.076, 39: 0.074, 40: 0.073,
    41: 0.071, 42: 0.069, 43: 0.068, 44: 0.066, 45: 0.065,
    46: 0.064, 47: 0.062, 48: 0.061, 49: 0.060, 50: 0.059,
    51: 0.058, 52: 0.056, 53: 0.055, 54: 0.054, 55: 0.054,
    56: 0.053, 57: 0.052, 58: 0.051, 59: 0.050, 60: 0.049,
}

# 위 표와 짝을 이루는 정액법 할푼리(참고/엑셀 상각률표 시트 표시용 — 실제 정액법
# 계산에는 쓰지 않는다. 정액법은 1/내용연수를 그대로 쓰는 것이 사용자 요청 사항).
STRAIGHT_LINE_RATE_TABLE = {
    2: 0.500, 3: 0.333, 4: 0.250, 5: 0.200, 6: 0.166,
    7: 0.142, 8: 0.125, 9: 0.111, 10: 0.100,
    11: 0.090, 12: 0.083, 13: 0.076, 14: 0.071, 15: 0.066,
    16: 0.062, 17: 0.058, 18: 0.055, 19: 0.052, 20: 0.050,
    21: 0.048, 22: 0.046, 23: 0.044, 24: 0.042, 25: 0.040,
    26: 0.039, 27: 0.037, 28: 0.036, 29: 0.035, 30: 0.034,
    31: 0.033, 32: 0.032, 33: 0.031, 34: 0.030, 35: 0.029,
    36: 0.028, 37: 0.027, 38: 0.027, 39: 0.026, 40: 0.025,
    41: 0.025, 42: 0.024, 43: 0.024, 44: 0.023, 45: 0.023,
    46: 0.022, 47: 0.022, 48: 0.021, 49: 0.021, 50: 0.020,
    51: 0.020, 52: 0.020, 53: 0.019, 54: 0.019, 55: 0.019,
    56: 0.018, 57: 0.018, 58: 0.018, 59: 0.017, 60: 0.017,
}


def classify_materiality(diff: int, threshold: int = MATERIALITY_THRESHOLD) -> str:
    return "유의한 차이" if abs(diff) >= threshold else "경미한 차이"


def normalize_method(raw) -> Optional[str]:
    """
    상각방법 표기를 METHOD_ALIASES 사전으로 정규화한다("정액" -> "정액법" 등).
    공백을 지우고 소문자로 바꾼 뒤 사전에서 찾으므로 "정액법"/"정액"/"SL"/" sl "
    모두 같은 값으로 취급한다. 사전에 없는 값은 원본을 그대로(strip만 해서)
    돌려준다 — 여기서 조용히 실패시키지 않고, validate_asset_inputs가 KNOWN_METHODS
    기준으로 "인식 불가"임을 명시적으로 잡아 데이터오류 시트로 격리하게 한다.
    """
    if raw is None:
        return raw
    text = str(raw).strip()
    key = text.replace(" ", "").lower()
    return METHOD_ALIASES.get(key, text)


def validate_asset_inputs(cost: float, salvage: float, life: int,
                           method: str = None, total_units: float = None,
                           reest_method: str = None,
                           capex_date=None, capex_amount: float = None,
                           susp_start=None, susp_end=None,
                           accum_reported: float = None,
                           prior_period_units: float = None) -> list:
    """
    재계산에 들어가기 전 취득원가/잔존가치/내용연수가 계산 가능한 값인지 검증한다.
    문제가 있으면 오류 사유 문자열 리스트를, 문제가 없으면 빈 리스트를 반환한다.
    여기서 걸러진 자산은 recalc_asset을 호출하지 않으므로(0/음수 내용연수로 인한
    ZeroDivisionError 등), 다른 자산의 계산에 영향을 주지 않고 안전하게 제외된다.

    생산량비례법(method="생산량비례법")은 내용연수를 계산에 쓰지 않고 대신
    총예정생산량을 쓰므로, 이 경우엔 내용연수 검증 대신 총예정생산량을 검증한다.

    자본적지출/상각중단/재추정후상각방법은 자산에 이벤트가 있을 때만 값이 채워지므로
    (선택 컬럼), 각각 "짝이 맞는지"와 "값 자체가 유효한지"만 검증한다.
    """
    errors = []
    if method is not None and method not in KNOWN_METHODS:
        errors.append(
            f"상각방법 오류(값={method}, 인식 가능한 값: "
            f"{'/'.join(KNOWN_METHODS)} 또는 그 약칭(정액/정률/이중체감/연수합계/생산량비례 등))"
        )
    if method != "생산량비례법" and life < 1:
        errors.append(f"내용연수 오류(내용연수={life}년, 1년 이상의 정수여야 함)")
    if cost <= 0:
        errors.append(f"취득원가 오류(취득원가={cost:,.0f}원, 0보다 커야 함)")
    if salvage >= cost:
        errors.append(f"잔존가치 오류(잔존가치={salvage:,.0f}원, 취득원가={cost:,.0f}원 미만이어야 함)")
    if method == "생산량비례법" and (total_units is None or total_units <= 0):
        errors.append(f"총예정생산량 오류(총예정생산량={total_units}, 0보다 커야 함)")
    if reest_method is not None and reest_method not in ("정액법", "정률법", "이중체감법", "연수합계법"):
        errors.append(f"재추정후상각방법 오류(값={reest_method}, 정액법/정률법/이중체감법/연수합계법 중 하나여야 함)")
    if (capex_date is None) != (capex_amount is None):
        errors.append("자본적지출 오류(자본적지출일과 자본적지출액은 함께 입력되어야 함)")
    elif capex_amount is not None and capex_amount <= 0:
        errors.append(f"자본적지출액 오류(자본적지출액={capex_amount:,.0f}원, 0보다 커야 함)")
    if (susp_start is None) != (susp_end is None):
        errors.append("상각중단기간 오류(상각중단시작일과 상각중단종료일은 함께 입력되어야 함)")
    elif susp_start is not None and susp_end is not None and susp_end < susp_start:
        errors.append(f"상각중단기간 오류(상각중단종료일={susp_end}이 상각중단시작일={susp_start}보다 빠름)")
    if accum_reported is not None and accum_reported < 0:
        errors.append(f"회사반영_전기말감가상각누계액 오류(값={accum_reported:,.0f}원, 0 이상이어야 함)")
    if prior_period_units is not None and prior_period_units < 0:
        errors.append(f"전기말누적생산량 오류(값={prior_period_units:,.0f}, 0 이상이어야 함)")
    return errors


def get_ai_estimated_cause(diff: int, elapsed_months: int, method: str,
                            is_reestimated: bool, is_disposed: bool) -> str:
    """
    "유의한 차이" 자산에 대해 Claude API로 차이 발생 원인을 한 줄로 추정한다.
    API 키가 없거나(.env 미설정) 호출 중 오류가 발생하면 이 기능만 건너뛰고
    "-"를 반환한다(나머지 재계산 로직에는 영향을 주지 않는다).
    """
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key or anthropic is None:
        return "-"

    try:
        client = anthropic.Anthropic(api_key=api_key)
        prompt = (
            "고정자산 감가상각비 재계산 검증에서 아래 자산의 회사반영 감가상각비와 "
            "재계산 감가상각비 사이에 '유의한 차이'가 발생했습니다.\n"
            f"- 차이금액(재계산-회사반영): {diff:,}원\n"
            f"- 경과개월수(취득일~기준일): {elapsed_months}개월\n"
            f"- 상각방법: {method}\n"
            f"- 내용연수 재추정 여부: {'있음' if is_reestimated else '없음'}\n"
            f"- 당기중 처분 여부: {'있음' if is_disposed else '없음'}\n\n"
            "회계/세무 실무자 관점에서 이런 차이가 발생했을 가능성이 가장 높은 원인을 "
            "한국어 한 줄(50자 이내)로 추정해서 답변하세요. 결론 한 줄 외 다른 설명은 출력하지 마세요."
        )
        response = client.messages.create(
            model=ANTHROPIC_MODEL,
            max_tokens=200,
            messages=[{"role": "user", "content": prompt}],
        )
        text = next((b.text for b in response.content if b.type == "text"), "").strip()
        return text if text else "-"
    except Exception as e:
        print(f"  [경고] AI 추정원인 호출 실패, 이 자산은 건너뜁니다: {e}")
        return "-"


def get_rule_based_cause(is_reestimated: bool, is_disposed: bool, note: str,
                          is_capex: bool = False, is_suspended: bool = False) -> Optional[str]:
    """
    이미 계산되어 있는 처분/재추정/자본적지출/상각중단/내용연수종료 정보만으로 원인이
    명확히 설명되는 케이스를 규칙으로 1차 분류한다. 분류 가능하면 원인 문자열을,
    애매하면 None을 반환해서 호출부가 AI 호출(get_ai_estimated_cause)로 폴백하도록
    한다(모든 API 호출을 규칙으로 대체할 수 있어 비용/속도를 아낀다).
    """
    if is_disposed and "당기중처분" in note:
        return "처분일 반영 오류 추정(처분일까지의 월할상각 미반영 가능성)"
    if is_disposed and "전기이전처분" in note:
        return "전기 이전 처분자산에 당기 상각비를 계속 반영했을 가능성"
    if is_reestimated:
        return "내용연수 재추정 반영 오류 추정(재추정시점 장부가액/신규 내용연수 미반영 가능성)"
    if is_capex and "자본적지출" in note:
        return "자본적지출 반영 오류 추정(지출액 가산 또는 잔여내용연수 재상각 미반영 가능성)"
    if is_suspended and "상각중단" in note:
        return "감가상각중단 반영 오류 추정(중단기간 0원 처리 또는 종료시점 연장 미반영 가능성)"
    if "내용연수종료" in note:
        return "내용연수 종료 자산에 상각비를 계속 반영했을 가능성"
    return None


def get_rate(life: int) -> float:
    if life in RATE_TABLE:
        return RATE_TABLE[life]
    return round(1 - 0.05 ** (1 / life), 3)


def _method_rate(method: str, life_years: int):
    """정률법/이중체감법처럼 '기초 장부가액 x 상각률'을 매년 곱해서 상각하는
    방법에서 쓸 상각률을 구한다. 정률법은 국세청 상각률표(get_rate), 이중체감법은
    정액법 상각률(1/내용연수)의 2배(2/내용연수)를 그대로 쓴다(세법상 표가 없다).
    그 외 방법(정액법/연수합계법/생산량비례법)은 이 상각률 개념 자체를 안 쓰므로 None."""
    if method == "정률법":
        return get_rate(life_years)
    if method == "이중체감법":
        return 2.0 / life_years
    return None


def _floor_threshold_for(seg: dict) -> float:
    """정률법/이중체감법 연단위 상각 루프에서 '미상각잔액이 이 값 이하가 되는
    과세기간에 잔여 전액을 상각'하는 특례의 기준값. 정률법은 세법상 취득가액
    (자본적지출 누계 반영)의 5%, 이중체감법은 그런 세법 특례가 없으므로 단순히
    잔존가치 자체를 하한으로 쓴다(장부가액이 잔존가치 밑으로 내려가지 않게만 함)."""
    if seg["method"] == "정률법":
        return seg["tax_cost"] * 0.05
    return seg["salvage"]


def month_index(year: int, month: int) -> int:
    return year * 12 + month


def idx_to_year(idx: int) -> int:
    """month_index(year, month)의 역함수 중 연도만 구한다(month는 1~12 범위 가정)."""
    return (idx - 1) // 12


def round_won(x: float) -> int:
    return int(Decimal(str(x)).quantize(0, rounding=ROUND_HALF_UP))


def to_date_or_none(v):
    if v is None or (isinstance(v, float) and pd.isna(v)) or pd.isna(v):
        return None
    return pd.to_datetime(v).date()


def elapsed_months_to_ref(acq: dt.date, ref: dt.date) -> int:
    """취득일부터 기준일까지 경과 개월수(취득월 포함, 월 단위 카운트)."""
    return month_index(ref.year, ref.month) - month_index(acq.year, acq.month) + 1


def fy_ref_date(year: int) -> dt.date:
    """
    다기간 비교에서 각 회계연도의 기준일을 반환한다. REF_DATE의 연도와 같으면
    REF_DATE를 그대로 쓰고(12/31이 아닌 기준일 설정도 존중), 그 외 연도는
    해당 연도의 12/31을 기준일로 본다.
    """
    return REF_DATE if year == FY_YEAR else dt.date(year, 12, 31)


# ---------------------------------------------------------------------------
# 정액법 재계산
# ---------------------------------------------------------------------------
def build_straight_line_segments(acq, cost, salvage, life_years, reest_date, reest_life_years):
    """
    정액법 상각 스케줄을 '구간(segment)' 목록으로 만든다.
    - 재추정이 없으면 구간은 1개: [취득월 ~ 취득월+내용연수*12-1], 기준가액=취득원가.
    - 내용연수 재추정이 있으면, 재추정월 직전까지가 1구간(기존 내용연수 기준),
      재추정월부터는 그 시점의 장부가액을 새 기준가액으로 하는 2구간(재추정후내용연수 기준)이
      새로 시작된다. 즉 "재추정 시점부터 남은 장부가액을 새 내용연수로 재상각"을 그대로 구현한다.
    """
    start_idx = month_index(acq.year, acq.month)
    life_months = life_years * 12

    if reest_date is None:
        return [dict(start_idx=start_idx, end_idx=start_idx + life_months - 1,
                      basis=cost, salvage=salvage, life_months=life_months)]

    reest_idx = month_index(reest_date.year, reest_date.month)
    monthly_dep1 = (cost - salvage) / life_months
    # 재추정 시점까지 원래 스케줄로 상각된 개월수(생애 종료를 넘지 않도록 방어)
    months_before_reest = max(0, min(reest_idx - start_idx, life_months))
    bv_at_reest = cost - monthly_dep1 * months_before_reest

    seg1 = dict(start_idx=start_idx, end_idx=reest_idx - 1,
                basis=cost, salvage=salvage, life_months=life_months)

    new_life_months = reest_life_years * 12
    seg2 = dict(start_idx=reest_idx, end_idx=reest_idx + new_life_months - 1,
                basis=bv_at_reest, salvage=salvage, life_months=new_life_months)
    return [seg1, seg2]


def straight_line_current_period_dep(segments, disposal_idx, fy_year):
    """
    당기 회계연도와 각 구간이 겹치는 개월수만큼 정액상각비를 계산해 합산한다.
    - 처분일이 있으면 당기 말(fy_end_idx)을 처분월로 앞당겨, 처분월 이후 개월은
      상각 개월수 계산에서 자동으로 제외된다(처분일까지만 상각).
    - 내용연수(구간의 end_idx)를 넘어서는 달은 겹침이 없어 자동으로 0개월 처리되므로
      내용연수가 끝난 자산은 별도 분기 없이도 더 이상 상각되지 않는다.
    """
    fy_start_idx = month_index(fy_year, 1)
    fy_end_idx = month_index(fy_year, 12)
    if disposal_idx is not None:
        fy_end_idx = min(fy_end_idx, disposal_idx)

    total_dep = 0.0
    total_months = 0
    for seg in segments:
        overlap_start = max(seg["start_idx"], fy_start_idx)
        overlap_end = min(seg["end_idx"], fy_end_idx)
        if overlap_start > overlap_end:
            continue
        months = overlap_end - overlap_start + 1
        monthly_dep = (seg["basis"] - seg["salvage"]) / seg["life_months"]
        total_dep += monthly_dep * months
        total_months += months
    return total_dep, total_months


# ---------------------------------------------------------------------------
# 정률법 재계산
# ---------------------------------------------------------------------------
def declining_balance_current_period_dep(acq, cost, salvage, life_years,
                                          reest_date, reest_life_years,
                                          disposal_date, fy_year):
    """
    정률법은 매년 '기초 장부가액 x 상각률'을 연 단위로 누적 적용하고,
    취득 첫 해/처분되는 해만 월할 비례한다(국세청 상각률표를 적용하는 일반적인 실무 방식과 동일).

    - 내용연수 재추정: 재추정일이 속한 회계연도의 1월 1일부터 새 내용연수/새 상각률을
      적용하는 것으로 단순화한다(재추정은 통상 회계연도 초부터 적용하는 것이 일반적이며,
      정액법과 달리 정률법은 상각률이 연 단위 장부가액에 곱해지는 구조라 재추정일을
      회계연도 중간의 특정 월로 정밀하게 나누는 것이 실무적 의미가 크지 않기 때문).
    - 처분: 처분월까지만 그 해의 상각 개월수에 포함한다.
    - 내용연수 종료: 그 해의 상각 종료월(active_end_idx)을 넘어서면 개월수가 0이 되어
      더 이상 상각되지 않는다.
    - 5% 특례(법인세법 시행규칙 [별표4] 관련 규정): 미상각잔액이 취득가액의 5% 이하로
      최초로 떨어지는 과세기간에 잔여 전액(회사가 입력한 잔존가치까지)을 그 해에
      상각한다. 회사가 입력한 잔존가치는 이 바닥값(5%)에 도달하기 전까지는 상각액
      계산에 영향을 주지 않는다.
    """
    rate = get_rate(life_years)
    start_idx = month_index(acq.year, acq.month)
    active_end_idx = start_idx + life_years * 12 - 1  # 상각 종료월(생애 종료월)
    active_rate = rate
    five_pct = cost * 0.05

    reest_year = reest_date.year if reest_date is not None else None
    disposal_idx = month_index(disposal_date.year, disposal_date.month) if disposal_date is not None else None

    book_value = float(cost)
    current_period_dep = 0.0
    current_period_months = 0
    opening_book_value = float(cost)

    y = acq.year
    while y <= fy_year:
        if reest_year is not None and y == reest_year:
            active_rate = get_rate(reest_life_years)
            active_end_idx = month_index(y, 1) + reest_life_years * 12 - 1

        if y == fy_year:
            opening_book_value = book_value

        year_start_idx = month_index(y, 1)
        year_end_idx = month_index(y, 12)
        eff_start = max(start_idx, year_start_idx)
        eff_end = min(active_end_idx, year_end_idx)
        if disposal_idx is not None:
            eff_end = min(eff_end, disposal_idx)

        months = eff_end - eff_start + 1 if eff_end >= eff_start else 0
        dep = 0.0
        if months > 0:
            dep = book_value * active_rate * months / 12
            if book_value - dep <= five_pct:
                dep = book_value - salvage
            book_value -= dep

        if y == fy_year:
            current_period_dep = dep
            current_period_months = months
        y += 1

    return max(current_period_dep, 0.0), current_period_months, opening_book_value


# ---------------------------------------------------------------------------
# 이벤트 기반 통합 스케줄 (자본적지출 / 재추정+방법변경 / 상각중단)
# 정액법 전용 함수(build_straight_line_segments) 하나로는 "재추정으로 상각방법이
# 정률법으로 바뀌는" 케이스를 표현할 수 없으므로, 여기서는 "구간(segment)마다
# method가 다를 수 있는" 통합 스케줄을 만든다. 이 경로는 자본적지출/상각중단/
# 방법이 실제로 바뀌는 재추정이 하나라도 있는 자산에 대해서만 사용되고
# (recalc_asset의 has_extended_events 분기), 그 외의 기존 자산은 기존
# build_straight_line_segments/declining_balance_current_period_dep를 그대로 탄다.
# ---------------------------------------------------------------------------
def _declining_balance_year_loop(start_idx, basis, salvage, rate, active_end_idx,
                                  disposal_idx, susp_start_idx, susp_end_idx, up_to_idx,
                                  floor_threshold=None):
    """
    declining_balance_current_period_dep(위)의 연단위 누적 로직을, 특정 구간이
    start_idx부터 up_to_idx(포함)까지 상각됐을 때의 (최종 장부가액, up_to_idx가
    속한 회계연도의 그 해 상각비, 그 해 상각개월수)를 구하도록 일반화한 버전이다.
    상각중단과 겹치는 개월을 그 해의 상각개월수에서 제외하는 점만 원본과 다르다.
    정률법/이중체감법(둘 다 '기초 장부가액 x 상각률'을 매년 곱하는 구조) 공통으로 쓴다.

    floor_threshold: 미상각잔액이 이 금액 이하로 내려가는 과세기간에 잔여 전액을
    상각하는 특례의 기준값(호출부가 _floor_threshold_for로 미리 계산해서 넘긴다 —
    정률법은 세무상 취득가액의 5%, 이중체감법은 잔존가치). 지정하지 않으면 basis의
    5%를 쓴다(과거 호출부 호환용 기본값).
    """
    if up_to_idx < start_idx:
        return float(basis), 0.0, 0

    threshold = floor_threshold if floor_threshold is not None else basis * 0.05
    start_year = idx_to_year(start_idx)
    up_to_year = idx_to_year(up_to_idx)

    book_value = float(basis)
    dep_in_target_year, months_in_target_year = 0.0, 0

    y = start_year
    while y <= up_to_year:
        year_start_idx = month_index(y, 1)
        year_end_idx = month_index(y, 12)
        eff_start = max(start_idx, year_start_idx)
        eff_end = min(active_end_idx, year_end_idx)
        if disposal_idx is not None:
            eff_end = min(eff_end, disposal_idx)
        if y == up_to_year:
            eff_end = min(eff_end, up_to_idx)

        months = eff_end - eff_start + 1 if eff_end >= eff_start else 0
        if susp_start_idx is not None and months > 0:
            s, e = max(eff_start, susp_start_idx), min(eff_end, susp_end_idx)
            if s <= e:
                months -= (e - s + 1)

        dep = 0.0
        if months > 0:
            dep = book_value * rate * months / 12
            if book_value - dep <= threshold:
                dep = book_value - salvage
            book_value -= dep

        if y == up_to_year:
            dep_in_target_year, months_in_target_year = dep, months
        y += 1

    return book_value, dep_in_target_year, months_in_target_year


def _syd_year_loop(start_idx, basis, salvage, life_years, origin_start_idx, active_end_idx,
                    disposal_idx, susp_start_idx, susp_end_idx, up_to_idx):
    """
    연수합계법(SYD) 버전의 _declining_balance_year_loop. 정률법/이중체감법과 달리
    상각률이 매년 달라진다 — origin_start_idx를 기준으로 몇 번째 해(k)인지에 따라
    그 해의 몫이 (내용연수-k+1)/(내용연수x(내용연수+1)/2)로 줄어든다(등차수열의 합
    공식). 이 몫은 그 해의 '기초 장부가액'이 아니라 항상 고정된 basis(=상각대상
    구간의 기준가액)에 곱해진다는 점이 정률법/이중체감법과 다르다(SYD 정의 자체가
    '기준가액 x 그 해의 몫'이지 '장부가액 x 상각률'이 아니기 때문).

    origin_start_idx: 이 SYD 스케줄의 1년차가 시작되는 달력월. 재추정이 있으면
    재추정월(완전히 새 스케줄), 자본적지출은 origin을 리셋하지 않는다(정액법의
    '잔여 내용연수로 계속 상각' 정책과 동일 — 자본적지출은 기준가액만 늘리고
    이미 진행 중인 연차는 그대로 이어간다).

    5%/잔존가치 하한(floor) 로직이 따로 필요 없다 — life_years년에 걸친 몫의 합이
    정확히 1이 되도록 설계된 공식이라, 스케줄이 끝나면(k가 1~life_years 범위를
    벗어나면) 그 해의 몫 자체가 0이 되어 자연히 잔존가치에서 멈춘다.
    """
    if up_to_idx < start_idx:
        return float(basis), 0.0, 0

    origin_year = idx_to_year(origin_start_idx)
    denom = life_years * (life_years + 1) / 2
    start_year = idx_to_year(start_idx)
    up_to_year = idx_to_year(up_to_idx)

    book_value = float(basis)
    dep_in_target_year, months_in_target_year = 0.0, 0

    y = start_year
    while y <= up_to_year:
        year_start_idx = month_index(y, 1)
        year_end_idx = month_index(y, 12)
        eff_start = max(start_idx, year_start_idx)
        eff_end = min(active_end_idx, year_end_idx)
        if disposal_idx is not None:
            eff_end = min(eff_end, disposal_idx)
        if y == up_to_year:
            eff_end = min(eff_end, up_to_idx)

        months = eff_end - eff_start + 1 if eff_end >= eff_start else 0
        if susp_start_idx is not None and months > 0:
            s, e = max(eff_start, susp_start_idx), min(eff_end, susp_end_idx)
            if s <= e:
                months -= (e - s + 1)

        k = y - origin_year + 1
        dep = 0.0
        if months > 0 and 1 <= k <= life_years:
            frac = (life_years - k + 1) / denom
            dep = (basis - salvage) * frac * months / 12
            book_value -= dep

        if y == up_to_year:
            dep_in_target_year, months_in_target_year = dep, months
        y += 1

    return book_value, dep_in_target_year, months_in_target_year


def _book_value_at(seg: dict, up_to_idx: int, susp_start_idx, susp_end_idx) -> float:
    """구간(seg) 시작부터 up_to_idx(달력월, 포함)까지 상각했을 때의 장부가액."""
    if up_to_idx < seg["start_idx"]:
        return seg["basis"]
    if seg["method"] == "정액법":
        # 달력월수에서 상각중단과 겹치는 개월수를 직접 빼는 방식(구간 시작월 자체가
        # 상각중단 구간 "안"에서 시작하는 경계 케이스 — 예: 재추정이 상각중단 도중에
        # 발생한 경우 — 에서도 정확하다). 시작/종료 두 지점을 각각 "명목 경과월"로
        # 변환해서 빼는 방식은 구간 시작월이 중단구간 안이면 그 시작월 자체가 중단
        # 시작 직전월로 고정돼버려 elapsed가 부정확해지므로 쓰지 않는다.
        total_months = up_to_idx - seg["start_idx"] + 1
        susp_months = 0
        if susp_start_idx is not None:
            s, e = max(seg["start_idx"], susp_start_idx), min(up_to_idx, susp_end_idx)
            if s <= e:
                susp_months = e - s + 1
        elapsed = max(0, min(total_months - susp_months, seg["life_months"]))
        monthly = (seg["basis"] - seg["salvage"]) / seg["life_months"]
        return seg["basis"] - monthly * elapsed
    elif seg["method"] == "연수합계법":
        book_value, _, _ = _syd_year_loop(
            seg["start_idx"], seg["basis"], seg["salvage"], seg["life_years"], seg["origin_start_idx"],
            seg["end_idx"], None, susp_start_idx, susp_end_idx, up_to_idx)
        return book_value
    else:
        book_value, _, _ = _declining_balance_year_loop(
            seg["start_idx"], seg["basis"], seg["salvage"], seg["rate"], seg["end_idx"],
            None, susp_start_idx, susp_end_idx, up_to_idx, floor_threshold=_floor_threshold_for(seg))
        return book_value


def build_depreciation_schedule(acq, cost, salvage, life_years, method,
                                 reest_date, reest_life_years, reest_method,
                                 capex_date, capex_amount,
                                 susp_start_idx=None, susp_end_idx=None) -> list:
    """
    자본적지출(최대 1건)과 재추정(+방법변경, 최대 1건)을 시간순으로 반영해
    구간(segment) 리스트를 만든다. 각 구간(dict)은 start_idx, end_idx(명목 종료월),
    method, basis, salvage, life_months, rate를 담고, method는 구간마다 다를 수 있다.

    - 같은 날짜에 자본적지출과 재추정이 겹치면 자본적지출을 먼저 반영한다(가산된
      장부가액을 재추정의 기준가로 삼는 것이 실무적으로 자연스럽기 때문).
    - 자본적지출: end_idx/method 불변, basis = 그 시점 장부가액 + 자본적지출액
      ("잔여 내용연수로 계속 상각"). 정률법 5% 특례의 기준이 되는 취득가액(tax_cost)도
      같은 금액만큼 늘어난다(자본적지출은 세무상 취득가액에 가산).
    - 재추정: basis = 그 시점 장부가액(가산 없음), method = reest_method or 기존 method,
      end_idx = 재추정월 + reest_life_years*12 - 1 (새 내용연수로 완전히 재설정).
      tax_cost는 재추정으로 바뀌지 않는다(재추정은 취득가액이 아니라 내용연수/방법만
      다시 잡는 이벤트이므로).
    """
    start_idx0 = month_index(acq.year, acq.month)
    cur = dict(start_idx=start_idx0, end_idx=start_idx0 + life_years * 12 - 1,
               method=method, basis=float(cost), salvage=float(salvage),
               life_months=life_years * 12, tax_cost=float(cost),
               rate=_method_rate(method, life_years),
               life_years=life_years, origin_start_idx=start_idx0)

    breakpoints = []
    if capex_date is not None:
        breakpoints.append(dict(kind="capex", idx=month_index(capex_date.year, capex_date.month),
                                 amount=capex_amount))
    if reest_date is not None:
        breakpoints.append(dict(kind="reest", idx=month_index(reest_date.year, reest_date.month),
                                 new_life_years=reest_life_years, new_method=reest_method))
    breakpoints.sort(key=lambda b: (b["idx"], 0 if b["kind"] == "capex" else 1))

    segments = []
    for bp in breakpoints:
        prev_end_idx = cur["end_idx"]  # capex는 이 종료월을 그대로 이어받는다
        book_value = _book_value_at(cur, bp["idx"] - 1, susp_start_idx, susp_end_idx)
        cur = dict(cur, end_idx=min(cur["end_idx"], bp["idx"] - 1))
        segments.append(cur)

        if bp["kind"] == "capex":
            # origin_start_idx는 그대로 이어받는다 — 자본적지출은 기준가액만 늘릴 뿐,
            # 연수합계법이 이미 진행 중인 연차(k) 스케줄을 리셋하지 않는다.
            cur = dict(start_idx=bp["idx"], end_idx=prev_end_idx, method=segments[-1]["method"],
                       basis=book_value + bp["amount"], salvage=segments[-1]["salvage"],
                       life_months=prev_end_idx - bp["idx"] + 1, rate=segments[-1]["rate"],
                       tax_cost=segments[-1]["tax_cost"] + bp["amount"],
                       life_years=segments[-1]["life_years"], origin_start_idx=segments[-1]["origin_start_idx"])
        else:
            new_method = bp["new_method"] or segments[-1]["method"]
            new_life_months = bp["new_life_years"] * 12
            cur = dict(start_idx=bp["idx"], end_idx=bp["idx"] + new_life_months - 1,
                       method=new_method, basis=book_value, salvage=segments[-1]["salvage"],
                       life_months=new_life_months, tax_cost=segments[-1]["tax_cost"],
                       rate=_method_rate(new_method, bp["new_life_years"]),
                       life_years=bp["new_life_years"], origin_start_idx=bp["idx"])

    segments.append(cur)
    return segments


def apply_suspension_extension(segments: list, susp_start_idx, susp_end_idx) -> list:
    """
    상각중단이 마지막 구간 안에서 발생한 경우, 그 구간의 종료월(end_idx)을 중단
    기간만큼 늘린다. 상각중단이 마지막이 아닌 중간 구간에서 발생하면(그 구간은 어차피
    다음 이벤트가 먼저 도래해 명목 내용연수가 끝나기 전에 끝나므로) 연장하지 않는다
    (v1 스코프 제한 — README에 명시).
    """
    if susp_start_idx is None:
        return segments
    last = segments[-1]
    if last["start_idx"] <= susp_start_idx <= last["end_idx"]:
        last["end_idx"] += (susp_end_idx - susp_start_idx + 1)
    return segments


def segments_current_period_dep(segments: list, disposal_idx, fy_year: int,
                                 susp_start_idx=None, susp_end_idx=None) -> tuple:
    """method가 섞인 구간 리스트를 받아 당기(fy_year)와 겹치는 상각비를 합산한다."""
    fy_start_idx = month_index(fy_year, 1)
    fy_end_idx = month_index(fy_year, 12)
    if disposal_idx is not None:
        fy_end_idx = min(fy_end_idx, disposal_idx)

    total_dep, total_months = 0.0, 0
    for seg in segments:
        overlap_start = max(seg["start_idx"], fy_start_idx)
        overlap_end = min(seg["end_idx"], fy_end_idx)
        if overlap_start > overlap_end:
            continue

        if seg["method"] == "정액법":
            susp_months = 0
            if susp_start_idx is not None:
                s, e = max(overlap_start, susp_start_idx), min(overlap_end, susp_end_idx)
                if s <= e:
                    susp_months = e - s + 1
            months = (overlap_end - overlap_start + 1) - susp_months
            if months <= 0:
                continue
            monthly_dep = (seg["basis"] - seg["salvage"]) / seg["life_months"]
            total_dep += monthly_dep * months
            total_months += months
        elif seg["method"] == "연수합계법":
            _, dep, months = _syd_year_loop(
                seg["start_idx"], seg["basis"], seg["salvage"], seg["life_years"], seg["origin_start_idx"],
                seg["end_idx"], None, susp_start_idx, susp_end_idx, overlap_end)
            total_dep += dep
            total_months += months
        else:
            _, dep, months = _declining_balance_year_loop(
                seg["start_idx"], seg["basis"], seg["salvage"], seg["rate"], seg["end_idx"],
                None, susp_start_idx, susp_end_idx, overlap_end, floor_threshold=_floor_threshold_for(seg))
            total_dep += dep
            total_months += months
    return total_dep, total_months


# ---------------------------------------------------------------------------
# 생산량비례법 재계산
# ---------------------------------------------------------------------------
def units_of_production_current_period_dep(cost, salvage, total_units, period_units,
                                             disposal, fy_year) -> tuple:
    """
    생산량비례법 당기상각비 = (취득원가-잔존가치) x (당기실제생산량/총예정생산량).
    - 처분일이 당기 이전 회계연도면 이미 상각이 종료된 것으로 보고 0원을 반환한다.
    - 취득 이후 누적생산량 이력은 추적하지 않으므로(당기 값만 입력받는 v1 단순화),
      상각누계액이 상각대상금액(취득원가-잔존가치)을 넘지 않도록만 방어적으로 캡을 건다.
    """
    if disposal is not None and disposal.year < fy_year:
        return 0.0, 0
    dep = (cost - salvage) * (period_units / total_units)
    dep = max(0.0, min(dep, cost - salvage))
    months = 12 if period_units and period_units > 0 else 0
    return dep, months


# ---------------------------------------------------------------------------
# 자산 1건 재계산
# ---------------------------------------------------------------------------
def recalc_asset(acq, cost, salvage, life, method, disposal, reest_date, reest_life, ref_date, fy_year,
                  total_units=None, period_units=None,
                  reest_method=None, capex_date=None, capex_amount=None,
                  susp_start=None, susp_end=None):
    # 자본적지출/상각중단/방법이 실제로 바뀌는 재추정이 하나라도 있으면 새 통합
    # 이벤트 엔진으로 넘어간다. 그 외(기존 52개 테스트가 타는 경로)는 기존
    # 정액법/정률법/생산량비례법 분기를 문자 그대로 그대로 사용한다(회귀 방지).
    has_extended_events = (
        capex_date is not None
        or susp_start is not None
        or (reest_method is not None and reest_method != method)
    )

    active_end_idx = None
    # 이중체감법/연수합계법은 신규 방법이라 기존 정률법처럼 "이벤트 없을 때만 쓰는
    # 단순경로"를 따로 만들지 않고, 이벤트 유무와 무관하게 항상 통합 엔진을 탄다.
    if (method in ("정액법", "정률법") and has_extended_events) or method in ("이중체감법", "연수합계법"):
        susp_s = month_index(susp_start.year, susp_start.month) if susp_start is not None else None
        susp_e = month_index(susp_end.year, susp_end.month) if susp_end is not None else None
        segments = build_depreciation_schedule(
            acq, cost, salvage, life, method, reest_date, reest_life, reest_method,
            capex_date, capex_amount, susp_s, susp_e)
        segments = apply_suspension_extension(segments, susp_s, susp_e)
        disposal_idx = month_index(disposal.year, disposal.month) if disposal is not None else None
        dep_raw, months = segments_current_period_dep(segments, disposal_idx, fy_year, susp_s, susp_e)
        active_end_idx = segments[-1]["end_idx"]
    elif method == "정액법":
        segments = build_straight_line_segments(acq, cost, salvage, life, reest_date, reest_life)
        disposal_idx = month_index(disposal.year, disposal.month) if disposal is not None else None
        dep_raw, months = straight_line_current_period_dep(segments, disposal_idx, fy_year)
        active_end_idx = segments[-1]["end_idx"]
    elif method == "정률법":
        dep_raw, months, _opening_bv = declining_balance_current_period_dep(
            acq, cost, salvage, life, reest_date, reest_life, disposal, fy_year)
        if reest_date is not None:
            active_end_idx = month_index(reest_date.year, 1) + reest_life * 12 - 1
        else:
            active_end_idx = month_index(acq.year, acq.month) + life * 12 - 1
    elif method == "생산량비례법":
        dep_raw, months = units_of_production_current_period_dep(
            cost, salvage, total_units, period_units, disposal, fy_year)
    else:
        raise ValueError(f"알 수 없는 상각방법: {method}")

    if active_end_idx is not None:
        ref_idx = month_index(ref_date.year, ref_date.month)
        life_ended = ref_idx > active_end_idx
    else:
        # 생산량비례법은 누적생산량 이력을 추적하지 않아(v1 단순화) 시간 기준
        # 생애종료 판정을 할 수 없으므로 항상 "종료 아님"으로 본다.
        life_ended = False

    notes = []
    if disposal is not None:
        if disposal.year < fy_year:
            notes.append(f"전기이전처분({disposal.isoformat()})")
        else:
            notes.append(f"당기중처분({disposal.isoformat()})")
    if reest_date is not None:
        method_note = f"→{reest_method}" if (reest_method is not None and reest_method != method) else ""
        notes.append(f"내용연수재추정({reest_date.isoformat()}→{reest_life}년{method_note})")
    if capex_date is not None:
        notes.append(f"자본적지출({capex_date.isoformat()}→{capex_amount:,.0f}원)")
    if susp_start is not None:
        notes.append(f"상각중단({susp_start.isoformat()}~{susp_end.isoformat()})")
    if life_ended and disposal is None:
        notes.append("내용연수종료")
    note = "; ".join(notes) if notes else "-"

    return round_won(dep_raw), months, life_ended, note


def recalc_accumulated_dep(acq, cost, salvage, life, method, disposal, reest_date, reest_life,
                            fy_year, total_units=None, period_units=None,
                            reest_method=None, capex_date=None, capex_amount=None,
                            susp_start=None, susp_end=None):
    """
    취득연도부터 전기(fy_year-1년)까지 매 연도의 재계산 상각비를 합산해 "전기말
    재계산 감가상각누계액"을 구한다. 회사가 제시한 전기말 누계액과 비교해서, 당기에
    갑자기 발생한 차이가 아니라 전기 이전부터 존재하던 차이인지 확인하는 참고용
    지표다(회사반영 당기 감가상각비와의 일치 여부 판정에는 영향을 주지 않는다).

    생산량비례법은 당기 실제생산량만 입력받고 과거 연도별 생산량 이력을 추적하지
    않으므로(v1 한계), 과거 연도별 상각액을 재구성할 수 없어 None을 반환한다.
    """
    if method == "생산량비례법":
        return None
    total = 0
    for y in range(acq.year, fy_year):
        dep, _, _, _ = recalc_asset(
            acq, cost, salvage, life, method, disposal, reest_date, reest_life,
            fy_ref_date(y), y, total_units=total_units, period_units=period_units,
            reest_method=reest_method, capex_date=capex_date, capex_amount=capex_amount,
            susp_start=susp_start, susp_end=susp_end)
        total += dep
    return total


def _segment_book_value_before(segments, cutoff_idx, susp_s, susp_e, default_method):
    """
    segments(정액법 또는 정률법 세그먼트 리스트) 안에서 cutoff_idx(달력월, 포함)
    시점까지 상각했을 때의 실제 장부가액을 구한다. cutoff_idx가 마지막 세그먼트
    종료월보다 뒤라면(이미 내용연수가 끝난 상태) 마지막 세그먼트의 종료 시점 장부가액
    (=잔존가치)을 반환한다. build_straight_line_segments가 만든 세그먼트는 "method"
    키가 없으므로 default_method로 채워 _book_value_at에 넘긴다.
    """
    if cutoff_idx < segments[0]["start_idx"]:
        return segments[0]["basis"]
    seg = next((s for s in segments if s["start_idx"] <= cutoff_idx <= s["end_idx"]), None)
    if seg is None:
        seg = segments[-1]
        cutoff_idx = seg["end_idx"]
    if "method" not in seg:
        seg = dict(seg, method=default_method)
    return _book_value_at(seg, cutoff_idx, susp_s, susp_e)


def get_period_formula_meta(acq, cost, salvage, life, method, disposal, reest_date, reest_life,
                             fy_year, reest_method=None, capex_date=None, capex_amount=None,
                             susp_start=None, susp_end=None, prior_period_units=None) -> dict:
    """
    재계산결과 시트에 엑셀 수식을 주입하기 위한 메타데이터를 계산한다(recalc_asset이
    반환하는 상각비 숫자 자체와는 별개로, "그 숫자를 단일 곱셈 수식으로 표현할 수 있는가"와
    "그 수식에 쓸 상각기준가액/장부가액"만 판단한다). recalc_asset의 계산 결과 자체는
    건드리지 않는다.

    - basis: 당기 상각비 계산의 기준가액. 정액법은 당기에 적용되는 세그먼트의 고정
      상각대상 기준액(재추정/자본적지출 전까지는 취득원가와 같음), 정률법은 당기
      1/1 시점의 실제 장부가액(매년 감소) — 이 경우 prior_fy_end_bv와 같은 값이다.
    - life_months: 정액법에서 상각률(=12/life_months) 계산에 쓰는, 당기에 적용되는
      세그먼트의 실제 상각기간(개월). 자본적지출만 있고 재추정이 없으면 "원 내용연수
      전체"가 아니라 "자본적지출 시점부터 원래 내용연수 종료월까지 남은 개월수"가
      되므로, 원 내용연수(년)를 그대로 쓰면 안 된다(자본적지출은 잔여 내용연수로
      재상각하는 정책이기 때문). 재추정이 있으면 재추정후내용연수(년)*12와 같다.
    - is_simple_current: 당기 회계연도 안에서 상각 기준(세그먼트)이 단 하나뿐이어서
      "(상각기준가액-잔존가치)*상각률*월수/12" 형태의 단일 셀 수식으로 표현 가능한지.
      당기 중에 재추정 또는 자본적지출이 발생해 같은 해 안에 기준이 바뀌는 경우만 False.
    - prior_fy_end_bv: 전기말(=당기 1/1 직전) 시점의 실제 장부가액(처분이 그 전에
      있었으면 처분 시점에 멈춘 값). "취득원가+자본적지출누계-prior_fy_end_bv"가 곧
      전기말 감가상각누계액과 같다는 항등식을 이용해 전기말 누계액을 수식화하는 데
      쓴다. 정액법/정률법 모두 계산 가능하고(당기 안에서 이벤트가 있어도 전기말은
      그보다 전 시점이라 영향받지 않는다), 생산량비례법은 장부가액 개념 자체가
      없어(생산량 비율로만 계산) None을 반환한다.
    - accum_formula_eligible: 전기말 감가상각누계액을 수식으로 표현할 수 있는지.
      정액법/정률법은 항상 True(위 항등식이 이벤트 유무와 무관하게 성립), 생산량비례법은
      prior_period_units(전기말누적생산량)이 주어졌을 때만 True.
    """
    if method == "생산량비례법":
        return dict(basis=cost, life_months=None, is_simple_current=True, prior_fy_end_bv=None,
                    accum_formula_eligible=prior_period_units is not None, syd_k=None, syd_life_years=None)

    susp_s = month_index(susp_start.year, susp_start.month) if susp_start is not None else None
    susp_e = month_index(susp_end.year, susp_end.month) if susp_end is not None else None
    has_extended = (
        capex_date is not None
        or susp_start is not None
        or (reest_method is not None and reest_method != method)
    )

    fy_start_idx = month_index(fy_year, 1)
    disposal_idx = month_index(disposal.year, disposal.month) if disposal is not None else None
    cutoff_idx = fy_start_idx - 1
    if disposal_idx is not None:
        cutoff_idx = min(cutoff_idx, disposal_idx)

    if method == "정률법" and not has_extended:
        # 구 경로(정률법, 자본적지출/상각중단 없음): 재추정을 항상 회계연도 1/1
        # 기준으로 적용하므로 한 회계연도 안에서 기준이 둘로 쪼개지는 경우가 없다.
        # life_months는 정률법 상각률 수식(VLOOKUP)에서 쓰지 않으므로 실사용되지 않는다.
        # declining_balance_current_period_dep의 연단위 루프는 처분월 이후 개월수가
        # 0이 되어 book_value가 자동으로 멈추므로 opening_book_value가 이미 처분
        # 시점 이후 값으로 고정돼 있어(disposal-safe) 별도 처리가 필요 없다.
        _, _, opening_bv = declining_balance_current_period_dep(
            acq, cost, salvage, life, reest_date, reest_life, disposal, fy_year)
        eff_life = reest_life if reest_date is not None else life
        return dict(basis=opening_bv, life_months=eff_life * 12, is_simple_current=True,
                    prior_fy_end_bv=opening_bv, accum_formula_eligible=True, syd_k=None, syd_life_years=None)

    # 이중체감법/연수합계법은 정률법과 달리 "단순경로"가 없어 이벤트가 하나도 없어도
    # 항상 통합 엔진(build_depreciation_schedule)을 탄다. 정액법만 이벤트가 없을 때
    # 기존 경량 경로(build_straight_line_segments)를 그대로 쓴다(회귀 방지).
    if method == "정액법" and not has_extended:
        segments = build_straight_line_segments(acq, cost, salvage, life, reest_date, reest_life)
    else:
        segments = build_depreciation_schedule(
            acq, cost, salvage, life, method, reest_date, reest_life, reest_method,
            capex_date, capex_amount, susp_s, susp_e)
        segments = apply_suspension_extension(segments, susp_s, susp_e)

    fy_end_idx = month_index(fy_year, 12)
    if disposal_idx is not None:
        fy_end_idx = min(fy_end_idx, disposal_idx)

    overlapping = [s for s in segments if s["start_idx"] <= fy_end_idx and s["end_idx"] >= fy_start_idx]
    is_simple_current = len(overlapping) <= 1

    opening_seg = next((s for s in segments if s["start_idx"] <= fy_start_idx <= s["end_idx"]), None)
    if opening_seg is None:
        opening_seg = overlapping[0] if overlapping else segments[0]

    opening_method = opening_seg.get("method", method)
    if opening_method in ("정액법", "연수합계법"):
        # 둘 다 "기초 장부가액"이 아니라 고정된 기준가액에 상각률(또는 몫)을 곱하는
        # 구조라 basis가 매년 줄어드는 개념이 아니다.
        basis = opening_seg["basis"]
    else:
        basis = _book_value_at(opening_seg, fy_start_idx - 1, susp_s, susp_e)

    prior_fy_end_bv = _segment_book_value_before(segments, cutoff_idx, susp_s, susp_e, method)

    syd_k, syd_life_years = None, None
    if opening_method == "연수합계법":
        syd_life_years = opening_seg["life_years"]
        syd_k = fy_year - idx_to_year(opening_seg["origin_start_idx"]) + 1

    return dict(basis=basis, life_months=opening_seg["life_months"], is_simple_current=is_simple_current,
                prior_fy_end_bv=prior_fy_end_bv, accum_formula_eligible=True,
                syd_k=syd_k, syd_life_years=syd_life_years)


# ---------------------------------------------------------------------------
# 컬럼 인식
# ---------------------------------------------------------------------------
# _suggest_columns_ai 프롬프트에서 각 필드가 무슨 뜻인지 Claude에게 설명해주는 문구.
# "취득원가"는 장부가액(순장부가액)과 다른 개념이라는 점을 명시해, AI가 헷갈리기 쉬운
# 컬럼("장부가액" 등)을 함부로 취득원가에 매칭하지 않도록 유도한다.
_FIELD_DESCRIPTIONS = {
    "자산명": "자산의 이름이나 고유 식별자(자산코드 등)",
    "자산분류": "유형자산/무형자산 같은 자산 대분류",
    "취득일": "자산을 취득한 날짜",
    "취득원가": "최초 취득 시 지급한 원가(취득가액). 장부가액/순장부가액과는 다른 개념이므로 혼동하지 말 것",
    "잔존가치": "내용연수 종료 시점의 잔존가치",
    "내용연수": "감가상각 내용연수(년)",
    "상각방법": "정액법/정률법/이중체감법/연수합계법/생산량비례법 등 감가상각방법",
    "회사반영상각비": "회사가 장부에 반영한 당기 감가상각비",
    "처분일": "당기 중 자산을 처분(매각)한 날짜",
    "재추정일": "내용연수를 재추정(변경)한 날짜",
    "재추정내용연수": "재추정 이후 적용하는 새 내용연수",
    "재추정후상각방법": "재추정과 함께 상각방법이 바뀌었다면 그 새 상각방법",
    "총예정생산량": "생산량비례법에서 쓰는 총 예정 생산량",
    "당기실제생산량": "생산량비례법에서 쓰는 당기 실제 생산량",
    "전기말누적생산량": "생산량비례법에서 쓰는 전기말까지의 누적 생산량",
    "자본적지출일": "자본적지출(추가 취득원가 반영)이 발생한 날짜",
    "자본적지출액": "자본적지출 금액",
    "상각중단시작일": "감가상각을 중단한 기간의 시작일",
    "상각중단종료일": "감가상각을 중단한 기간의 종료일",
    "회사반영누계상각액": "회사가 반영한 전기말 감가상각누계액",
}


def _find_column(key: str, df_columns, mapped_name):
    """
    COLUMN_MAP에 지정된 이름과 정확히 일치하는 컬럼을 먼저 찾고, 없으면
    COLUMN_SYNONYMS 목록에서 찾는다. (실제 컬럼명, 매칭소스) 튜플을 반환하며,
    소스는 "정확일치"/"동의어" 중 하나. 둘 다 못 찾으면 (None, None).
    """
    if mapped_name and mapped_name in df_columns:
        return mapped_name, "정확일치"
    for syn in COLUMN_SYNONYMS.get(key, []):
        if syn in df_columns:
            return syn, "동의어"
    return None, None


def _suggest_columns_fuzzy(missing_keys: list, df_columns) -> dict:
    """
    difflib(문자열 유사도)로 파일의 실제 컬럼명 중 각 필드와 비슷한 이름을 추천한다.
    반환값은 참고용 메시지에만 쓰이고 절대 COLUMN_MAP/resolved에 자동 반영되지
    않는다(잘못 추측된 컬럼이 재무 계산에 조용히 흘러들어가는 것을 막기 위함).
    """
    cols = [str(c) for c in df_columns]
    suggestions = {}
    for key in missing_keys:
        candidates = difflib.get_close_matches(key, cols, n=2, cutoff=0.4)
        if candidates:
            suggestions[key] = candidates
    return suggestions


def _suggest_columns_ai(missing_keys: list, df_columns) -> dict:
    """
    ANTHROPIC_API_KEY가 있으면 Claude에게 파일의 실제 컬럼 헤더 목록을 보여주고
    누락된 각 필드에 가장 그럴듯한 컬럼을 파일당 1회만 추천받는다(get_ai_estimated_cause와
    동일한 "키 없거나 호출 실패하면 조용히 빈 dict 반환" 폴백 패턴 — 호출부가
    _suggest_columns_fuzzy로 이어서 대체한다). 이 함수의 결과도 어디까지나 참고용
    추천 메시지에만 쓰이고, 절대 자동으로 COLUMN_MAP에 반영되지 않는다 — 컬럼 매칭이
    틀리면 코멘트 한 줄이 아니라 전체 재계산 결과가 조용히 틀어지기 때문이다.
    """
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key or anthropic is None:
        return {}

    cols = [str(c) for c in df_columns]
    fields_text = "\n".join(f"- {k}: {_FIELD_DESCRIPTIONS.get(k, k)}" for k in missing_keys)
    columns_text = ", ".join(f'"{c}"' for c in cols)
    prompt = (
        "고정자산대장 엑셀 파일에서 아래 항목에 해당하는 컬럼을 실제 파일의 컬럼명 목록에서 "
        "찾아주세요. 개념이 다른 컬럼(예: 취득원가와 장부가액/순장부가액)을 섞어서 추천하면 안 됩니다.\n\n"
        f"찾아야 할 항목:\n{fields_text}\n\n"
        f"실제 파일의 컬럼명 목록: {columns_text}\n\n"
        "각 항목마다 정확히 이 형식으로 한 줄씩만 답하세요(그 외 설명 문장 없이):\n"
        "항목명: 추천컬럼명 | 신뢰도(상/중/하) | 이유(20자 이내)\n"
        "해당하는 컬럼이 전혀 없어 보이면 추천컬럼명 자리에 없음이라고 쓰세요."
    )
    try:
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model=ANTHROPIC_MODEL, max_tokens=500,
            messages=[{"role": "user", "content": prompt}],
        )
        text = next((b.text for b in response.content if b.type == "text"), "")
        suggestions = {}
        for line in text.splitlines():
            if ":" not in line or "|" not in line:
                continue
            key_part, rest = line.split(":", 1)
            key = key_part.strip()
            if key not in missing_keys:
                continue
            parts = [p.strip() for p in rest.split("|")]
            col_guess = parts[0] if parts else ""
            if col_guess and col_guess != "없음" and col_guess in cols:
                confidence = parts[1] if len(parts) > 1 else "-"
                reason = parts[2] if len(parts) > 2 else "-"
                suggestions[key] = [f"{col_guess} (AI추천, 신뢰도:{confidence}, {reason})"]
        return suggestions
    except Exception as e:
        print(f"  [경고] AI 컬럼 매칭 추천 호출 실패, 퍼지매칭으로 대체합니다: {e}")
        return {}


def resolve_columns(df):
    print("=== 컬럼 인식 결과 ===")
    resolved = {}
    missing_required = []
    for key in REQUIRED_KEYS:
        actual, source = _find_column(key, df.columns, COLUMN_MAP.get(key))
        if actual is None:
            missing_required.append(key)
            print(f"  {key} = {COLUMN_MAP.get(key)}  [!! 파일에서 찾을 수 없음]")
        else:
            tag = "" if source == "정확일치" else f"  ({source} 매칭: '{actual}')"
            print(f"  {key} = {actual}{tag}")
        resolved[key] = actual

    missing_optional = []
    for key in OPTIONAL_KEYS:
        actual, source = _find_column(key, df.columns, COLUMN_MAP.get(key))
        if actual is None:
            missing_optional.append(key)
            print(f"  {key} = {COLUMN_MAP.get(key)} (선택, 파일에 없어 미적용)")
        else:
            tag = "(선택)" if source == "정확일치" else f"(선택, {source} 매칭: '{actual}')"
            print(f"  {key} = {actual} {tag}")
        resolved[key] = actual

    if missing_required:
        suggestions = _suggest_columns_ai(missing_required, df.columns)
        if not suggestions:
            suggestions = _suggest_columns_fuzzy(missing_required, df.columns)
        detail_lines = [
            f"  - {key}: 혹시 이 컬럼 아닌가요? {', '.join(suggestions[key])}"
            for key in missing_required if suggestions.get(key)
        ]
        detail = ("\n" + "\n".join(detail_lines)) if detail_lines else ""
        raise KeyError(
            "다음 필수 컬럼을 파일에서 찾을 수 없습니다: "
            + ", ".join(missing_required)
            + "  → 스크립트 상단 COLUMN_MAP 에서 실제 컬럼명으로 수정하세요."
            + detail
        )

    if missing_optional:
        opt_suggestions = _suggest_columns_fuzzy(missing_optional, df.columns)
        for key in missing_optional:
            cands = opt_suggestions.get(key)
            if cands:
                print(f"  [참고] {key}를 찾지 못했는데 혹시 이 컬럼 아닌가요? {', '.join(cands)}")

    print()
    return resolved


def build_category_summary(result_df: pd.DataFrame, error_df: pd.DataFrame) -> pd.DataFrame:
    """자산분류별 전체/정상/불일치/유의한차이/오류 건수와 불일치율을 집계한다."""
    if len(result_df) > 0:
        normal = result_df.groupby("자산분류").agg(
            정상계산건수=("자산명", "count"),
            불일치건수=("차이(재계산-회사반영)", lambda s: (s != 0).sum()),
            유의한차이건수=("중요성구분", lambda s: (s == "유의한 차이").sum()),
        )
    else:
        normal = pd.DataFrame(columns=["정상계산건수", "불일치건수", "유의한차이건수"])
        normal.index.name = "자산분류"

    if len(error_df) > 0:
        error_counts = error_df.groupby("자산분류").agg(오류건수=("자산명", "count"))
    else:
        error_counts = pd.DataFrame(columns=["오류건수"])
        error_counts.index.name = "자산분류"

    summary = normal.join(error_counts, how="outer").fillna(0)
    for col in ("정상계산건수", "불일치건수", "유의한차이건수", "오류건수"):
        summary[col] = summary[col].astype(int)
    summary["전체건수"] = summary["정상계산건수"] + summary["오류건수"]
    summary["불일치율(%)"] = summary.apply(
        lambda row: round(row["불일치건수"] / row["정상계산건수"] * 100, 1) if row["정상계산건수"] > 0 else 0.0,
        axis=1,
    )
    summary = summary.reset_index().rename(columns={"자산분류": "자산분류"})
    return summary[["자산분류", "전체건수", "정상계산건수", "불일치건수", "불일치율(%)", "유의한차이건수", "오류건수"]]


def parse_asset_row(r, cols) -> dict:
    """엑셀 한 행에서 계산에 필요한 원시값을 파싱해 dict로 반환한다."""
    total_units_raw = r[cols["총예정생산량"]] if cols["총예정생산량"] else None
    period_units_raw = r[cols["당기실제생산량"]] if cols["당기실제생산량"] else None
    prior_period_units_raw = r[cols["전기말누적생산량"]] if cols["전기말누적생산량"] else None
    reest_life_raw = r[cols["재추정내용연수"]] if cols["재추정내용연수"] else None
    reest_method_raw = r[cols["재추정후상각방법"]] if cols["재추정후상각방법"] else None
    capex_amount_raw = r[cols["자본적지출액"]] if cols["자본적지출액"] else None
    accum_reported_raw = r[cols["회사반영누계상각액"]] if cols["회사반영누계상각액"] else None
    return dict(
        자산명=r[cols["자산명"]],
        자산분류=r[cols["자산분류"]],
        acq=to_date_or_none(r[cols["취득일"]]),
        cost=float(r[cols["취득원가"]]),
        salvage=float(r[cols["잔존가치"]]),
        life=int(r[cols["내용연수"]]),
        method=normalize_method(r[cols["상각방법"]]),
        reported=round_won(r[cols["회사반영상각비"]]),
        disposal=to_date_or_none(r[cols["처분일"]]) if cols["처분일"] else None,
        reest_date=to_date_or_none(r[cols["재추정일"]]) if cols["재추정일"] else None,
        reest_life=int(reest_life_raw) if (reest_life_raw is not None and not pd.isna(reest_life_raw)) else None,
        reest_method=(normalize_method(reest_method_raw)
                      if (reest_method_raw is not None and not pd.isna(reest_method_raw)) else None),
        total_units=float(total_units_raw) if (total_units_raw is not None and not pd.isna(total_units_raw)) else None,
        period_units=float(period_units_raw) if (period_units_raw is not None and not pd.isna(period_units_raw)) else None,
        prior_period_units=(float(prior_period_units_raw)
                             if (prior_period_units_raw is not None and not pd.isna(prior_period_units_raw)) else None),
        capex_date=to_date_or_none(r[cols["자본적지출일"]]) if cols["자본적지출일"] else None,
        capex_amount=(float(capex_amount_raw)
                      if (capex_amount_raw is not None and not pd.isna(capex_amount_raw)) else None),
        susp_start=to_date_or_none(r[cols["상각중단시작일"]]) if cols["상각중단시작일"] else None,
        susp_end=to_date_or_none(r[cols["상각중단종료일"]]) if cols["상각중단종료일"] else None,
        accum_reported=(float(accum_reported_raw)
                        if (accum_reported_raw is not None and not pd.isna(accum_reported_raw)) else None),
    )


def build_multi_year_trend_df(df: pd.DataFrame, cols: dict, years: list) -> pd.DataFrame:
    """
    자산별로 years에 지정된 각 회계연도의 재계산 상각비를 구해 long format으로 쌓는다.
    자산명이 중복될 수 있으므로 그룹핑용 키는 자산명이 아니라 원본 행 순번(자산ID)을 쓴다.
    데이터 오류(계산 불가능한 값)로 걸러지는 자산은 단일연도 처리와 동일하게 제외한다.
    """
    rows = []
    for asset_id, (_, r) in enumerate(df.iterrows()):
        p = parse_asset_row(r, cols)
        errors = validate_asset_inputs(
            p["cost"], p["salvage"], p["life"], method=p["method"], total_units=p["total_units"],
            reest_method=p["reest_method"], capex_date=p["capex_date"], capex_amount=p["capex_amount"],
            susp_start=p["susp_start"], susp_end=p["susp_end"])
        if errors:
            continue
        for y in years:
            recalced, months, life_ended, note = recalc_asset(
                p["acq"], p["cost"], p["salvage"], p["life"], p["method"], p["disposal"],
                p["reest_date"], p["reest_life"], fy_ref_date(y), y,
                total_units=p["total_units"], period_units=p["period_units"],
                reest_method=p["reest_method"], capex_date=p["capex_date"], capex_amount=p["capex_amount"],
                susp_start=p["susp_start"], susp_end=p["susp_end"])
            rows.append({
                "자산ID": asset_id, "자산명": p["자산명"], "자산분류": p["자산분류"],
                "상각방법": p["method"], "회계연도": y,
                "재계산_당기감가상각비": recalced, "당기해당월수": months, "비고": note,
            })
    return pd.DataFrame(rows)


def detect_yoy_anomalies(trend_df: pd.DataFrame, threshold_pct: float = YOY_ANOMALY_THRESHOLD_PCT) -> pd.DataFrame:
    """
    자산ID별로 회계연도 순 정렬 후 전년대비 증감률을 계산해 이상탐지 컬럼을 추가한다.
    전년도 상각비가 0원이면 %증감이 무의미하므로(0으로 나누기) '신규발생'/'-'로 표기한다.
    """
    df = trend_df.sort_values(["자산ID", "회계연도"]).reset_index(drop=True).copy()
    prev = df.groupby("자산ID")["재계산_당기감가상각비"].shift(1)
    pct = (df["재계산_당기감가상각비"] - prev) / prev.replace(0, float("nan")) * 100
    df["전년대비증감률(%)"] = pct.round(1)

    def _flag(row_prev, row_cur, row_pct):
        if pd.isna(row_prev):
            return "-"
        if row_prev == 0:
            return "신규발생" if row_cur > 0 else "-"
        return "경고" if abs(row_pct) >= threshold_pct else "-"

    df["이상탐지"] = [
        _flag(p, c, pc) for p, c, pc in zip(prev, df["재계산_당기감가상각비"], pct)
    ]
    return df


def main():
    df = pd.read_excel(IN_PATH, sheet_name=0)
    df.columns = [str(c).strip() for c in df.columns]

    cols = resolve_columns(df)

    error_cols = ["자산명", "자산분류", "취득일", "취득원가", "잔존가치", "내용연수(년)",
                  "상각방법", "회사반영_당기감가상각비", "총예정생산량", "당기실제생산량",
                  "전기말누적생산량", "자본적지출일", "자본적지출액", "상각중단시작일", "상각중단종료일",
                  "회사반영_전기말감가상각누계액", "오류사유"]
    rows = []
    formula_flags = []
    error_rows = []
    for _, r in df.iterrows():
        p = parse_asset_row(r, cols)
        acq, cost, salvage, life, method, reported = (
            p["acq"], p["cost"], p["salvage"], p["life"], p["method"], p["reported"])
        disposal, reest_date, reest_life, reest_method = (
            p["disposal"], p["reest_date"], p["reest_life"], p["reest_method"])
        total_units, period_units = p["total_units"], p["period_units"]
        prior_period_units = p["prior_period_units"]
        capex_date, capex_amount = p["capex_date"], p["capex_amount"]
        susp_start, susp_end = p["susp_start"], p["susp_end"]
        accum_reported = p["accum_reported"]

        validation_errors = validate_asset_inputs(
            cost, salvage, life, method=method, total_units=total_units,
            reest_method=reest_method, capex_date=capex_date, capex_amount=capex_amount,
            susp_start=susp_start, susp_end=susp_end, accum_reported=accum_reported,
            prior_period_units=prior_period_units)
        if validation_errors:
            error_rows.append({
                "자산명": p["자산명"],
                "자산분류": p["자산분류"],
                "취득일": acq,
                "취득원가": cost,
                "잔존가치": salvage,
                "내용연수(년)": life,
                "상각방법": method,
                "회사반영_당기감가상각비": reported,
                "총예정생산량": total_units,
                "당기실제생산량": period_units,
                "전기말누적생산량": prior_period_units,
                "자본적지출일": capex_date,
                "자본적지출액": capex_amount,
                "상각중단시작일": susp_start,
                "상각중단종료일": susp_end,
                "회사반영_전기말감가상각누계액": accum_reported,
                "오류사유": "; ".join(validation_errors),
            })
            continue

        elapsed = elapsed_months_to_ref(acq, REF_DATE)

        recalced, cur_months, life_ended, note = recalc_asset(
            acq, cost, salvage, life, method, disposal, reest_date, reest_life, REF_DATE, FY_YEAR,
            total_units=total_units, period_units=period_units,
            reest_method=reest_method, capex_date=capex_date, capex_amount=capex_amount,
            susp_start=susp_start, susp_end=susp_end)

        # 전기말 감가상각누계액 비교(참고용): 회사반영 당기 감가상각비의 일치/중요성
        # 판정에는 영향을 주지 않는다 — 당기 차이(또는 누계액 차이)가 당기에 갑자기
        # 발생한 게 아니라 전기 이전부터 있었는지 확인하는 별도 절차이기 때문이다.
        accum_recalc = recalc_accumulated_dep(
            acq, cost, salvage, life, method, disposal, reest_date, reest_life, FY_YEAR,
            total_units=total_units, period_units=period_units,
            reest_method=reest_method, capex_date=capex_date, capex_amount=capex_amount,
            susp_start=susp_start, susp_end=susp_end)
        accum_diff = (accum_recalc - accum_reported) if (accum_recalc is not None and accum_reported is not None) else None

        # 엑셀 수식 주입용 메타데이터: 상각률/감가상각대상금액/재계산 수치를 실제 셀
        # 수식으로 연결하기 위해 필요한 상각기준가액과 "단일 수식으로 표현 가능한가"를
        # 계산한다(재계산 숫자 자체는 위 recalc_asset/recalc_accumulated_dep 결과를
        # 그대로 쓰고, 여기서는 표시 방식만 결정한다).
        meta = get_period_formula_meta(
            acq, cost, salvage, life, method, disposal, reest_date, reest_life, FY_YEAR,
            reest_method=reest_method, capex_date=capex_date, capex_amount=capex_amount,
            susp_start=susp_start, susp_end=susp_end, prior_period_units=prior_period_units)
        basis = meta["basis"]
        subject_amount = basis - salvage
        life_months_applied = meta["life_months"]
        prior_fy_end_bv = meta["prior_fy_end_bv"]

        # 당기말 누적생산량(생산량비례법 참고용)과 당기말 감가상각누계액(모든 상각방법
        # 공통 롤포워드: 당기말 = 전기말 + 당기감가상각비)은 회사반영/재계산 양쪽 모두
        # 이미 있는 값들의 단순 합산이라 항상 수식으로 연결할 수 있다.
        cur_period_units = (prior_period_units + period_units
                             if (prior_period_units is not None and period_units is not None) else None)
        accum_reported_end = accum_reported + reported if accum_reported is not None else None
        accum_recalc_end = accum_recalc + recalced if accum_recalc is not None else None
        accum_end_diff = (accum_recalc_end - accum_reported_end
                           if (accum_recalc_end is not None and accum_reported_end is not None) else None)

        eff_method = reest_method if (reest_date is not None and reest_method is not None) else method
        eff_life = reest_life if reest_date is not None else life
        if eff_method == "생산량비례법":
            rate_value = "-"
        elif eff_method == "정률법":
            rate_value = get_rate(eff_life)
        elif eff_method == "이중체감법":
            rate_value = 2 / eff_life
        elif eff_method == "연수합계법" and meta["syd_k"] is not None:
            n = meta["syd_life_years"]
            rate_value = (n - meta["syd_k"] + 1) / (n * (n + 1) / 2)
        else:
            # 정액법(그리고 syd_k를 못 구한 예외적 연수합계법)은 원 내용연수(년)가
            # 아니라 "당기에 적용되는 세그먼트의 실제 상각기간"으로 상각률을 구해야
            # 한다 — 자본적지출만 있고 재추정이 없으면 잔여 내용연수(개월)가 원
            # 내용연수*12보다 짧기 때문(자본적지출 시점부터 원래 내용연수 종료월까지
            # 남은 개월수로 재상각하는 정책).
            rate_value = 12 / life_months_applied

        # "일치여부"류/"수식여부"류 열은 재계산결과 시트에는 더 이상 내보내지 않지만
        # (차이금액 열로 대체 가능해 중복이라 사용자 요청으로 삭제),
        # 엑셀 수식 주입(_inject_recalc_formulas) 시 "수식으로 표현 가능한 케이스인지"
        # 판단은 여전히 필요하므로 별도 리스트(formula_flags)에만 기록해둔다.
        formula_flags.append(dict(
            current_ok=meta["is_simple_current"],
            accum_ok=meta["accum_formula_eligible"],
            has_capex=capex_date is not None,
            # 상각중단은 "마지막 구간 안에서 발생하면 그만큼 내용연수 종료월을
            # 늘린다"(apply_suspension_extension)는 규칙이 있어 단순 MONTH() 산식
            # 하나로 재현하기 어렵고, 생산량비례법은 "내용연수(년)" 자체가 더미값이라
            # 개월수 개념이 없다 — 둘 다 적용내용연수(개월)/당기해당월수는 파이썬 값을
            # 그대로 둔다(자본적지출과 동일한 취급).
            has_susp=susp_start is not None,
            is_units=(method == "생산량비례법"),
        ))

        diff = recalced - reported
        match = "일치" if diff == 0 else "불일치"
        materiality = classify_materiality(diff, MATERIALITY_THRESHOLD)

        # 규칙 기반 분류는 비용이 들지 않으므로 "불일치"인 모든 자산에 시도하고,
        # 규칙으로 설명이 안 되는 애매한 경우에 한해서만(그리고 유의한 차이일 때만)
        # 비용이 드는 AI 호출로 폴백한다.
        cause = "-"
        cause_source = "-"
        if match == "불일치":
            rule_cause = get_rule_based_cause(
                reest_date is not None, disposal is not None, note,
                is_capex=capex_date is not None, is_suspended=susp_start is not None)
            if rule_cause is not None:
                cause, cause_source = rule_cause, "규칙"
            elif materiality == "유의한 차이":
                cause = get_ai_estimated_cause(
                    diff, elapsed, method, reest_date is not None, disposal is not None)
                cause_source = "AI" if cause != "-" else "-"

        rows.append({
            "자산명": p["자산명"],
            "자산분류": p["자산분류"],
            "취득일": acq,
            "취득원가": cost,
            "잔존가치": salvage,
            "상각방법": method,
            "내용연수(년)": life,
            "상각률": rate_value,
            "처분일": disposal,
            "내용연수재추정일": reest_date,
            "재추정후내용연수(년)": reest_life,
            "재추정후상각방법": reest_method,
            "총예정생산량": total_units,
            "당기실제생산량": period_units,
            "전기말누적생산량": prior_period_units,
            "당기말누적생산량": cur_period_units,
            "자본적지출일": capex_date,
            "자본적지출액": capex_amount,
            "상각중단시작일": susp_start,
            "상각중단종료일": susp_end,
            "경과개월수(취득일~기준일)": elapsed,
            "당기해당월수": cur_months,
            "적용내용연수(개월)": life_months_applied,
            "상각기준가액(당기)": basis,
            "감가상각대상금액": subject_amount,
            "전기말장부가액": prior_fy_end_bv,
            "내용연수종료여부": "종료" if life_ended else "-",
            "비고": note,
            "회사반영_당기감가상각비": reported,
            "재계산_당기감가상각비": recalced,
            "차이(재계산-회사반영)": diff,
            "회사반영_전기말감가상각누계액": accum_reported,
            "재계산_전기말감가상각누계액": accum_recalc,
            "누계액차이(재계산-회사반영)": accum_diff,
            "회사반영_당기말감가상각누계액": accum_reported_end,
            "재계산_당기말감가상각누계액": accum_recalc_end,
            "당기말누계액차이(재계산-회사반영)": accum_end_diff,
            "중요성구분": materiality,
            "추정원인": cause,
            "추정원인출처": cause_source,
        })

    result_df = pd.DataFrame(rows)
    diff_df = result_df[result_df["차이(재계산-회사반영)"] != 0].copy()
    diff_df = diff_df.reindex(diff_df["차이(재계산-회사반영)"].abs().sort_values(ascending=False).index)
    # 감사자가 직접 채워 넣는 검토란. diff_df에서 파생되는 material_diff_df에도
    # 그대로 이어지므로("차이자산"/"유의한차이자산" 두 시트 모두 대상) 여기 한 곳에만 추가한다.
    diff_df["검토자"] = ""
    diff_df["검토의견"] = ""
    diff_df["검토일"] = pd.NaT
    material_diff_df = diff_df[diff_df["중요성구분"] == "유의한 차이"].copy()
    error_df = pd.DataFrame(error_rows, columns=error_cols)
    category_summary_df = build_category_summary(result_df, error_df)

    # 다기간(연도별) 비교: COMPARISON_YEARS가 [FY_YEAR] 하나뿐이면(기본값) 이 블록은
    # 건너뛰고, 기존 4개+통계 시트 출력이 이전과 완전히 동일하게 유지된다.
    trend_df = None
    pivot_df = None
    if len(COMPARISON_YEARS) > 1:
        trend_df = detect_yoy_anomalies(build_multi_year_trend_df(df, cols, COMPARISON_YEARS),
                                         threshold_pct=YOY_ANOMALY_THRESHOLD_PCT)
        pivot_df = trend_df.pivot_table(
            index=["자산ID", "자산명", "자산분류", "상각방법"],
            columns="회계연도", values="재계산_당기감가상각비",
        ).reset_index()

    # "재계산결과" 시트의 상각률/재계산 수치를 실제 엑셀 수식으로 연결하기 위한
    # 참조용 시트 2개. 법인세법 시행규칙 [별표4] 상각률표 원문(정률법 VLOOKUP 원본)과,
    # 기준일/전기말일자/중요성기준금액 등 수식에서 참조할 기준값을 담는다.
    rate_table_df = pd.DataFrame(
        {
            "내용연수(년)": list(range(2, 61)),
            "정액법할푼리": [STRAIGHT_LINE_RATE_TABLE[n] for n in range(2, 61)],
            "정률법할푼리": [RATE_TABLE[n] for n in range(2, 61)],
        }
    )

    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
        result_df.to_excel(writer, sheet_name="재계산결과", index=False)
        diff_df.to_excel(writer, sheet_name="차이자산", index=False)
        material_diff_df.to_excel(writer, sheet_name="유의한차이자산", index=False)
        error_df.to_excel(writer, sheet_name="데이터오류", index=False)
        category_summary_df.to_excel(writer, sheet_name="자산군별통계", index=False)
        if trend_df is not None:
            trend_df.to_excel(writer, sheet_name="연도별추이", index=False)
            pivot_df.to_excel(writer, sheet_name="연도별요약", index=False)
        rate_table_df.to_excel(writer, sheet_name=RATE_TABLE_SHEET, index=False)
        _write_info_sheet(writer.book, REF_DATE, FY_YEAR, dt.date(FY_YEAR - 1, 12, 31),
                           MATERIALITY_THRESHOLD, OVERALL_MATERIALITY, PERFORMANCE_MATERIALITY,
                           has_trend=trend_df is not None)
        _inject_recalc_formulas(writer.book["재계산결과"], result_df, formula_flags)
        _format_workbook(writer, result_df, diff_df, material_diff_df, error_df, category_summary_df,
                          trend_df, pivot_df, rate_table_df)

    print("DONE:", OUT_PATH)
    print(f"기준일: {REF_DATE}, 당기: {FY_YEAR}년, AMPT(유의한 차이 판정기준): {MATERIALITY_THRESHOLD:,}원")
    print(f"총 {len(result_df) + len(error_df)}건 중 정상 계산 {len(result_df)}건, "
          f"데이터 오류로 계산 제외 {len(error_df)}건")
    print(f"정상 계산 {len(result_df)}건 중 불일치 {len(diff_df)}건 "
          f"(유의한 차이 {len(material_diff_df)}건 / 경미한 차이 {len(diff_df) - len(material_diff_df)}건)")
    if len(error_df) > 0:
        print(f"[!!] 데이터 오류로 제외된 자산 {len(error_df)}건 → '데이터오류' 시트에서 사유를 확인하세요.")
    print("=== 자산군별 통계 ===")
    print(category_summary_df.to_string(index=False))
    if trend_df is not None:
        warn_count = len(trend_df[trend_df["이상탐지"] == "경고"])
        print(f"=== 다기간 비교({COMPARISON_YEARS}) === "
              f"전년대비 {YOY_ANOMALY_THRESHOLD_PCT}% 이상 증감 경고 {warn_count}건")


def _inject_recalc_formulas(ws, df, formula_flags):
    """
    '재계산결과' 시트의 파생 컬럼(경과개월수/당기해당월수/적용내용연수/상각률/
    감가상각대상금액/재계산_당기감가상각비/차이/재계산_전기말감가상각누계액/누계액차이/
    당기말누적생산량/회사반영·재계산_당기말감가상각누계액/당기말누계액차이)을 실제 엑셀
    셀 수식으로 덮어쓴다. df는 이미 to_excel로 값이 쓰인 것과 같은 DataFrame(순서/값
    동일)이고, formula_flags(main()에서 만든, df와 같은 순서의 리스트)의
    current_ok/accum_ok/has_capex로 단순(수식 가능)/복합(값 유지) 케이스만 구분한다
    — 상각비 숫자 자체는 이미 값으로 들어가 있으므로, 수식으로 바꿔도 동일한 값이
    나오도록 재현하는 것이 목표다(실제 계산 로직은 여전히 Python이 담당).

    전기말 감가상각누계액은 "(취득원가+자본적지출)-전기말장부가액"이라는 항등식을
    이용한다(투입금액 총액에서 현재 남은 장부가액을 빼면 그동안 상각된 누계액과
    같다는 관계 — 정액법/정률법/이중체감법/연수합계법 모두, 이벤트 유무와 무관하게
    성립). 당기말은 모든 상각방법 공통으로 "전기말+당기감가상각비" 롤포워드로 구한다.

    "중요성구분" 등 분류 컬럼은 건드리지 않으므로, 이 함수는 _format_workbook의
    "불일치 행 빨간칠" 조건부서식(값 기준 판정)보다 앞이든 뒤든 순서에 상관없이
    호출해도 된다.
    """
    col_names = list(df.columns)
    letter = {name: get_column_letter(col_names.index(name) + 1) for name in col_names}

    for i in range(len(df)):
        r = i + 2  # 엑셀 행(1행은 헤더)
        row = df.iloc[i]
        flags = formula_flags[i]

        acq = f'{letter["취득일"]}{r}'
        cost = f'{letter["취득원가"]}{r}'
        salvage = f'{letter["잔존가치"]}{r}'
        method = f'{letter["상각방법"]}{r}'
        life = f'{letter["내용연수(년)"]}{r}'
        rate = f'{letter["상각률"]}{r}'
        reest = f'{letter["내용연수재추정일"]}{r}'
        reest_life = f'{letter["재추정후내용연수(년)"]}{r}'
        reest_method = f'{letter["재추정후상각방법"]}{r}'
        elapsed = f'{letter["경과개월수(취득일~기준일)"]}{r}'
        months = f'{letter["당기해당월수"]}{r}'
        applied_life_months = f'{letter["적용내용연수(개월)"]}{r}'
        basis = f'{letter["상각기준가액(당기)"]}{r}'
        subject = f'{letter["감가상각대상금액"]}{r}'
        prior_bv = f'{letter["전기말장부가액"]}{r}'
        disposal_dt = f'{letter["처분일"]}{r}'
        capex_dt = f'{letter["자본적지출일"]}{r}'
        capex_amt = f'{letter["자본적지출액"]}{r}'
        susp_start_dt = f'{letter["상각중단시작일"]}{r}'
        susp_end_dt = f'{letter["상각중단종료일"]}{r}'
        period_units = f'{letter["당기실제생산량"]}{r}'
        total_units = f'{letter["총예정생산량"]}{r}'
        prior_units = f'{letter["전기말누적생산량"]}{r}'
        cur_units = f'{letter["당기말누적생산량"]}{r}'
        reported = f'{letter["회사반영_당기감가상각비"]}{r}'
        recalced = f'{letter["재계산_당기감가상각비"]}{r}'
        accum_reported = f'{letter["회사반영_전기말감가상각누계액"]}{r}'
        accum_recalc = f'{letter["재계산_전기말감가상각누계액"]}{r}'
        accum_reported_end = f'{letter["회사반영_당기말감가상각누계액"]}{r}'
        accum_recalc_end = f'{letter["재계산_당기말감가상각누계액"]}{r}'

        eff_method = f'IF({reest}<>"",IF({reest_method}<>"",{reest_method},{method}),{method})'
        eff_life = f'IF({reest}<>"",{reest_life},{life})'
        # 당기 상각비 수식(정률법/이중체감법 하한 판정)에서 쓰는 취득가액: 이 수식은
        # "당기수식여부=수식"인 행에서만 쓰이는데, 그 조건 자체가 "당기 회계연도 안에서
        # 자본적지출이 발생하지 않음(=발생했다면 반드시 그 전 회계연도)"을 내포하므로
        # 자본적지출액을 무조건 더해도 된다.
        tax_cost_expr = f'({cost}+IF({capex_amt}<>"",{capex_amt},0))'
        # 전기말 누계액 수식에서 쓰는 취득가액은 다르다 — 자본적지출이 당기(또는 그
        # 이후)에 발생했다면 전기말 시점에는 아직 반영되지 않았어야 하므로, 자본적지출일이
        # 전기말일자 이전일 때만 더한다.
        tax_cost_prior_expr = (
            f'({cost}+IF(AND({capex_amt}<>"",{capex_dt}<=\'{INFO_SHEET}\'!$E$4),{capex_amt},0))'
        )

        # 경과개월수(취득일~기준일)는 이벤트와 무관한 순수 달력 계산이라 예외 없이
        # 항상 수식으로 넣는다.
        ws[elapsed] = (
            f"=(YEAR('{INFO_SHEET}'!$E$2)-YEAR({acq}))*12"
            f"+MONTH('{INFO_SHEET}'!$E$2)-MONTH({acq})+1"
        )

        # 연수합계법의 그 해 몫(N-k+1)/(N(N+1)/2) 계산에 쓰는 k(취득 또는 재추정
        # 시점부터 몇 번째 해인지). 자본적지출은 이 기준연도를 바꾸지 않는다.
        syd_k_expr = f"('{INFO_SHEET}'!$E$3-YEAR(IF({reest}<>\"\",{reest},{acq}))+1)"

        # 정률법은 VLOOKUP(내용연수-년)을 그대로 쓰고, 이중체감법은 2/내용연수를
        # 직접 계산하며, 연수합계법은 위 k로 그 해의 몫을 계산한다. 정액법은 원
        # 내용연수(년)가 아니라 "적용내용연수(개월)"(자본적지출로 잔여기간이
        # 줄었을 수 있음)로 상각률을 구해야 한다 — main()의 rate_value 계산과 동일.
        ws[rate] = (
            f'=IF({eff_method}="정률법",VLOOKUP({eff_life},\'{RATE_TABLE_SHEET}\'!$A$2:$C$60,3,FALSE),'
            f'IF({eff_method}="생산량비례법","-",'
            f'IF({eff_method}="이중체감법",2/{eff_life},'
            f'IF({eff_method}="연수합계법",({eff_life}-{syd_k_expr}+1)/({eff_life}*({eff_life}+1)/2),'
            f'12/{applied_life_months}))))'
        )
        ws[subject] = f"={basis}-{salvage}"
        ws[f'{letter["차이(재계산-회사반영)"]}{r}'] = f"={recalced}-{reported}"

        if flags["current_ok"]:
            ws[recalced] = (
                f'=IF({eff_method}="생산량비례법",({cost}-{salvage})*{period_units}/{total_units},'
                f'IF({eff_method}="정률법",'
                f'IF({basis}-{basis}*{rate}*{months}/12<={tax_cost_expr}*0.05,{subject},'
                f'{basis}*{rate}*{months}/12),'
                f'IF({eff_method}="이중체감법",'
                f'IF({basis}-{basis}*{rate}*{months}/12<={salvage},{subject},'
                f'{basis}*{rate}*{months}/12),'
                f'{subject}*{rate}*{months}/12)))'
            )

        # 적용내용연수(개월)/당기해당월수는 자본적지출·상각중단이 없고 생산량비례법도
        # 아닌 "단순" 행에서만 MONTH() 수식으로 넣는다. 자본적지출은 잔여 내용연수를
        # 바꿔버리고, 상각중단은 마지막 구간 안에서 발생하면 내용연수 종료월 자체를
        # 늘리며, 생산량비례법은 "내용연수(년)"가 계산에 안 쓰이는 더미값이라 —
        # 셋 다 세그먼트 재구성 로직을 엑셀 수식 하나로 그대로 옮기기엔 지나치게
        # 복잡해지므로, 그 경우는 지금처럼 파이썬이 계산한 값을 그대로 둔다.
        if flags["current_ok"] and not flags["has_capex"] and not flags["has_susp"] and not flags["is_units"]:
            ws[applied_life_months] = f'=IF({reest}<>"",{reest_life},{life})*12'

            seg_start_idx = f'IF({reest}<>"",YEAR({reest})*12+MONTH({reest}),YEAR({acq})*12+MONTH({acq}))'
            seg_end_idx = f'({seg_start_idx}+{applied_life_months}-1)'
            fy_start_idx = f"('{INFO_SHEET}'!$E$3*12+1)"
            fy_end_idx = f"('{INFO_SHEET}'!$E$3*12+12)"
            disposal_idx = f'IF({disposal_dt}<>"",YEAR({disposal_dt})*12+MONTH({disposal_dt}),9999999)'
            overlap_start = f'MAX({seg_start_idx},{fy_start_idx})'
            overlap_end = f'MIN({seg_end_idx},{fy_end_idx},{disposal_idx})'
            raw_months = f'MAX(0,{overlap_end}-{overlap_start}+1)'
            susp_s_idx = f'YEAR({susp_start_dt})*12+MONTH({susp_start_dt})'
            susp_e_idx = f'YEAR({susp_end_dt})*12+MONTH({susp_end_dt})'
            susp_overlap = (
                f'IF({susp_start_dt}<>"",MAX(0,MIN({overlap_end},{susp_e_idx})'
                f'-MAX({overlap_start},{susp_s_idx})+1),0)'
            )
            ws[months] = f'=MAX(0,{raw_months}-{susp_overlap})'

        # 당기말누적생산량(생산량비례법 참고용, 다른 방법은 두 값 다 비어있어 자동으로 빈 값)
        ws[cur_units] = f'=IF(OR({prior_units}="",{period_units}=""),"",{prior_units}+{period_units})'

        if flags["accum_ok"]:
            ws[accum_recalc] = (
                f'=IF({method}="생산량비례법",IF({prior_units}<>"",({cost}-{salvage})*{prior_units}/{total_units},""),'
                f'{tax_cost_prior_expr}-{prior_bv})'
            )

        if pd.notna(row["누계액차이(재계산-회사반영)"]):
            ws[f'{letter["누계액차이(재계산-회사반영)"]}{r}'] = f"={accum_recalc}-{accum_reported}"

        # 당기말 감가상각누계액 = 전기말 + 당기감가상각비 롤포워드(모든 상각방법 공통).
        # 전기말 셀이 비어 있으면(예: 생산량비례법인데 전기말누적생산량 미입력) 그대로 공란.
        ws[accum_reported_end] = f'=IF({accum_reported}="","",{accum_reported}+{reported})'
        ws[accum_recalc_end] = f'=IF({accum_recalc}="","",{accum_recalc}+{recalced})'

        if pd.notna(row["당기말누계액차이(재계산-회사반영)"]):
            ws[f'{letter["당기말누계액차이(재계산-회사반영)"]}{r}'] = f"={accum_recalc_end}-{accum_reported_end}"


def _write_info_sheet(wb, ref_date, fy_year, prior_fy_end, materiality_threshold,
                       overall_materiality, performance_materiality, has_trend=False):
    """
    '기준정보' 시트를 워크북 맨 앞(0번 탭)에 만든다. A열에는 이 프로그램이 어떻게
    동작하는지 / 결과 파일이 어떤 과정을 거쳐 만들어지는지 / 시트별로 무엇을
    보여주는지를 설명하는 안내문을, D~E열에는 이번 실행에 쓰인 기준값(기준일 등)
    표를 둔다. 설명문 길이가 늘어나도 D~E열 표 위치(1~4행)는 영향받지 않도록
    분리해뒀다 — _inject_recalc_formulas()가 이 표의 전기말일자 셀(E4)을
    '기준정보'!$E$4로 직접 참조하므로, 이 표의 4행까지는 위치를 바꾸면 안 된다
    (5~7행 AMPT/중요성/수행중요성은 뒤에 이어붙인 것이라 안전).
    """
    ws = wb.create_sheet(INFO_SHEET, 0)
    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)

    lines = [
        ("이 결과 파일 읽는 법", title_font),
        ("", None),
        ("1. 이 프로그램은 무엇을 하나요?", bold),
        ("회사가 작성한 고정자산대장(엑셀) 하나를 입력받아, 자산마다 정액법/정률법/이중체감법/연수합계법/", None),
        ("생산량비례법 등 자기 상각방법에 맞는 당기 감가상각비를 이 프로그램이 회사와 무관하게 독립적으로 다시 계산합니다.", None),
        ("그 재계산 금액과 회사가 장부에 반영한 금액을 자산별로 대조해서, 금액이 다른 자산만 골라 보여줍니다.", None),
        ("즉 \"회사가 계산한 감가상각비가 맞는지\"를 사람이 일일이 엑셀 수식으로 대조하지 않고 자동으로 검증하는 도구입니다.", None),
        ("", None),
        ("2. 결과 파일은 어떤 과정을 거쳐 만들어지나요?", bold),
        ("① 컬럼 인식 — 입력 파일에서 취득일·취득원가·상각방법 등 필요한 컬럼을 찾습니다(컬럼명이 달라도 매핑 가능).", None),
        ("② 입력값 검증 — 내용연수가 0 이하이거나 잔존가치가 취득원가보다 큰 것처럼 계산 자체가 불가능한", None),
        ("   자산은 재계산을 시도하지 않고 [데이터오류] 시트로 분리합니다.", None),
        ("③ 자산별 재계산 — 나머지 자산은 취득일부터 기준일까지 월할상각(정액법), 연도별 장부가액 감소", None),
        ("   (정률법·이중체감법), 연차별 몫 배분(연수합계법), 생산량 비율(생산량비례법)로 당기 감가상각비를", None),
        ("   다시 계산합니다. 처분·내용연수 재추정·자본적지출·감가상각중단 같은 이벤트가 있는 자산도 그", None),
        ("   이벤트를 반영해 계산합니다.", None),
        ("④ 비교 및 분류 — 회사반영 금액과 재계산 금액을 비교해 일치/불일치, 중요성 기준(AMPT, 아래 표) 이상", None),
        ("   차이나면 유의한 차이로 구분합니다. 전기말 감가상각누계액도 참고용으로 같은 방식으로 비교합니다", None),
        ("   (당기 감가상각비 일치 판정에는 영향 없음).", None),
        ("⑤ 원인 추정 — 불일치 자산 중 처분·재추정 등 이미 계산된 정보로 설명되는 경우는 규칙으로, 애매한", None),
        ("   경우는 (설정된 경우) AI로 차이 원인을 한 줄 추정합니다.", None),
        ("⑥ 엑셀 저장 — 위 결과를 시트별로 정리해 저장합니다. 상각률·재계산 감가상각비 등 재계산 관련 수치는", None),
        ("   고정된 값이 아니라 실제 엑셀 수식으로 연결돼 있어, 셀을 눌러보면 어떻게 계산됐는지 확인하거나", None),
        ("   내용연수 같은 입력값을 바꿔가며 재계산 결과가 어떻게 달라지는지 직접 시험해볼 수 있습니다.", None),
        ("", None),
        ("3. 시트별 설명", bold),
        ("기준정보 — 지금 보고 계신 이 시트. 프로그램 안내와 이 결과 파일을 만들 때 쓰인 기준값.", None),
        ("   아래 표의 AMPT가 실제 \"유의한 차이\" 판정에 쓰이는 값이고, 중요성/수행중요성(PM)은 참고용", None),
        ("   정보로만 표시됩니다(재계산결과 시트의 판정 로직에는 관여하지 않습니다).", None),
        ("재계산결과 — 전체 자산의 상세 재계산 내역(입력값·계산과정·회사반영/재계산 금액·차이 등). 차이 열이", None),
        ("   0이면 일치, 0이 아니면 불일치인 자산입니다.", None),
        ("차이자산 — [재계산결과] 중 회사반영 금액과 재계산 금액이 다른 자산만 모은 시트(차이금액 큰 순).", None),
        ("유의한차이자산 — [차이자산] 중에서도 중요성 기준(AMPT, 아래 표) 이상 차이나는 자산만 별도로 정리한 시트.", None),
        ("데이터오류 — 취득원가·잔존가치·내용연수 값 자체가 계산 불가능해 재계산을 시도하지 않은 자산과 그 사유.", None),
        ("자산군별통계 — 유형자산/무형자산 등 자산분류별 전체건수·불일치건수·불일치율 등 집계.", None),
        ("상각률표(별표4) — 정률법 재계산에 쓰인 법인세법 시행규칙 [별표4] 상각률표 원문. [재계산결과]의", None),
        ("   \"상각률\" 수식이 이 표를 참조합니다.", None),
        ("연도별추이 / 연도별요약 — 여러 회계연도를 비교 설정했을 때만 생성되는, 자산별 연도별 상각비 추이와" +
         ("(이번 결과 파일에 포함됨)" if has_trend else "(이번 결과 파일에는 없음 — 단일 연도만 계산)"), None),
        ("   전년대비 이상 증감 탐지 시트.", None),
    ]
    for row_idx, (text, font) in enumerate(lines, start=1):
        c = ws.cell(row=row_idx, column=1, value=text)
        if font is not None:
            c.font = font

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_cell = ws.cell(row=1, column=4, value="이 결과 파일의 기준값")
    title_cell.font, title_cell.fill = header_font, header_fill
    ws.cell(row=1, column=5).fill = header_fill
    # 전기말일자는 _inject_recalc_formulas()가 '기준정보'!$E$4로 직접 참조하므로
    # 이 표의 행 위치(1행=제목, 2~5행=값)를 바꿀 때는 그 참조도 함께 고쳐야 한다.
    table = [
        ("기준일", ref_date, "yyyy-mm-dd"),
        ("당기회계연도", fy_year, None),
        ("전기말일자", prior_fy_end, "yyyy-mm-dd"),
        ("AMPT", materiality_threshold, "#,##0"),
        ("중요성", overall_materiality, "#,##0"),
        ("수행중요성(PM)", performance_materiality, "#,##0"),
    ]
    for i, (label, value, fmt) in enumerate(table, start=2):
        ws.cell(row=i, column=4, value=label)
        vc = ws.cell(row=i, column=5, value=value)
        if fmt is not None:
            vc.number_format = fmt

    ws.column_dimensions["A"].width = 105
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16


def _format_workbook(writer, result_df, diff_df, material_diff_df, error_df, category_summary_df,
                      trend_df=None, pivot_df=None, rate_table_df=None):
    wb = writer.book
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    mismatch_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    error_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    review_input_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # "재계산결과" 시트 전용 헤더 색상 — 취득~비고(입력값/이벤트/계산과정), 당기 비교,
    # 전기말 누계 비교, 당기말 누계 비교, 분류·메타 5개 논리 그룹을 구분해서 보여준다.
    # 하드코딩된 열문자 대신 컬럼명으로 그룹 경계를 찾으므로 컬럼 구성이 또 바뀌어도
    # (그룹 시작 컬럼명 자체가 없어지지만 않으면) 깨지지 않는다.
    result_header_fills = [
        header_fill,  # 그룹1: 자산명~비고(취득값/이벤트/계산과정)
        PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid"),  # 그룹2: 당기 비교
        PatternFill(start_color="548235", end_color="548235", fill_type="solid"),  # 그룹3: 전기말 누계 비교
        PatternFill(start_color="31859C", end_color="31859C", fill_type="solid"),  # 그룹4: 당기말 누계 비교
        PatternFill(start_color="7F7F7F", end_color="7F7F7F", fill_type="solid"),  # 그룹5: 분류/메타
    ]

    date_cols = ["취득일", "처분일", "내용연수재추정일", "검토일",
                 "자본적지출일", "상각중단시작일", "상각중단종료일"]
    money_cols = ["취득원가", "잔존가치", "회사반영_당기감가상각비", "재계산_당기감가상각비",
                  "차이(재계산-회사반영)", "자본적지출액", "상각기준가액(당기)", "감가상각대상금액",
                  "전기말장부가액", "총예정생산량", "당기실제생산량", "전기말누적생산량", "당기말누적생산량",
                  "회사반영_전기말감가상각누계액", "재계산_전기말감가상각누계액",
                  "누계액차이(재계산-회사반영)", "회사반영_당기말감가상각누계액", "재계산_당기말감가상각누계액",
                  "당기말누계액차이(재계산-회사반영)"]
    rate_cols = ["상각률", "정액법할푼리", "정률법할푼리"]

    sheets = [
        ("재계산결과", result_df), ("차이자산", diff_df),
        ("유의한차이자산", material_diff_df), ("데이터오류", error_df),
        ("자산군별통계", category_summary_df),
    ]
    if trend_df is not None:
        sheets += [("연도별추이", trend_df), ("연도별요약", pivot_df)]
    if rate_table_df is not None:
        sheets.append((RATE_TABLE_SHEET, rate_table_df))

    for sheet_name, df in sheets:
        ws = wb[sheet_name]
        n_rows, n_cols = df.shape
        col_names = list(df.columns)

        group_bounds = None
        if sheet_name == "재계산결과":
            try:
                group_bounds = (
                    col_names.index("회사반영_당기감가상각비"),
                    col_names.index("회사반영_전기말감가상각누계액"),
                    col_names.index("회사반영_당기말감가상각누계액"),
                    col_names.index("중요성구분"),
                )
            except ValueError:
                group_bounds = None  # 컬럼 구성이 달라지면 안전하게 단일 헤더색으로 폴백

        for col_idx in range(1, n_cols + 1):
            c = ws.cell(row=1, column=col_idx)
            c.font = header_font
            if group_bounds is not None:
                idx0 = col_idx - 1
                g2, g3, g4, g5 = group_bounds
                if idx0 < g2:
                    c.fill = result_header_fills[0]
                elif idx0 < g3:
                    c.fill = result_header_fills[1]
                elif idx0 < g4:
                    c.fill = result_header_fills[2]
                elif idx0 < g5:
                    c.fill = result_header_fills[3]
                else:
                    c.fill = result_header_fills[4]
            else:
                c.fill = header_fill
            c.alignment = header_align

        for col_idx, col_name in enumerate(col_names, start=1):
            letter = get_column_letter(col_idx)
            # "연도별요약"은 회계연도(정수)가 컬럼 헤더가 되는 pivot 결과라 money_cols
            # 이름 목록으로는 못 잡으므로, 이 시트에서는 정수 헤더를 금액으로 취급한다.
            is_pivot_year_col = sheet_name == "연도별요약" and isinstance(col_name, int)
            if col_name in date_cols:
                for row_idx in range(2, n_rows + 2):
                    ws.cell(row=row_idx, column=col_idx).number_format = "yyyy-mm-dd"
            elif col_name in money_cols or is_pivot_year_col:
                for row_idx in range(2, n_rows + 2):
                    ws.cell(row=row_idx, column=col_idx).number_format = "#,##0"
            elif col_name in rate_cols:
                for row_idx in range(2, n_rows + 2):
                    ws.cell(row=row_idx, column=col_idx).number_format = "0.000"
            if col_name in ("추정원인", "오류사유", "검토의견"):
                ws.column_dimensions[letter].width = 60
            else:
                ws.column_dimensions[letter].width = max(14, len(str(col_name)) + 4)

        # 불일치 판정은 "차이(재계산-회사반영)"의 실제 파이썬 값(df)으로 한다 — 이
        # 시트의 워크셀 값 자체는 _inject_recalc_formulas가 수식 문자열로 덮어써서
        # openpyxl로는 계산 결과를 알 수 없기 때문이다(엑셀이 열어야 계산됨).
        if "차이(재계산-회사반영)" in col_names and n_rows > 0:
            for pos in range(n_rows):
                diff_value = df.iloc[pos]["차이(재계산-회사반영)"]
                if pd.notna(diff_value) and diff_value != 0:
                    row_idx = pos + 2
                    for c_idx in range(1, n_cols + 1):
                        ws.cell(row=row_idx, column=c_idx).fill = mismatch_fill

        if sheet_name == "데이터오류" and n_rows > 0:
            for row_idx in range(2, n_rows + 2):
                for c_idx in range(1, n_cols + 1):
                    ws.cell(row=row_idx, column=c_idx).fill = error_fill

        # 검토란(검토자/검토의견/검토일)은 감사자가 직접 입력하는 영역이므로,
        # 위의 "불일치 행 전체 빨간색" 칠보다 뒤에 별도 색으로 덧칠해 구분한다.
        if sheet_name in ("차이자산", "유의한차이자산") and n_rows > 0:
            for col_name in ("검토자", "검토의견", "검토일"):
                if col_name in col_names:
                    col_idx = col_names.index(col_name) + 1
                    for row_idx in range(2, n_rows + 2):
                        ws.cell(row=row_idx, column=col_idx).fill = review_input_fill

        ws.freeze_panes = "A2"
        if n_rows > 0:
            ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{n_rows + 1}"


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="고정자산 감가상각비 재계산 검증 도구")
    parser.add_argument("--config", "-c", default=DEFAULT_CONFIG_PATH,
                         help=f"설정 파일 경로 (기본값: {DEFAULT_CONFIG_PATH})")
    parser.add_argument("--input", "-i", default=None,
                         help="입력 고정자산대장 엑셀 경로 (기본값: 설정 파일의 input_path)")
    parser.add_argument("--output", "-o", default=None,
                         help="결과 엑셀 저장 경로 (기본값: 설정 파일의 output_path)")
    args = parser.parse_args()

    # --config로 지정한 설정 파일을 다시 읽어 모든 설정을 갱신한다(기본값과 같은
    # 경로라도 한 번 더 읽을 뿐이라 멱등하게 안전하다 — 특별 분기 불필요).
    settings = resolve_settings(load_config(args.config))
    IN_PATH = settings["IN_PATH"]
    OUT_PATH = settings["OUT_PATH"]
    REF_DATE = settings["REF_DATE"]
    FY_YEAR = settings["FY_YEAR"]
    MATERIALITY_THRESHOLD = settings["MATERIALITY_THRESHOLD"]
    OVERALL_MATERIALITY = settings["OVERALL_MATERIALITY"]
    PERFORMANCE_MATERIALITY = settings["PERFORMANCE_MATERIALITY"]
    ANTHROPIC_MODEL = settings["ANTHROPIC_MODEL"]
    COMPARISON_YEARS = settings["COMPARISON_YEARS"]
    YOY_ANOMALY_THRESHOLD_PCT = settings["YOY_ANOMALY_THRESHOLD_PCT"]
    COLUMN_MAP = settings["COLUMN_MAP"]

    # --input/--output은 설정 파일보다 항상 우선한다.
    if args.input:
        IN_PATH = args.input
    if args.output:
        OUT_PATH = args.output

    main()
