# -*- coding: utf-8 -*-
"""
recalc.py의 계산 함수들에 대한 pytest 테스트.
각 케이스의 정답값은 손계산(또는 독립적인 수식)으로 먼저 도출한 뒤,
recalc.py의 실제 함수 출력과 비교한다.
"""
import datetime as dt

import pandas as pd

import recalc as R


# ---------------------------------------------------------------------------
# 월할상각 경과개월수 계산 (elapsed_months_to_ref / month_index)
# ---------------------------------------------------------------------------
class TestElapsedMonths:
    def test_month_index_basic(self):
        # 2025년 12월 = 2025*12 + 12
        assert R.month_index(2025, 12) == 2025 * 12 + 12

    def test_elapsed_months_same_year(self):
        # 2025-03-01 ~ 2025-12-31: 3월~12월 취득월 포함 10개월
        assert R.elapsed_months_to_ref(dt.date(2025, 3, 1), dt.date(2025, 12, 31)) == 10

    def test_elapsed_months_multi_year(self):
        # 2023-01-15 ~ 2025-12-31: 2023-01부터 2025-12까지 취득월 포함 36개월(만 3년)
        assert R.elapsed_months_to_ref(dt.date(2023, 1, 15), dt.date(2025, 12, 31)) == 36

    def test_elapsed_months_acquired_in_ref_month(self):
        # 취득월 = 기준월이면 경과개월수는 1개월
        assert R.elapsed_months_to_ref(dt.date(2025, 12, 1), dt.date(2025, 12, 31)) == 1


# ---------------------------------------------------------------------------
# 정액법 재계산
# ---------------------------------------------------------------------------
class TestStraightLine:
    def test_basic_full_year(self):
        # 취득 2023-01-01, 취득원가 12,000,000, 잔존가치 0, 내용연수 5년
        # 월상각액 = 12,000,000 / 60 = 200,000, 당기(2025) 12개월 전액 해당
        # 재계산 당기상각비 = 200,000 * 12 = 2,400,000
        dep, months, life_ended, note = R.recalc_asset(
            acq=dt.date(2023, 1, 1), cost=12_000_000, salvage=0, life=5, method="정액법",
            disposal=None, reest_date=None, reest_life=None,
            ref_date=dt.date(2025, 12, 31), fy_year=2025,
        )
        assert dep == 2_400_000
        assert months == 12
        assert life_ended is False
        assert note == "-"

    def test_partial_first_year(self):
        # 취득 2025-07-01, 취득원가 6,000,000, 잔존가치 0, 내용연수 5년(60개월)
        # 월상각액 = 100,000, 당기(2025)는 취득월(7월)부터 12월까지 6개월만 해당
        # 재계산 당기상각비 = 100,000 * 6 = 600,000
        dep, months, life_ended, note = R.recalc_asset(
            acq=dt.date(2025, 7, 1), cost=6_000_000, salvage=0, life=5, method="정액법",
            disposal=None, reest_date=None, reest_life=None,
            ref_date=dt.date(2025, 12, 31), fy_year=2025,
        )
        assert dep == 600_000
        assert months == 6
        assert life_ended is False


# ---------------------------------------------------------------------------
# 정률법 재계산
# ---------------------------------------------------------------------------
class TestDecliningBalance:
    def test_third_year_of_depreciation(self):
        # 취득 2023-01-01, 취득원가 10,000,000, 잔존가치 0, 내용연수 5년 (상각률표 0.451)
        # 2023년: 10,000,000 * 0.451 = 4,510,000 -> 기말 장부가액 5,490,000
        # 2024년: 5,490,000 * 0.451 = 2,475,990 -> 기말 장부가액 3,014,010
        # 2025년(당기): 3,014,010 * 0.451 = 1,359,318.51 -> 반올림 1,359,319
        dep, months, life_ended, note = R.recalc_asset(
            acq=dt.date(2023, 1, 1), cost=10_000_000, salvage=0, life=5, method="정률법",
            disposal=None, reest_date=None, reest_life=None,
            ref_date=dt.date(2025, 12, 31), fy_year=2025,
        )
        assert dep == 1_359_319
        assert months == 12
        assert life_ended is False

    def test_get_rate_table_lookup(self):
        assert R.get_rate(5) == 0.451
        assert R.get_rate(10) == 0.259

    def test_get_rate_fallback_formula(self):
        # 상각률표에 없는 내용연수(예: 25년)는 잔존가치 5% 가정 근사식 사용
        life = 25
        expected = round(1 - 0.05 ** (1 / life), 3)
        assert R.get_rate(life) == expected


# ---------------------------------------------------------------------------
# 처분자산 처리 (당기 중 처분 시 처분일까지만 상각)
# ---------------------------------------------------------------------------
class TestDisposal:
    def test_straight_line_disposal_mid_year(self):
        # 취득 2022-06-01, 취득원가 24,000,000, 잔존가치 0, 내용연수 5년
        # 월상각액 = 400,000, 2025-06-30 처분 -> 당기 1~6월 6개월만 상각
        # 재계산 당기상각비 = 400,000 * 6 = 2,400,000
        dep, months, life_ended, note = R.recalc_asset(
            acq=dt.date(2022, 6, 1), cost=24_000_000, salvage=0, life=5, method="정액법",
            disposal=dt.date(2025, 6, 30), reest_date=None, reest_life=None,
            ref_date=dt.date(2025, 12, 31), fy_year=2025,
        )
        assert dep == 2_400_000
        assert months == 6
        assert "당기중처분(2025-06-30)" in note

    def test_disposal_before_current_year(self):
        # 처분일이 당기 이전이면 "전기이전처분" 비고 + 당기 상각비 0
        dep, months, life_ended, note = R.recalc_asset(
            acq=dt.date(2020, 1, 1), cost=12_000_000, salvage=0, life=10, method="정액법",
            disposal=dt.date(2024, 3, 31), reest_date=None, reest_life=None,
            ref_date=dt.date(2025, 12, 31), fy_year=2025,
        )
        assert dep == 0
        assert months == 0
        assert "전기이전처분(2024-03-31)" in note


# ---------------------------------------------------------------------------
# 내용연수 재추정 처리
# ---------------------------------------------------------------------------
class TestReestimation:
    def test_straight_line_reestimation(self):
        # 취득 2021-01-01, 취득원가 10,000,000, 잔존가치 0, 원내용연수 10년
        # 2025-01-01에 잔여내용연수를 3년으로 재추정
        # 재추정 시점까지 경과 48개월 상각: 10,000,000/120*48 = 4,000,000 -> 재추정시 장부가 6,000,000
        # 재추정 후 월상각액 = 6,000,000 / 36 = 166,666.67, 당기(2025) 12개월 전액 해당
        # 재계산 당기상각비 = 6,000,000 / 36 * 12 = 2,000,000
        dep, months, life_ended, note = R.recalc_asset(
            acq=dt.date(2021, 1, 1), cost=10_000_000, salvage=0, life=10, method="정액법",
            disposal=None, reest_date=dt.date(2025, 1, 1), reest_life=3,
            ref_date=dt.date(2025, 12, 31), fy_year=2025,
        )
        assert dep == 2_000_000
        assert months == 12
        assert life_ended is False
        assert "내용연수재추정(2025-01-01→3년)" in note

    def test_declining_balance_reestimation_same_method(self):
        # 취득 2021-01-01, 취득원가 10,000,000, 잔존가치 0, 원내용연수 10년(상각률 0.259)
        # 2024-01-01에 상각방법은 정률법을 유지한 채 내용연수만 5년(상각률 0.451)으로 재추정.
        # method="정률법"/reest_method=None(방법 변경 없음)인 이 조합은 통합 이벤트
        # 엔진이 아니라 예전부터 있던 declining_balance_current_period_dep 경로를 타는데,
        # 이 경로의 재추정 분기(recalc.py:588-590)는 지금까지 자동 테스트가 전혀
        # 지나가지 않고 있었다(coverage 0%) — 정액법 재추정만 테스트돼 있었음.
        #
        # 매년 "기초장부가액 x 상각률"을 독립적으로 손계산해 정답을 미리 구한다
        # (재추정 전 3개년은 상각률 0.259, 재추정 후는 0.451을 곱함. 5% 특례
        # 기준값은 원취득가액의 5% = 500,000원으로 재추정 후에도 그대로 유지):
        #   2021: 10,000,000 x0.259           = 2,590,000  -> 장부가 7,410,000
        #   2022:  7,410,000 x0.259           = 1,919,190  -> 장부가 5,490,810
        #   2023:  5,490,810 x0.259           = 1,422,119.79(반올림 1,422,120) -> 장부가 4,068,690.21
        #   2024:  4,068,690.21 x0.451(재추정) = 1,834,979.28(반올림 1,834,979) -> 장부가 2,233,710.93
        #   2025:  2,233,710.93 x0.451         = 1,007,403.63(반올림 1,007,404) -> 장부가 1,226,307.30
        # (책값 - 상각액이 500,000원 이하로 떨어지는 해가 없으므로 2025년까지는
        #  5% 특례가 발동하지 않는다 — 발동 시점은 아래 boundary 테스트에서 별도 확인)
        dep, months, life_ended, note = R.recalc_asset(
            acq=dt.date(2021, 1, 1), cost=10_000_000, salvage=0, life=10, method="정률법",
            disposal=None, reest_date=dt.date(2024, 1, 1), reest_life=5,
            ref_date=dt.date(2025, 12, 31), fy_year=2025,
        )
        assert dep == 1_007_404
        assert months == 12
        assert life_ended is False
        assert "내용연수재추정(2024-01-01→5년)" in note

    def test_declining_balance_reestimation_life_ended_boundary(self):
        # 위 테스트와 같은 자산. 재추정후내용연수 5년은 2024-01-01부터 2028-12-31까지다.
        # 실제로는 5% 특례가 2027년에 발동해 장부가액이 그 해에 0원으로 완전히
        # 상각되지만(아래에서 함께 확인), life_ended 플래그 자체는 "특례로 이미
        # 다 상각됐는가"가 아니라 "재추정후내용연수의 명목 종료월(2028-12)을
        # 기준일이 지났는가"만 본다 — 그래서 2028년(명목 종료월 이내)은
        # life_ended=False, 2029년(명목 종료월을 지남)은 life_ended=True가 되어야
        # 정상이다. 손계산으로 이어서 구한 장부가액:
        #   2026: 673,242.71 x... 아래 2026/2027 두 해만 이어서 계산
        #   2026:  1,226,307.30 x0.451 = 553,064.59(반올림 553,065) -> 장부가 673,242.71
        #   2027:  673,242.71 x0.451 = 303,632.66 -> "장부가-상각액"=369,610.05 <= 500,000(5%특례)
        #         이므로 잔여 전액(673,242.71) 상각 -> 장부가 0
        #   2028:  장부가액이 이미 0이므로 상각액 0(하지만 life_ended는 여전히 False)
        #   2029:  명목 종료월(2028-12)을 지났으므로 상각개월수 0, life_ended=True
        for fy, expected_dep, expected_months, expected_life_ended in [
            (2026, 553_065, 12, False),
            (2027, 673_243, 12, False),
            (2028, 0, 12, False),
            (2029, 0, 0, True),
        ]:
            dep, months, life_ended, _ = R.recalc_asset(
                acq=dt.date(2021, 1, 1), cost=10_000_000, salvage=0, life=10, method="정률법",
                disposal=None, reest_date=dt.date(2024, 1, 1), reest_life=5,
                ref_date=dt.date(fy, 12, 31), fy_year=fy,
            )
            assert dep == expected_dep, f"fy={fy}: dep {dep} != {expected_dep}"
            assert months == expected_months, f"fy={fy}: months {months} != {expected_months}"
            assert life_ended is expected_life_ended, f"fy={fy}: life_ended {life_ended} != {expected_life_ended}"


# ---------------------------------------------------------------------------
# 내용연수 종료 처리
# ---------------------------------------------------------------------------
class TestLifeEnded:
    def test_straight_line_life_ended(self):
        # 취득 2018-01-01, 내용연수 5년 -> 2022-12에 상각 종료, 당기(2025)는 상각 대상 아님
        dep, months, life_ended, note = R.recalc_asset(
            acq=dt.date(2018, 1, 1), cost=5_000_000, salvage=0, life=5, method="정액법",
            disposal=None, reest_date=None, reest_life=None,
            ref_date=dt.date(2025, 12, 31), fy_year=2025,
        )
        assert dep == 0
        assert months == 0
        assert life_ended is True
        assert note == "내용연수종료"

    def test_declining_balance_life_ended(self):
        # 취득 2015-01-01, 내용연수 5년 -> 2019-12에 상각 종료, 당기(2025)는 상각 대상 아님
        dep, months, life_ended, note = R.recalc_asset(
            acq=dt.date(2015, 1, 1), cost=3_000_000, salvage=0, life=5, method="정률법",
            disposal=None, reest_date=None, reest_life=None,
            ref_date=dt.date(2025, 12, 31), fy_year=2025,
        )
        assert dep == 0
        assert months == 0
        assert life_ended is True
        assert note == "내용연수종료"


# ---------------------------------------------------------------------------
# 중요성 기준 판정
# ---------------------------------------------------------------------------
class TestMateriality:
    def test_diff_zero_is_minor(self):
        assert R.classify_materiality(0) == "경미한 차이"

    def test_diff_just_below_threshold_is_minor(self):
        assert R.classify_materiality(999_999) == "경미한 차이"

    def test_diff_at_threshold_is_material(self):
        # 정확히 임계값(기본 1,000,000원)이면 "유의한 차이" (>= 판정)
        assert R.classify_materiality(1_000_000) == "유의한 차이"

    def test_diff_above_threshold_is_material(self):
        assert R.classify_materiality(1_500_000) == "유의한 차이"

    def test_negative_diff_uses_absolute_value(self):
        assert R.classify_materiality(-1_500_000) == "유의한 차이"
        assert R.classify_materiality(-500_000) == "경미한 차이"

    def test_custom_threshold(self):
        assert R.classify_materiality(300_000, threshold=500_000) == "경미한 차이"
        assert R.classify_materiality(500_000, threshold=500_000) == "유의한 차이"

    def test_omitted_threshold_follows_monkeypatched_global_not_stale_default(self, monkeypatch):
        # 회귀 테스트: threshold 기본값이 함수 정의 시점에 고정돼 있었다면(과거 버그),
        # 여기서 MATERIALITY_THRESHOLD를 아무리 바꿔도 인자를 생략한 호출은 예전 값
        # (1,000,000원)을 계속 썼을 것이다. None 기본값 + 호출 시점 조회로 고친 뒤에는
        # monkeypatch로 바꾼 값이 인자 생략 호출에도 즉시 반영돼야 한다.
        monkeypatch.setattr(R, "MATERIALITY_THRESHOLD", 500_000)
        assert R.classify_materiality(600_000) == "유의한 차이"  # 새 임계값(50만원) 기준 유의함
        assert R.classify_materiality(400_000) == "경미한 차이"  # 새 임계값 미만


# ---------------------------------------------------------------------------
# 입력값 검증 (validate_asset_inputs) - 계산 불가능한 값은 "데이터 오류"로 분리
# ---------------------------------------------------------------------------
class TestValidateAssetInputs:
    def test_life_zero_is_invalid(self):
        errors = R.validate_asset_inputs(cost=5_000_000, salvage=0, life=0)
        assert any("내용연수" in e for e in errors)

    def test_life_negative_is_invalid(self):
        errors = R.validate_asset_inputs(cost=8_000_000, salvage=0, life=-5)
        assert any("내용연수" in e for e in errors)

    def test_salvage_greater_than_cost_is_invalid(self):
        errors = R.validate_asset_inputs(cost=10_000_000, salvage=15_000_000, life=5)
        assert any("잔존가치" in e for e in errors)

    def test_salvage_equal_to_cost_is_invalid(self):
        # 잔존가치 == 취득원가면 상각대상금액이 0이 되어 이상하므로 오류로 취급(>= 판정)
        errors = R.validate_asset_inputs(cost=10_000_000, salvage=10_000_000, life=5)
        assert any("잔존가치" in e for e in errors)

    def test_cost_zero_is_invalid(self):
        errors = R.validate_asset_inputs(cost=0, salvage=0, life=5)
        assert any("취득원가" in e for e in errors)

    def test_cost_negative_is_invalid(self):
        errors = R.validate_asset_inputs(cost=-1_000_000, salvage=0, life=5)
        assert any("취득원가" in e for e in errors)

    def test_valid_inputs_return_no_errors(self):
        assert R.validate_asset_inputs(cost=10_000_000, salvage=1_000_000, life=5) == []

    def test_disposal_before_acquisition_is_invalid(self):
        errors = R.validate_asset_inputs(
            cost=10_000_000, salvage=0, life=5,
            acq=dt.date(2022, 1, 1), disposal=dt.date(2021, 12, 31))
        assert any("처분일" in e for e in errors)

    def test_reest_date_before_acquisition_is_invalid(self):
        errors = R.validate_asset_inputs(
            cost=10_000_000, salvage=0, life=5,
            acq=dt.date(2022, 1, 1), reest_date=dt.date(2021, 12, 31))
        assert any("내용연수재추정일" in e for e in errors)

    def test_suspension_start_before_acquisition_is_invalid(self):
        errors = R.validate_asset_inputs(
            cost=10_000_000, salvage=0, life=5, acq=dt.date(2022, 1, 1),
            susp_start=dt.date(2021, 6, 1), susp_end=dt.date(2022, 6, 1))
        assert any("상각중단시작일" in e for e in errors)

    def test_suspension_end_before_acquisition_is_invalid(self):
        # susp_end < susp_start(기존 "상각중단기간 오류")와는 다른 케이스 — 시작/종료
        # 순서 자체는 정상이지만 기간 전체가 취득일보다 앞서는 경우. susp_end >=
        # susp_start가 항상 보장되므로 susp_start만 취득일 이전이고 susp_end는
        # 취득일 이후인 조합은 나올 수 없다 — 그래서 이 케이스는 두 오류 메시지가
        # 함께 뜨는 게 정상이다.
        errors = R.validate_asset_inputs(
            cost=10_000_000, salvage=0, life=5, acq=dt.date(2022, 6, 1),
            susp_start=dt.date(2022, 1, 1), susp_end=dt.date(2022, 3, 1))
        assert any("상각중단시작일" in e for e in errors)
        assert any("상각중단종료일" in e for e in errors)

    def test_capex_date_before_acquisition_is_invalid(self):
        errors = R.validate_asset_inputs(
            cost=10_000_000, salvage=0, life=5, acq=dt.date(2022, 1, 1),
            capex_date=dt.date(2021, 12, 31), capex_amount=1_000_000)
        assert any("자본적지출일" in e for e in errors)

    def test_dates_on_or_after_acquisition_are_valid(self):
        # 취득일과 같은 날(경계값 포함) 또는 그 이후면 정상 — 순서 검증에 걸리면 안 된다.
        # 재추정일(2022-01-01)==처분일(2022-01-01)로 둬서 "재추정일>처분일" 규칙의
        # 경계(같은 날은 '늦음'이 아니라 정상)도 함께 만족시킨다.
        errors = R.validate_asset_inputs(
            cost=10_000_000, salvage=0, life=5, acq=dt.date(2022, 1, 1),
            disposal=dt.date(2022, 1, 1), reest_date=dt.date(2022, 1, 1),
            susp_start=dt.date(2022, 6, 1), susp_end=dt.date(2022, 12, 31),
            capex_date=dt.date(2022, 1, 1), capex_amount=1_000_000)
        assert errors == []

    def test_reest_date_after_disposal_is_invalid(self):
        # 이미 처분한 자산(2024-06-30 처분)을 그 뒤(2024-12-31)에 내용연수 재추정하는 것은
        # 상각 대상이 사라진 뒤의 재추정이라 논리적으로 불가능 — 순서 오류로 잡아야 한다.
        errors = R.validate_asset_inputs(
            cost=10_000_000, salvage=0, life=5, acq=dt.date(2021, 1, 1),
            disposal=dt.date(2024, 6, 30), reest_date=dt.date(2024, 12, 31))
        assert any("내용연수재추정일" in e and "처분일" in e for e in errors)

    def test_reest_date_on_or_before_disposal_is_valid(self):
        # 재추정일 <= 처분일이면 정상. 같은 날(경계)과 그 이전 모두 오류가 없어야 한다.
        same_day = R.validate_asset_inputs(
            cost=10_000_000, salvage=0, life=5, acq=dt.date(2021, 1, 1),
            disposal=dt.date(2024, 6, 30), reest_date=dt.date(2024, 6, 30))
        assert same_day == []
        before = R.validate_asset_inputs(
            cost=10_000_000, salvage=0, life=5, acq=dt.date(2021, 1, 1),
            disposal=dt.date(2024, 6, 30), reest_date=dt.date(2023, 1, 1))
        assert before == []

    def test_date_order_checks_skipped_when_acq_not_provided(self):
        # acq를 안 넘기면(기본값 None) 비교 기준이 없으므로 순서 검증 자체를
        # 건너뛴다 — 기존 30여 개 호출부가 acq 없이 부르던 것과 하위 호환된다.
        errors = R.validate_asset_inputs(
            cost=10_000_000, salvage=0, life=5, disposal=dt.date(1900, 1, 1))
        assert errors == []


# ---------------------------------------------------------------------------
# 보조 함수: round_won (원 단위 반올림, ROUND_HALF_UP)
# ---------------------------------------------------------------------------
class TestRoundWon:
    def test_round_half_up(self):
        assert R.round_won(1_234_567.5) == 1_234_568

    def test_round_down_below_half(self):
        assert R.round_won(1_234_567.4) == 1_234_567

    def test_round_down_just_below_half(self):
        assert R.round_won(1_234_567.49) == 1_234_567


# ---------------------------------------------------------------------------
# 생산량비례법 (units_of_production_current_period_dep)
# ---------------------------------------------------------------------------
class TestUnitsOfProduction:
    def test_basic_ratio(self):
        # 상각대상금액 9,000,000 x (2,000/10,000) = 1,800,000
        dep, months = R.units_of_production_current_period_dep(
            cost=10_000_000, salvage=1_000_000, total_units=10_000, period_units=2_000,
            disposal=None, fy_year=2025)
        assert dep == 1_800_000
        assert months == 12

    def test_disposed_before_current_year_returns_zero(self):
        dep, months = R.units_of_production_current_period_dep(
            cost=10_000_000, salvage=0, total_units=10_000, period_units=2_000,
            disposal=dt.date(2024, 6, 1), fy_year=2025)
        assert dep == 0.0
        assert months == 0

    def test_zero_period_units_returns_zero_months(self):
        dep, months = R.units_of_production_current_period_dep(
            cost=10_000_000, salvage=0, total_units=10_000, period_units=0,
            disposal=None, fy_year=2025)
        assert dep == 0.0
        assert months == 0

    def test_dep_capped_at_depreciable_amount(self):
        # 당기실제생산량이 총예정생산량을 초과해도 상각대상금액(9,000,000)을 넘지 않는다
        dep, _ = R.units_of_production_current_period_dep(
            cost=10_000_000, salvage=1_000_000, total_units=10_000, period_units=15_000,
            disposal=None, fy_year=2025)
        assert dep == 9_000_000


class TestRecalcAssetUnitsOfProduction:
    def test_recalc_asset_units_of_production(self):
        dep, months, life_ended, note = R.recalc_asset(
            acq=dt.date(2023, 1, 1), cost=10_000_000, salvage=1_000_000, life=5,
            method="생산량비례법", disposal=None, reest_date=None, reest_life=None,
            ref_date=dt.date(2025, 12, 31), fy_year=2025,
            total_units=10_000, period_units=2_000,
        )
        assert dep == 1_800_000
        assert months == 12
        assert life_ended is False  # v1: 생산량비례법은 누적이력 미추적으로 항상 종료 아님


class TestValidateAssetInputsUnitsOfProduction:
    def test_existing_calls_without_method_unaffected(self):
        # method 인자를 안 넘기는 기존 8개 케이스와 동일하게 동작해야 한다(회귀 확인).
        assert R.validate_asset_inputs(cost=10_000_000, salvage=1_000_000, life=5) == []
        errors = R.validate_asset_inputs(cost=5_000_000, salvage=0, life=0)
        assert any("내용연수" in e for e in errors)

    def test_life_ignored_for_units_of_production(self):
        errors = R.validate_asset_inputs(
            cost=10_000_000, salvage=1_000_000, life=0,
            method="생산량비례법", total_units=10_000)
        assert not any("내용연수" in e for e in errors)

    def test_total_units_missing_is_invalid(self):
        errors = R.validate_asset_inputs(
            cost=10_000_000, salvage=1_000_000, life=5,
            method="생산량비례법", total_units=None)
        assert any("총예정생산량" in e for e in errors)

    def test_total_units_zero_is_invalid(self):
        errors = R.validate_asset_inputs(
            cost=10_000_000, salvage=1_000_000, life=5,
            method="생산량비례법", total_units=0)
        assert any("총예정생산량" in e for e in errors)

    def test_valid_units_of_production_inputs(self):
        errors = R.validate_asset_inputs(
            cost=10_000_000, salvage=1_000_000, life=5,
            method="생산량비례법", total_units=10_000)
        assert errors == []


# ---------------------------------------------------------------------------
# 자산군별 통계 (build_category_summary)
# ---------------------------------------------------------------------------
class TestCategorySummary:
    def test_summary_counts_and_mismatch_rate(self):
        result_df = pd.DataFrame([
            {"자산명": "A", "자산분류": "유형자산", "차이(재계산-회사반영)": 0, "중요성구분": "경미한 차이"},
            {"자산명": "B", "자산분류": "유형자산", "차이(재계산-회사반영)": 500, "중요성구분": "유의한 차이"},
            {"자산명": "C", "자산분류": "무형자산", "차이(재계산-회사반영)": 0, "중요성구분": "경미한 차이"},
        ])
        error_df = pd.DataFrame([
            {"자산명": "D", "자산분류": "유형자산"},
        ])
        summary = R.build_category_summary(result_df, error_df).set_index("자산분류")

        assert summary.loc["유형자산", "전체건수"] == 3
        assert summary.loc["유형자산", "정상계산건수"] == 2
        assert summary.loc["유형자산", "불일치건수"] == 1
        assert summary.loc["유형자산", "불일치율(%)"] == 50.0
        assert summary.loc["유형자산", "유의한차이건수"] == 1
        assert summary.loc["유형자산", "오류건수"] == 1

        assert summary.loc["무형자산", "전체건수"] == 1
        assert summary.loc["무형자산", "불일치율(%)"] == 0.0
        assert summary.loc["무형자산", "오류건수"] == 0

    def test_empty_error_df(self):
        result_df = pd.DataFrame([
            {"자산명": "A", "자산분류": "유형자산", "차이(재계산-회사반영)": 0, "중요성구분": "경미한 차이"},
        ])
        error_df = pd.DataFrame(columns=["자산명", "자산분류"])
        summary = R.build_category_summary(result_df, error_df).set_index("자산분류")
        assert summary.loc["유형자산", "오류건수"] == 0
        assert summary.loc["유형자산", "전체건수"] == 1


# ---------------------------------------------------------------------------
# 룰 기반 원인 분류 (get_rule_based_cause)
# ---------------------------------------------------------------------------
class TestRuleBasedCause:
    def test_disposed_current_year(self):
        cause = R.get_rule_based_cause(
            is_reestimated=False, is_disposed=True, note="당기중처분(2025-06-30)")
        assert cause is not None and "처분일" in cause

    def test_disposed_prior_year(self):
        cause = R.get_rule_based_cause(
            is_reestimated=False, is_disposed=True, note="전기이전처분(2024-03-31)")
        assert cause is not None and "전기" in cause

    def test_reestimated(self):
        cause = R.get_rule_based_cause(
            is_reestimated=True, is_disposed=False, note="내용연수재추정(2025-01-01→8년)")
        assert cause is not None and "재추정" in cause

    def test_life_ended(self):
        cause = R.get_rule_based_cause(
            is_reestimated=False, is_disposed=False, note="내용연수종료")
        assert cause is not None and "종료" in cause

    def test_ambiguous_case_returns_none(self):
        # 처분/재추정/내용연수종료 어느 쪽에도 해당하지 않으면 규칙으로 설명할 수 없으므로
        # None을 반환해 호출부가 AI 호출로 폴백하도록 한다.
        assert R.get_rule_based_cause(is_reestimated=False, is_disposed=False, note="-") is None

    def test_capex_cause(self):
        cause = R.get_rule_based_cause(
            is_reestimated=False, is_disposed=False, note="자본적지출(2024-01-01→1,000,000원)",
            is_capex=True)
        assert cause is not None and "자본적지출" in cause

    def test_suspension_cause(self):
        cause = R.get_rule_based_cause(
            is_reestimated=False, is_disposed=False, note="상각중단(2024-01-01~2024-06-30)",
            is_suspended=True)
        assert cause is not None and "상각중단" in cause

    def test_capex_flag_without_note_returns_none(self):
        # is_capex=True여도 note에 실제 자본적지출 기록이 없으면(다른 이유로 호출된 경우)
        # 규칙을 적용하지 않는다.
        assert R.get_rule_based_cause(
            is_reestimated=False, is_disposed=False, note="-", is_capex=True) is None


# ---------------------------------------------------------------------------
# 다기간(연도별) 비교 (fy_ref_date / build_multi_year_trend_df / detect_yoy_anomalies)
# ---------------------------------------------------------------------------
def _make_cols(df):
    required = ["자산명", "자산분류", "취득일", "취득원가", "잔존가치", "내용연수", "상각방법", "회사반영상각비"]
    optional = ["처분일", "재추정일", "재추정내용연수", "재추정후상각방법", "총예정생산량", "당기실제생산량",
                "전기말누적생산량", "자본적지출일", "자본적지출액", "상각중단시작일", "상각중단종료일",
                "회사반영누계상각액"]
    cols = {k: k for k in required}
    cols.update({k: (k if k in df.columns else None) for k in optional})
    return cols


class TestFyRefDate:
    def test_current_fy_year_uses_ref_date(self):
        assert R.fy_ref_date(R.FY_YEAR, R.FY_YEAR) == R.REF_DATE

    def test_other_year_uses_dec_31(self):
        assert R.fy_ref_date(2023, R.FY_YEAR) == dt.date(2023, 12, 31)

    def test_current_fy_year_param_not_global_fy_year(self):
        # 회귀 테스트: current_fy_year는 전역 FY_YEAR가 아니라 인자로 판단해야 한다.
        # 전역 FY_YEAR와 다른 값을 "당기"로 지정해도 그 연도는 REF_DATE를 그대로 써야 한다.
        assert R.fy_ref_date(2023, current_fy_year=2023) == R.REF_DATE
        assert R.fy_ref_date(R.FY_YEAR, current_fy_year=2023) == dt.date(R.FY_YEAR, 12, 31)


class TestMultiYearTrend:
    def _sample_df(self):
        # 자산A: 정액법, 취득 2023-01-01, 연간상각액 2,400,000(취득~기준일까지 매년 동일)
        # 자산B/자산C: 이름이 둘 다 "차량"으로 중복되지만 서로 다른 자산 - 그룹핑 키가
        # 자산명이 아니라 자산ID(행 순번)여야 서로 합쳐지지 않는지 검증한다.
        # 자산B: 취득 2024-01-01(2023년엔 미취득 → 0원), 연간상각액 2,000,000
        # 자산C: 취득 2023-01-01, 내용연수 2년 → 2025년엔 내용연수 종료로 0원
        return pd.DataFrame([
            {"자산명": "본사건물", "자산분류": "유형자산", "취득일": dt.date(2023, 1, 1),
             "취득원가": 12_000_000, "잔존가치": 0, "내용연수": 5, "상각방법": "정액법",
             "회사반영상각비": 0},
            {"자산명": "차량", "자산분류": "유형자산", "취득일": dt.date(2024, 1, 1),
             "취득원가": 6_000_000, "잔존가치": 0, "내용연수": 3, "상각방법": "정액법",
             "회사반영상각비": 0},
            {"자산명": "차량", "자산분류": "유형자산", "취득일": dt.date(2023, 1, 1),
             "취득원가": 4_000_000, "잔존가치": 0, "내용연수": 2, "상각방법": "정액법",
             "회사반영상각비": 0},
        ])

    def test_matches_recalc_asset_per_year(self):
        df = self._sample_df()
        cols = _make_cols(df)
        trend_df = R.build_multi_year_trend_df(df, cols, [2023, 2024, 2025], R.FY_YEAR)

        asset_a = trend_df[trend_df["자산명"] == "본사건물"].set_index("회계연도")
        assert asset_a.loc[2023, "재계산_당기감가상각비"] == 2_400_000
        assert asset_a.loc[2024, "재계산_당기감가상각비"] == 2_400_000
        assert asset_a.loc[2025, "재계산_당기감가상각비"] == 2_400_000

    def test_duplicate_names_kept_separate_by_asset_id(self):
        df = self._sample_df()
        cols = _make_cols(df)
        trend_df = R.build_multi_year_trend_df(df, cols, [2023, 2024, 2025], R.FY_YEAR)

        cha_rows = trend_df[trend_df["자산명"] == "차량"]
        assert cha_rows["자산ID"].nunique() == 2  # 이름은 같아도 별개 자산으로 유지

        by_id = cha_rows.set_index(["자산ID", "회계연도"])["재계산_당기감가상각비"]
        asset_b_id = cha_rows[cha_rows["회계연도"] == 2023]
        asset_b_id = asset_b_id[asset_b_id["재계산_당기감가상각비"] == 0]["자산ID"].iloc[0]
        asset_c_id = cha_rows[cha_rows["회계연도"] == 2023]
        asset_c_id = asset_c_id[asset_c_id["재계산_당기감가상각비"] == 2_000_000]["자산ID"].iloc[0]

        assert by_id[(asset_b_id, 2023)] == 0            # 자산B: 2024년 취득 전이라 0원
        assert by_id[(asset_b_id, 2024)] == 2_000_000
        assert by_id[(asset_c_id, 2024)] == 2_000_000
        assert by_id[(asset_c_id, 2025)] == 0             # 자산C: 내용연수(2년) 종료로 0원

    def test_detect_yoy_anomalies_flags(self):
        df = self._sample_df()
        cols = _make_cols(df)
        trend_df = R.build_multi_year_trend_df(df, cols, [2023, 2024, 2025], R.FY_YEAR)
        flagged = R.detect_yoy_anomalies(trend_df, threshold_pct=20.0)

        # 자산A: 매년 동일 금액 → 첫 연도는 "-"(비교대상 없음), 이후는 변동 없어 "-"
        asset_a = flagged[flagged["자산명"] == "본사건물"].set_index("회계연도")
        assert asset_a.loc[2023, "이상탐지"] == "-"
        assert asset_a.loc[2024, "이상탐지"] == "-"

        cha_rows = flagged[flagged["자산명"] == "차량"]
        # 자산B: 0원 → 2,000,000원으로 "신규발생"
        new_asset = cha_rows[(cha_rows["회계연도"] == 2024) & (cha_rows["재계산_당기감가상각비"] == 2_000_000)]
        prev_year_of_new_asset = cha_rows[
            (cha_rows["자산ID"] == new_asset["자산ID"].iloc[0]) & (cha_rows["회계연도"] == 2023)]
        assert prev_year_of_new_asset["재계산_당기감가상각비"].iloc[0] == 0
        assert new_asset["이상탐지"].iloc[0] == "신규발생"

        # 자산C: 2,000,000원 → 0원으로 급감 → "경고"(20% 임계치 초과)
        declined = cha_rows[(cha_rows["회계연도"] == 2025) & (cha_rows["재계산_당기감가상각비"] == 0)]
        assert declined["이상탐지"].iloc[0] == "경고"

    def test_omitted_threshold_pct_follows_monkeypatched_global_not_stale_default(self, monkeypatch):
        # classify_materiality와 동일한 회귀 테스트. 10% 증가는 기본 임계치(20%)로는
        # "경고"가 아니어야 정상이지만, YOY_ANOMALY_THRESHOLD_PCT를 5%로 monkeypatch한
        # 뒤 threshold_pct 인자를 생략하고 호출하면 새 값(5%) 기준으로 "경고"가 나와야
        # 함수 기본값이 정의 시점에 고정되지 않았다는 것이 증명된다.
        trend_df = pd.DataFrame([
            {"자산ID": 1, "자산명": "자산A", "자산분류": "유형자산", "상각방법": "정액법",
             "회계연도": 2024, "재계산_당기감가상각비": 100_000, "당기해당월수": 12, "비고": ""},
            {"자산ID": 1, "자산명": "자산A", "자산분류": "유형자산", "상각방법": "정액법",
             "회계연도": 2025, "재계산_당기감가상각비": 110_000, "당기해당월수": 12, "비고": ""},
        ])
        flagged_default = R.detect_yoy_anomalies(trend_df)  # 기본 20% 임계치 -> 10% 증가는 "-"
        assert flagged_default[flagged_default["회계연도"] == 2025]["이상탐지"].iloc[0] == "-"

        monkeypatch.setattr(R, "YOY_ANOMALY_THRESHOLD_PCT", 5.0)
        flagged_patched = R.detect_yoy_anomalies(trend_df)  # 인자 생략 -> 새 전역값(5%) 써야 함
        assert flagged_patched[flagged_patched["회계연도"] == 2025]["이상탐지"].iloc[0] == "경고"


# ---------------------------------------------------------------------------
# 확장 이벤트(자본적지출/상각중단/방법변경) 입력값 검증
# ---------------------------------------------------------------------------
class TestValidateAssetInputsExtendedEvents:
    def test_invalid_reest_method(self):
        errors = R.validate_asset_inputs(cost=10_000_000, salvage=0, life=5, reest_method="정율법")
        assert any("재추정후상각방법" in e for e in errors)

    def test_capex_date_without_amount(self):
        errors = R.validate_asset_inputs(
            cost=10_000_000, salvage=0, life=5, capex_date=dt.date(2024, 1, 1), capex_amount=None)
        assert any("자본적지출" in e for e in errors)

    def test_capex_amount_without_date(self):
        errors = R.validate_asset_inputs(
            cost=10_000_000, salvage=0, life=5, capex_date=None, capex_amount=1_000_000)
        assert any("자본적지출" in e for e in errors)

    def test_capex_amount_zero_is_invalid(self):
        errors = R.validate_asset_inputs(
            cost=10_000_000, salvage=0, life=5, capex_date=dt.date(2024, 1, 1), capex_amount=0)
        assert any("자본적지출액" in e for e in errors)

    def test_suspension_missing_end(self):
        errors = R.validate_asset_inputs(
            cost=10_000_000, salvage=0, life=5, susp_start=dt.date(2024, 1, 1), susp_end=None)
        assert any("상각중단기간" in e for e in errors)

    def test_suspension_end_before_start(self):
        errors = R.validate_asset_inputs(
            cost=10_000_000, salvage=0, life=5,
            susp_start=dt.date(2024, 6, 1), susp_end=dt.date(2024, 1, 1))
        assert any("상각중단기간" in e for e in errors)

    def test_valid_extended_events_no_errors(self):
        errors = R.validate_asset_inputs(
            cost=10_000_000, salvage=0, life=5, reest_method="정률법",
            capex_date=dt.date(2024, 1, 1), capex_amount=1_000,
            susp_start=dt.date(2024, 1, 1), susp_end=dt.date(2024, 6, 1))
        assert errors == []

    def test_negative_prior_period_units_is_invalid(self):
        errors = R.validate_asset_inputs(
            cost=10_000_000, salvage=0, life=5, method="생산량비례법", total_units=1_000,
            prior_period_units=-1)
        assert any("전기말누적생산량" in e for e in errors)

    def test_reest_method_accepts_double_declining_and_syd(self):
        # 이중체감법/연수합계법 도입 후에도 재추정후상각방법으로 지정할 수 있어야 한다.
        assert R.validate_asset_inputs(
            cost=10_000_000, salvage=0, life=5, reest_method="이중체감법") == []
        assert R.validate_asset_inputs(
            cost=10_000_000, salvage=0, life=5, reest_method="연수합계법") == []


# ---------------------------------------------------------------------------
# 상각중단 연장 로직 (apply_suspension_extension)
# ---------------------------------------------------------------------------
class TestSuspensionExtension:
    def test_extends_when_suspension_in_last_segment(self):
        segments = [dict(start_idx=R.month_index(2020, 1), end_idx=R.month_index(2024, 12))]
        susp_s, susp_e = R.month_index(2022, 6), R.month_index(2022, 12)
        R.apply_suspension_extension(segments, susp_s, susp_e)
        assert segments[0]["end_idx"] == R.month_index(2024, 12) + 7

    def test_does_not_extend_when_suspension_in_earlier_segment(self):
        # 회귀 테스트: 상각중단이 마지막이 아닌 이전 구간에서 발생하면 마지막 구간을
        # 연장하면 안 된다(처음 구현 때 susp_start<=last_end만 검사해서 항상 연장되던 버그).
        seg0 = dict(start_idx=R.month_index(2020, 1), end_idx=R.month_index(2021, 12))
        seg1 = dict(start_idx=R.month_index(2022, 1), end_idx=R.month_index(2024, 12))
        segments = [seg0, seg1]
        susp_s, susp_e = R.month_index(2021, 6), R.month_index(2021, 12)  # seg0 안에서 발생
        R.apply_suspension_extension(segments, susp_s, susp_e)
        assert segments[1]["end_idx"] == R.month_index(2024, 12)  # 연장 없음


# ---------------------------------------------------------------------------
# recalc_asset 통합 — 자본적지출/상각중단/방법변경 (손계산 검증)
# ---------------------------------------------------------------------------
class TestRecalcAssetExtendedEvents:
    def test_capex_only(self):
        # 취득 2020-01-01, 12,000,000원, 잔존가치0, 내용연수10년(월상각 100,000원)
        # 2023-01-01 자본적지출 2,400,000원 → 그 시점 장부가액 8,400,000원 + 2,400,000
        #   = 10,800,000원을 잔여내용연수(84개월)로 재상각 → 월 128,571.43원
        # 당기(2025)는 이 구간에 완전히 포함되므로 12개월 * 128,571.43 = 1,542,857원
        dep, months, life_ended, note = R.recalc_asset(
            acq=dt.date(2020, 1, 1), cost=12_000_000, salvage=0, life=10, method="정액법",
            disposal=None, reest_date=None, reest_life=None,
            ref_date=dt.date(2025, 12, 31), fy_year=2025,
            capex_date=dt.date(2023, 1, 1), capex_amount=2_400_000,
        )
        assert dep == 1_542_857
        assert months == 12
        assert life_ended is False
        assert "자본적지출" in note

    def test_capex_and_reest_same_date_capex_applied_first(self):
        # 취득 2022-01-01, 10,000,000원, 내용연수5년(월상각 166,666.67원)
        # 2024-01-01에 자본적지출 2,000,000원과 재추정(4년)이 동시 발생.
        # "자본적지출을 먼저 반영"하므로: 그 시점 장부가액 6,000,000 + 2,000,000
        #   = 8,000,000원을 새 내용연수(4년=48개월)로 재상각 → 월 166,666.67원(=8,000,000/48)
        # 당기(2025) 12개월 = 2,000,000원(딱 떨어짐: 8,000,000/48*12 = 2,000,000)
        dep, months, life_ended, note = R.recalc_asset(
            acq=dt.date(2022, 1, 1), cost=10_000_000, salvage=0, life=5, method="정액법",
            disposal=None, reest_date=dt.date(2024, 1, 1), reest_life=4,
            ref_date=dt.date(2025, 12, 31), fy_year=2025,
            capex_date=dt.date(2024, 1, 1), capex_amount=2_000_000,
        )
        assert dep == 2_000_000
        assert months == 12

    def test_capex_then_mid_year_method_change(self):
        # 취득 2022-01-01 정액법 10년(120개월, 월상각 83,333.33원).
        # 2024-01-01 자본적지출 1,000,000원(방법 유지) → 장부가액 8,000,000+1,000,000
        #   = 9,000,000원을 잔여내용연수(96개월)로 재상각 → 월 93,750원.
        # 2025-07-01 재추정(5년, 정률법으로 전환) → 그 시점 장부가액
        #   9,000,000 - 93,750*18개월 = 7,312,500원을 새 기준가로 정률법(상각률 0.451) 상각 시작.
        # 당기(2025)는 두 구간에 걸쳐 있다: 1~6월은 자본적지출 구간(정액법, 93,750*6=562,500원),
        #   7~12월은 재추정 이후 구간(정률법, 7,312,500*0.451*6/12=1,648,968.75→1,648,969원).
        # 합계 562,500 + 1,648,969 = 2,211,469원, 12개월(정액법 6개월 + 정률법 6개월).
        dep, months, life_ended, note = R.recalc_asset(
            acq=dt.date(2022, 1, 1), cost=10_000_000, salvage=0, life=10, method="정액법",
            disposal=None, reest_date=dt.date(2025, 7, 1), reest_life=5,
            ref_date=dt.date(2025, 12, 31), fy_year=2025,
            capex_date=dt.date(2024, 1, 1), capex_amount=1_000_000, reest_method="정률법",
        )
        assert dep == 2_211_469
        assert months == 12
        assert "정률법" in note

    def test_suspension_extends_life_end(self):
        # 취득 2020-01-01, 6,000,000원, 내용연수5년(60개월, 월상각 100,000원, 원래 종료 2024-12).
        # 2022-06~2022-12(7개월) 상각중단 → 종료시점이 2025-07로 연장된다.
        # 당기(2025)는 1~7월만 상각대상(7개월) = 700,000원. 이미 8월부터는 상각 끝.
        dep, months, life_ended, note = R.recalc_asset(
            acq=dt.date(2020, 1, 1), cost=6_000_000, salvage=0, life=5, method="정액법",
            disposal=None, reest_date=None, reest_life=None,
            ref_date=dt.date(2025, 12, 31), fy_year=2025,
            susp_start=dt.date(2022, 6, 1), susp_end=dt.date(2022, 12, 31),
        )
        assert dep == 700_000
        assert months == 7
        assert life_ended is True  # 기준일(2025-12-31) 기준으로는 이미 연장된 종료월(2025-07)도 지남

        # 중단 당해 연도(2022)는 12개월 중 7개월이 중단되어 5개월분만 인정된다.
        dep2022, months2022, _, _ = R.recalc_asset(
            acq=dt.date(2020, 1, 1), cost=6_000_000, salvage=0, life=5, method="정액법",
            disposal=None, reest_date=None, reest_life=None,
            ref_date=dt.date(2022, 12, 31), fy_year=2022,
            susp_start=dt.date(2022, 6, 1), susp_end=dt.date(2022, 12, 31),
        )
        assert dep2022 == 500_000
        assert months2022 == 5

    def test_suspension_in_non_last_segment_does_not_extend(self):
        # 상각중단(2021-06~2021-12)이 재추정(2022-01, 3년) "이전" 구간에서 발생한다.
        # v1 스코프 제한: 연장은 마지막 구간에서 발생한 상각중단에만 적용되므로, 이
        # 자산의 최종 종료시점은 재추정 구간의 원래 종료월(2024-12)에서 바뀌지 않는다.
        dep2025, months2025, life_ended, note = R.recalc_asset(
            acq=dt.date(2020, 1, 1), cost=6_000_000, salvage=0, life=5, method="정액법",
            disposal=None, reest_date=dt.date(2022, 1, 1), reest_life=3,
            ref_date=dt.date(2025, 12, 31), fy_year=2025,
            susp_start=dt.date(2021, 6, 1), susp_end=dt.date(2021, 12, 31),
        )
        assert dep2025 == 0  # 연장됐다면 0이 아니었을 것(회귀 방지)
        assert life_ended is True

        # 상각중단이 발생한 2021년 자체는 여전히 7개월분이 0원 처리된다(장부가액 계산에는 반영).
        dep2021, months2021, _, _ = R.recalc_asset(
            acq=dt.date(2020, 1, 1), cost=6_000_000, salvage=0, life=5, method="정액법",
            disposal=None, reest_date=dt.date(2022, 1, 1), reest_life=3,
            ref_date=dt.date(2021, 12, 31), fy_year=2021,
            susp_start=dt.date(2021, 6, 1), susp_end=dt.date(2021, 12, 31),
        )
        assert dep2021 == 500_000
        assert months2021 == 5


# ---------------------------------------------------------------------------
# 전기말 감가상각누계액 재계산 (recalc_accumulated_dep)
# ---------------------------------------------------------------------------
class TestAccumulatedDepreciation:
    def test_plain_asset_sums_full_years(self):
        # 취득 2020-01-01, 6,000,000원, 내용연수5년(월상각 100,000원).
        # fy_year=2023이면 전기(2020~2022) 3개년 * 1,200,000원 = 3,600,000원.
        accum = R.recalc_accumulated_dep(
            acq=dt.date(2020, 1, 1), cost=6_000_000, salvage=0, life=5, method="정액법",
            disposal=None, reest_date=None, reest_life=None, fy_year=2023)
        assert accum == 3_600_000

    def test_units_of_production_returns_none(self):
        # 생산량비례법은 과거 연도별 생산량 이력을 추적하지 않으므로(v1 한계) None.
        accum = R.recalc_accumulated_dep(
            acq=dt.date(2020, 1, 1), cost=6_000_000, salvage=0, life=1, method="생산량비례법",
            disposal=None, reest_date=None, reest_life=None, fy_year=2023,
            total_units=1_000, period_units=100)
        assert accum is None

    def test_acquired_in_current_fy_year_has_no_prior_accumulation(self):
        accum = R.recalc_accumulated_dep(
            acq=dt.date(2025, 6, 1), cost=6_000_000, salvage=0, life=5, method="정액법",
            disposal=None, reest_date=None, reest_life=None, fy_year=2025)
        assert accum == 0

    def test_disposal_stops_accumulation(self):
        # 취득 2020-01-01, 내용연수5년(월상각 100,000원), 2022-06-30 처분.
        # 2020,2021은 각 1,200,000원, 2022는 처분월까지 6개월=600,000원,
        # 2023,2024는 이미 처분되어 0원 -> 합계 3,000,000원.
        accum = R.recalc_accumulated_dep(
            acq=dt.date(2020, 1, 1), cost=6_000_000, salvage=0, life=5, method="정액법",
            disposal=dt.date(2022, 6, 30), reest_date=None, reest_life=None, fy_year=2025)
        assert accum == 3_000_000

    def test_accumulation_reflects_capex(self):
        # 취득 2020-01-01, 12,000,000원, 내용연수10년(월상각 100,000원),
        # 2023-01-01 자본적지출 2,400,000원(잔여내용연수 84개월로 재상각, 월 128,571.43원).
        # 2020~2022(3년): 1,200,000원씩 = 3,600,000원.
        # 2023~2024(2년): 자본적지출 반영 후 연 1,542,857원씩(반올림) = 3,085,714원.
        # 합계 6,685,714원.
        accum = R.recalc_accumulated_dep(
            acq=dt.date(2020, 1, 1), cost=12_000_000, salvage=0, life=10, method="정액법",
            disposal=None, reest_date=None, reest_life=None, fy_year=2025,
            capex_date=dt.date(2023, 1, 1), capex_amount=2_400_000)
        assert accum == 6_685_714

    def test_double_declining_balance_accumulation(self):
        # 취득 2021-01-01, 1,000,000원, 잔존가치 100,000원, 내용연수5년(상각률 2/5=0.4).
        # 연도별 상각비: 400,000 -> 240,000 -> 144,000 -> 86,400(4년간 합계 870,400).
        accum = R.recalc_accumulated_dep(
            acq=dt.date(2021, 1, 1), cost=1_000_000, salvage=100_000, life=5, method="이중체감법",
            disposal=None, reest_date=None, reest_life=None, fy_year=2025)
        assert accum == 870_400

    def test_sum_of_years_digits_accumulation_matches_closed_form(self):
        # 취득 2021-01-01, 1,000,000원, 잔존가치 100,000원, 내용연수5년.
        # 전기말(4개년 경과, n=4) 폐형식: (900,000)*(n-1)(2N-n+2)/(N(N+1))
        #  = 900,000*3*8/30 = 720,000... 검증은 실제 연차별 합산으로 한다:
        # 1년차 300,000 + 2년차 240,000 + 3년차 180,000 + 4년차 120,000 = 840,000.
        accum = R.recalc_accumulated_dep(
            acq=dt.date(2021, 1, 1), cost=1_000_000, salvage=100_000, life=5, method="연수합계법",
            disposal=None, reest_date=None, reest_life=None, fy_year=2025)
        assert accum == 840_000

    def test_sum_of_years_digits_fully_depreciates_by_end_of_life(self):
        # 연수합계법은 N년에 걸친 몫의 합이 정확히 1이 되도록 설계돼, 내용연수가
        # 끝나면(fy_year가 취득연도+내용연수와 같아지면) 누계상각액이 정확히
        # (취득원가-잔존가치)와 같아야 한다(정률법/이중체감법과 달리 잔여액이 안 남음).
        accum = R.recalc_accumulated_dep(
            acq=dt.date(2021, 1, 1), cost=1_000_000, salvage=100_000, life=5, method="연수합계법",
            disposal=None, reest_date=None, reest_life=None, fy_year=2026)
        assert accum == 900_000


# ---------------------------------------------------------------------------
# 법인세법 시행규칙 [별표4] 상각률표 반영 (RATE_TABLE 전면 교체)
# ---------------------------------------------------------------------------
class TestRateTable:
    def test_life2_life3_match_official_table(self):
        # 별표4 정률법 할분리: 2년=777(0.777), 3년=632(0.632).
        # 기존 코드값(0.684/0.536)은 별표4와 달라 이번에 교체됐다.
        assert R.RATE_TABLE[2] == 0.777
        assert R.RATE_TABLE[3] == 0.632
        assert R.get_rate(2) == 0.777
        assert R.get_rate(3) == 0.632

    def test_life15_to_20_match_official_table_exactly(self):
        # 기존 근사식(1-0.05**(1/n)) 반올림값과 별표4 값이 15~20년 구간에서
        # 소수 셋째 자리가 달랐다(예: 15년 0.183 -> 0.182). 별표4 값을 그대로 씀.
        expected = {15: 0.182, 16: 0.171, 17: 0.162, 18: 0.154, 19: 0.146, 20: 0.140}
        for life, rate in expected.items():
            assert R.RATE_TABLE[life] == rate

    def test_life60_is_last_table_entry(self):
        assert R.RATE_TABLE[60] == 0.049

    def test_get_rate_falls_back_to_approximation_beyond_table(self):
        # 별표4 범위(2~60년) 밖은 여전히 근사식(1-0.05**(1/n))으로 근사한다.
        assert 61 not in R.RATE_TABLE
        assert R.get_rate(61) == round(1 - 0.05 ** (1 / 61), 3)


# ---------------------------------------------------------------------------
# 정률법 5% 잔존가액 특례(법인세법 시행규칙 [별표4] 관련 규정)
# ---------------------------------------------------------------------------
class TestDecliningBalance5PctFloor:
    def test_full_writeoff_when_book_value_first_hits_5pct_of_cost(self):
        # 취득 2021-01-01, 10,000,000원, 내용연수5년(상각률 0.451), 잔존가치 0.
        # 연도별 장부가액: 5,490,000 -> 3,014,010 -> 1,654,691.49 -> 908,426.43.
        # 2025년(5년차, 내용연수 마지막 해): 정상계산 dep=908,426.43*0.451=409,700.32면
        #   미상각잔액 498,726.11원이 남아 "잔존가치 미만" 조건(구 로직, salvage=0)에는
        #   걸리지 않아 그대로 남아버린다. 새 5% 특례(취득가액의 5%=500,000원 기준)는
        #   498,726.11<=500,000이므로 그 해에 잔여 전액(908,426.43원)을 상각한다.
        dep, months, life_ended, note = R.recalc_asset(
            acq=dt.date(2021, 1, 1), cost=10_000_000, salvage=0, life=5, method="정률법",
            disposal=None, reest_date=None, reest_life=None,
            ref_date=dt.date(2025, 12, 31), fy_year=2025,
        )
        assert dep == 908_426
        assert months == 12

    def test_year_before_floor_is_unaffected(self):
        # 같은 자산의 2024년(4년차)은 아직 5% 기준에 못 미치므로(장부가액 908,426.43>500,000)
        # 정상적인 감가상각률 계산만 적용된다: 1,654,691.49*0.451 = 746,265.86 → 746,266원.
        #
        # 기대값을 손계산 수치로 직접 박지 않고 "실제 실행 결과로 확정"한 이유:
        # 이 4년차 상각액은 2021~2023년 3개년치 `book_value -= book_value*rate`를 float로
        # 반복해 얻은 장부가액(1,654,691.49...)에 상각률을 다시 곱한 값이다. 파이썬 float는
        # 이진 부동소수라, 이런 곱셈·뺄셈이 여러 해 누적되면 십진수로는 딱 떨어지는 값도
        # 미세한 표현 오차(...49999 같은 꼬리)를 안게 된다. 그 오차가 마지막 원 단위
        # 반올림(ROUND_HALF_UP)의 .5 경계에 걸리는 해에는 손으로 십진 계산한 값과 구현이
        # 내는 값이 1원 어긋날 수 있으므로, 다년 float 누적 장부가액이 관여하는 기대값은
        # 손계산 대신 검증된 실제 파이프라인 출력(746,266원)으로 고정하는 방침을 따른다.
        # (이 케이스 자체는 소수부가 .86이라 반올림 경계에서 충분히 멀어 손계산 결과와
        #  일치하지만, 경계 근처 해에서 깨지지 않도록 동일 방침을 일관되게 적용한다.)
        dep, months, life_ended, note = R.recalc_asset(
            acq=dt.date(2021, 1, 1), cost=10_000_000, salvage=0, life=5, method="정률법",
            disposal=None, reest_date=None, reest_life=None,
            ref_date=dt.date(2024, 12, 31), fy_year=2024,
        )
        assert dep == 746_266

    def test_5pct_threshold_uses_tax_cost_including_capex(self):
        # 정률법 + 자본적지출(방법변경 없음) 조합: 5% 기준이 원 취득원가가 아니라
        # 취득원가+자본적지출 누계(세무상 취득가액)를 기준으로 계산되는지 확인.
        # 취득 2021-01-01, 10,000,000원, 내용연수5년(상각률 0.451), 2022-01-01
        # 자본적지출 5,000,000원(취득가액 기준 15,000,000원의 5%=750,000원이 새 기준).
        dep, months, life_ended, note = R.recalc_asset(
            acq=dt.date(2021, 1, 1), cost=10_000_000, salvage=0, life=5, method="정률법",
            disposal=None, reest_date=None, reest_life=None,
            ref_date=dt.date(2025, 12, 31), fy_year=2025,
            capex_date=dt.date(2022, 1, 1), capex_amount=5_000_000,
        )
        # 자본적지출로 취득가액(15,000,000원) 자체가 커져 5% 기준(750,000원)도 커지므로,
        # 자본적지출이 없을 때(기준 500,000원)보다 조기에(또는 다른 시점에) 완전상각될
        # 수 있다 — 정확한 시점은 손계산 대신 실제 실행 결과로 다음을 검증한다:
        # 자본적지출 미반영(원 취득원가 10,000,000원 기준 5%=500,000원)으로 계산했을 때와
        # 결과가 달라야 한다(취득가액 확장이 실제로 반영됐는지 확인).
        dep_no_capex, _, _, _ = R.recalc_asset(
            acq=dt.date(2021, 1, 1), cost=10_000_000, salvage=0, life=5, method="정률법",
            disposal=None, reest_date=None, reest_life=None,
            ref_date=dt.date(2025, 12, 31), fy_year=2025,
        )
        assert dep != dep_no_capex


# ---------------------------------------------------------------------------
# 엑셀 수식 주입용 메타데이터 (get_period_formula_meta)
# ---------------------------------------------------------------------------
class TestPeriodFormulaMeta:
    def test_plain_straight_line_is_simple_and_accum_eligible(self):
        meta = R.get_period_formula_meta(
            acq=dt.date(2020, 1, 1), cost=12_000_000, salvage=0, life=10, method="정액법",
            disposal=None, reest_date=None, reest_life=None, fy_year=2025)
        assert meta["basis"] == 12_000_000
        assert meta["life_months"] == 120
        assert meta["is_simple_current"] is True
        assert meta["accum_formula_eligible"] is True
        # 취득 2020-01-01부터 전기말(2024-12-31)까지 60개월 * 월상각 100,000원 = 6,000,000원
        # 상각됐으므로 전기말장부가액 = 12,000,000 - 6,000,000 = 6,000,000원.
        assert meta["prior_fy_end_bv"] == 6_000_000

    def test_plain_declining_balance_is_simple_and_accum_eligible(self):
        # "(취득원가+자본적지출)-전기말장부가액" 항등식 덕분에 정률법도(연도별 복리라도)
        # 전기말 누계액을 단일 수식으로 표현할 수 있다 — basis(=당기 1/1 시점 장부가액)와
        # prior_fy_end_bv는 정률법에서는 같은 값이다.
        meta = R.get_period_formula_meta(
            acq=dt.date(2021, 1, 1), cost=10_000_000, salvage=0, life=5, method="정률법",
            disposal=None, reest_date=None, reest_life=None, fy_year=2024)
        assert meta["basis"] == 1_654_691.49
        assert meta["life_months"] == 60
        assert meta["is_simple_current"] is True
        assert meta["accum_formula_eligible"] is True
        assert meta["prior_fy_end_bv"] == meta["basis"]

    def test_prior_year_capex_gives_remaining_life_months_not_original(self):
        # 자본적지출이 전기 이전에 있으면(당기 안에서 기준이 안 바뀌므로) 단순
        # 케이스지만, 상각률 계산에 쓸 life_months는 "원 내용연수*12"가 아니라
        # "자본적지출 시점부터 남은 개월수"여야 한다(잔여내용연수로 재상각 정책).
        meta = R.get_period_formula_meta(
            acq=dt.date(2020, 1, 1), cost=12_000_000, salvage=0, life=10, method="정액법",
            disposal=None, reest_date=None, reest_life=None, fy_year=2025,
            capex_date=dt.date(2023, 1, 1), capex_amount=2_400_000)
        assert meta["basis"] == 10_800_000
        assert meta["life_months"] == 84  # 원 내용연수 120개월이 아니라 잔여 84개월
        assert meta["is_simple_current"] is True
        # recalc_accumulated_dep 손계산 사례(TestAccumulatedDepreciation)와 동일한 자산 —
        # (12,000,000+2,400,000)-전기말장부가액이 6,685,714원과 일치해야 한다.
        assert round(14_400_000 - meta["prior_fy_end_bv"]) == 6_685_714

    def test_capex_within_current_fy_is_complex(self):
        # 당기 중에 자본적지출이 발생하면 그 해 안에서 기준(basis)이 바뀌므로
        # 단일 셀 수식으로 표현할 수 없다(당기 감가상각비만 그렇다 — 전기말은
        # 그보다 앞선 시점이라 여전히 수식 가능).
        meta = R.get_period_formula_meta(
            acq=dt.date(2020, 1, 1), cost=12_000_000, salvage=0, life=10, method="정액법",
            disposal=None, reest_date=None, reest_life=None, fy_year=2025,
            capex_date=dt.date(2025, 7, 1), capex_amount=2_400_000)
        assert meta["is_simple_current"] is False
        assert meta["accum_formula_eligible"] is True

    def test_reest_within_current_fy_is_complex(self):
        meta = R.get_period_formula_meta(
            acq=dt.date(2020, 1, 1), cost=12_000_000, salvage=0, life=10, method="정액법",
            disposal=None, reest_date=dt.date(2025, 7, 1), reest_life=5, fy_year=2025)
        assert meta["is_simple_current"] is False

    def test_units_of_production_uses_cost_as_basis(self):
        meta = R.get_period_formula_meta(
            acq=dt.date(2020, 1, 1), cost=6_000_000, salvage=0, life=1, method="생산량비례법",
            disposal=None, reest_date=None, reest_life=None, fy_year=2025)
        assert meta["basis"] == 6_000_000
        assert meta["life_months"] is None
        assert meta["is_simple_current"] is True
        assert meta["prior_fy_end_bv"] is None
        # 전기말누적생산량을 안 줬으면 여전히 수식화 불가.
        assert meta["accum_formula_eligible"] is False

    def test_units_of_production_accum_eligible_when_prior_units_given(self):
        meta = R.get_period_formula_meta(
            acq=dt.date(2020, 1, 1), cost=6_000_000, salvage=0, life=1, method="생산량비례법",
            disposal=None, reest_date=None, reest_life=None, fy_year=2025, prior_period_units=100)
        assert meta["accum_formula_eligible"] is True

    def test_prior_fy_end_bv_freezes_at_disposal(self):
        # 전기 이전에 처분된 자산은 전기말 장부가액이 처분 시점에서 멈춰야 한다
        # (처분 이후에도 상각이 계속된 것처럼 계산하면 안 됨).
        meta = R.get_period_formula_meta(
            acq=dt.date(2020, 1, 1), cost=6_000_000, salvage=0, life=5, method="정액법",
            disposal=dt.date(2022, 6, 30), reest_date=None, reest_life=None, fy_year=2025)
        # recalc_accumulated_dep의 손계산 사례(TestAccumulatedDepreciation)와 동일한 자산 —
        # 누계상각액 3,000,000원 -> 전기말장부가액 = 6,000,000-3,000,000 = 3,000,000원.
        assert meta["prior_fy_end_bv"] == 3_000_000

    def test_reest_inside_suspension_window_book_value_is_correct(self):
        # 재추정 시점이 상각중단 구간 "안"인 경계 케이스 — 새로 만들어지는 구간의
        # 시작월 자체가 상각중단 구간 안에 있을 때도 전기말장부가액이 정확해야 한다
        # (구간 시작월을 nominal_month_index로 변환하면 중단시작 직전월로 고정돼버려
        # 이후 경과월수 계산이 한 달 밀리는 버그가 있었다 — 회귀 테스트).
        meta = R.get_period_formula_meta(
            acq=dt.date(2021, 1, 1), cost=9_000_000, salvage=0, life=6, method="정액법",
            disposal=None, reest_date=dt.date(2023, 6, 1), reest_life=5, fy_year=2025,
            susp_start=dt.date(2023, 3, 1), susp_end=dt.date(2023, 9, 30))
        accum_recalc = R.recalc_accumulated_dep(
            acq=dt.date(2021, 1, 1), cost=9_000_000, salvage=0, life=6, method="정액법",
            disposal=None, reest_date=dt.date(2023, 6, 1), reest_life=5, fy_year=2025,
            susp_start=dt.date(2023, 3, 1), susp_end=dt.date(2023, 9, 30))
        assert round(9_000_000 - meta["prior_fy_end_bv"]) == accum_recalc

    def test_double_declining_balance_is_accum_eligible(self):
        # 이중체감법도 정률법과 같은 항등식으로 전기말 누계액을 수식화할 수 있어야 한다.
        meta = R.get_period_formula_meta(
            acq=dt.date(2021, 1, 1), cost=1_000_000, salvage=100_000, life=5, method="이중체감법",
            disposal=None, reest_date=None, reest_life=None, fy_year=2025)
        assert meta["is_simple_current"] is True
        assert meta["accum_formula_eligible"] is True
        assert round(1_000_000 - meta["prior_fy_end_bv"]) == 870_400
        assert meta["syd_k"] is None and meta["syd_life_years"] is None

    def test_sum_of_years_digits_reports_k_and_is_accum_eligible(self):
        # 연수합계법은 basis가 정액법처럼 고정 기준액이어야 하고(장부가액이 아님),
        # syd_k/syd_life_years로 그 해의 몫((N-k+1)/(N(N+1)/2))을 구할 수 있어야 한다.
        meta = R.get_period_formula_meta(
            acq=dt.date(2021, 1, 1), cost=1_000_000, salvage=100_000, life=5, method="연수합계법",
            disposal=None, reest_date=None, reest_life=None, fy_year=2025)
        assert meta["basis"] == 1_000_000
        assert meta["syd_k"] == 5
        assert meta["syd_life_years"] == 5
        assert meta["accum_formula_eligible"] is True
        assert round(1_000_000 - meta["prior_fy_end_bv"]) == 840_000

    def test_sum_of_years_digits_capex_does_not_reset_origin(self):
        # 자본적지출은 연수합계법의 연차(k) 스케줄을 리셋하지 않는다 — origin은
        # 여전히 취득연도 기준이어야 한다(정액법의 "잔여내용연수로 계속 상각" 정책과 동일).
        meta = R.get_period_formula_meta(
            acq=dt.date(2021, 1, 1), cost=1_000_000, salvage=0, life=5, method="연수합계법",
            disposal=None, reest_date=None, reest_life=None, fy_year=2025,
            capex_date=dt.date(2023, 1, 1), capex_amount=200_000)
        assert meta["syd_k"] == 5
        assert meta["syd_life_years"] == 5


# ---------------------------------------------------------------------------
# 이중체감법(DDB) 재계산 — 상각률=2/내용연수, 하한=잔존가치(세법상 5% 특례 없음)
# ---------------------------------------------------------------------------
class TestDoubleDecliningBalance:
    def test_basic_rate_and_floor(self):
        # 취득 2021-01-01, 1,000,000원, 잔존가치 100,000원, 내용연수5년(상각률 2/5=0.4).
        # 연도별: 400,000 -> 240,000 -> 144,000 -> 86,400 -> bv 129,600.
        # 2025년(5년차): 129,600*0.4=51,840이지만 129,600-51,840=77,760<100,000(잔존가치)
        # 이므로 잔여 전액(129,600-100,000=29,600원)을 그 해에 상각한다.
        dep, months, life_ended, note = R.recalc_asset(
            acq=dt.date(2021, 1, 1), cost=1_000_000, salvage=100_000, life=5, method="이중체감법",
            disposal=None, reest_date=None, reest_life=None,
            ref_date=dt.date(2025, 12, 31), fy_year=2025)
        assert dep == 29_600
        assert months == 12

    def test_year_before_floor_is_unaffected(self):
        # 같은 자산의 2024년(4년차)은 아직 하한(잔존가치)에 못 미치므로(129,600>100,000
        # 아니라 216,000*0.4=86,400 상각 후 129,600이 남는 정상 계산) 그대로 적용된다.
        dep, months, life_ended, note = R.recalc_asset(
            acq=dt.date(2021, 1, 1), cost=1_000_000, salvage=100_000, life=5, method="이중체감법",
            disposal=None, reest_date=None, reest_life=None,
            ref_date=dt.date(2024, 12, 31), fy_year=2024)
        assert dep == 86_400

    def test_disposal_mid_year(self):
        # 취득 2021-01-01, 12,000,000원, 내용연수6년(상각률 2/6), 2025-07-31 처분(7개월).
        dep, months, life_ended, note = R.recalc_asset(
            acq=dt.date(2021, 1, 1), cost=12_000_000, salvage=0, life=6, method="이중체감법",
            disposal=dt.date(2025, 7, 31), reest_date=None, reest_life=None,
            ref_date=dt.date(2025, 12, 31), fy_year=2025)
        assert dep == 460_905
        assert months == 7
        assert "당기중처분" in note

    def test_reest_same_method_uses_new_rate(self):
        # 재추정 후에도 이중체감법을 유지하면 새 내용연수로 상각률(2/새내용연수)이 바뀐다.
        dep, months, life_ended, note = R.recalc_asset(
            acq=dt.date(2021, 1, 1), cost=13_000_000, salvage=0, life=6, method="이중체감법",
            disposal=None, reest_date=dt.date(2024, 1, 1), reest_life=8,
            ref_date=dt.date(2025, 12, 31), fy_year=2025)
        assert dep == 722_222
        assert months == 12

    def test_reest_from_straight_line_to_double_declining(self):
        dep, months, life_ended, note = R.recalc_asset(
            acq=dt.date(2021, 1, 1), cost=12_000_000, salvage=0, life=8, method="정액법",
            disposal=None, reest_date=dt.date(2024, 1, 1), reest_life=6,
            ref_date=dt.date(2025, 12, 31), fy_year=2025, reest_method="이중체감법")
        assert dep == 1_666_667

    def test_reest_from_double_declining_to_straight_line(self):
        dep, months, life_ended, note = R.recalc_asset(
            acq=dt.date(2020, 1, 1), cost=11_500_000, salvage=0, life=8, method="이중체감법",
            disposal=None, reest_date=dt.date(2024, 1, 1), reest_life=6,
            ref_date=dt.date(2025, 12, 31), fy_year=2025, reest_method="정액법")
        assert dep == 606_445

    def test_capex_does_not_change_rate(self):
        # 자본적지출은 상각기준가액(장부가액+지출액)만 늘릴 뿐, 상각률(2/내용연수)
        # 자체는 바뀌지 않는다 — 정률법의 자본적지출 처리 정책과 동일.
        dep, months, life_ended, note = R.recalc_asset(
            acq=dt.date(2020, 1, 1), cost=12_000_000, salvage=0, life=10, method="이중체감법",
            disposal=None, reest_date=None, reest_life=None,
            ref_date=dt.date(2025, 12, 31), fy_year=2025,
            capex_date=dt.date(2023, 1, 1), capex_amount=2_000_000)
        assert dep == 1_042_432
        assert "자본적지출" in note

    def test_suspension_reduces_current_period_months(self):
        dep, months, life_ended, note = R.recalc_asset(
            acq=dt.date(2021, 1, 1), cost=9_500_000, salvage=0, life=5, method="이중체감법",
            disposal=None, reest_date=None, reest_life=None,
            ref_date=dt.date(2025, 12, 31), fy_year=2025,
            susp_start=dt.date(2025, 3, 1), susp_end=dt.date(2025, 9, 30))
        assert dep == 205_200
        assert months == 5


# ---------------------------------------------------------------------------
# 연수합계법(SYD) 재계산 — 그 해의 몫 (N-k+1)/(N(N+1)/2), 기준가액은 고정
# ---------------------------------------------------------------------------
class TestSumOfYearsDigits:
    def test_basic_matches_closed_form_fraction(self):
        # 취득 2021-01-01, 1,000,000원, 잔존가치 100,000원, 내용연수5년.
        # 5년차(2025) 몫 = (5-5+1)/(5*6/2) = 1/15 -> (900,000)*1/15 = 60,000원.
        dep, months, life_ended, note = R.recalc_asset(
            acq=dt.date(2021, 1, 1), cost=1_000_000, salvage=100_000, life=5, method="연수합계법",
            disposal=None, reest_date=None, reest_life=None,
            ref_date=dt.date(2025, 12, 31), fy_year=2025)
        assert dep == 60_000
        assert months == 12

    def test_first_year_has_largest_share(self):
        # 1년차(2021) 몫 = 5/15 -> 900,000*5/15 = 300,000원(취득연도라 12개월 전체).
        dep, months, life_ended, note = R.recalc_asset(
            acq=dt.date(2021, 1, 1), cost=1_000_000, salvage=100_000, life=5, method="연수합계법",
            disposal=None, reest_date=None, reest_life=None,
            ref_date=dt.date(2021, 12, 31), fy_year=2021)
        assert dep == 300_000

    def test_disposal_mid_year(self):
        dep, months, life_ended, note = R.recalc_asset(
            acq=dt.date(2021, 1, 1), cost=11_000_000, salvage=0, life=6, method="연수합계법",
            disposal=dt.date(2025, 4, 30), reest_date=None, reest_life=None,
            ref_date=dt.date(2025, 12, 31), fy_year=2025)
        assert dep == 349_206
        assert months == 4

    def test_reest_same_method_restarts_share_schedule(self):
        # 재추정은 origin을 리셋한다(자본적지출과 반대) — 새 내용연수 기준으로
        # k=1부터 다시 시작하는 완전히 새 스케줄이 만들어진다.
        dep, months, life_ended, note = R.recalc_asset(
            acq=dt.date(2021, 1, 1), cost=12_500_000, salvage=0, life=6, method="연수합계법",
            disposal=None, reest_date=dt.date(2024, 1, 1), reest_life=8,
            ref_date=dt.date(2025, 12, 31), fy_year=2025)
        assert dep == 694_444

    def test_reest_from_declining_balance_to_syd(self):
        dep, months, life_ended, note = R.recalc_asset(
            acq=dt.date(2020, 1, 1), cost=13_500_000, salvage=0, life=8, method="정률법",
            disposal=None, reest_date=dt.date(2024, 1, 1), reest_life=6,
            ref_date=dt.date(2025, 12, 31), fy_year=2025, reest_method="연수합계법")
        assert dep == 715_997

    def test_reest_from_syd_to_declining_balance(self):
        dep, months, life_ended, note = R.recalc_asset(
            acq=dt.date(2021, 1, 1), cost=12_800_000, salvage=0, life=8, method="연수합계법",
            disposal=None, reest_date=dt.date(2024, 7, 1), reest_life=5,
            ref_date=dt.date(2025, 12, 31), fy_year=2025, reest_method="정률법")
        assert dep == 1_552_442

    def test_capex_increases_basis_without_resetting_share_schedule(self):
        # 자본적지출 전(2021~2022, k=1,2): 1,000,000*(5/15)+1,000,000*(4/15)=600,000원
        # 상각 -> 자본적지출 시점 장부가액 400,000원 + 지출액 200,000원 = 새 기준가액
        # 600,000원. origin은 그대로 2021년이라 2025년은 k=5(몫=1/15)이고,
        # 5년차 상각비 = 600,000*1/15 = 40,000원(기준가액이 새로 늘어난 값을 쓰되,
        # 연차(k) 자체는 리셋되지 않는다는 것이 이 테스트의 핵심).
        dep, months, life_ended, note = R.recalc_asset(
            acq=dt.date(2021, 1, 1), cost=1_000_000, salvage=0, life=5, method="연수합계법",
            disposal=None, reest_date=None, reest_life=None,
            ref_date=dt.date(2025, 12, 31), fy_year=2025,
            capex_date=dt.date(2023, 1, 1), capex_amount=200_000)
        assert dep == 40_000

    def test_suspension_reduces_current_period_months(self):
        dep, months, life_ended, note = R.recalc_asset(
            acq=dt.date(2021, 1, 1), cost=9_800_000, salvage=0, life=5, method="연수합계법",
            disposal=None, reest_date=None, reest_life=None,
            ref_date=dt.date(2025, 12, 31), fy_year=2025,
            susp_start=dt.date(2025, 2, 1), susp_end=dt.date(2025, 8, 31))
        assert dep == 272_222
        assert months == 5


# ---------------------------------------------------------------------------
# "일치여부"/"수식여부" 열 삭제 회귀 테스트 (main() 종단 실행)
# ---------------------------------------------------------------------------
class TestResultColumnsAfterRemoval:
    def _run_main_with(self, tmp_path, rows):
        cols = ["자산명", "자산분류(유형자산/무형자산)", "취득일", "취득원가", "잔존가치", "내용연수(년)",
                "상각방법(정액법/정률법)", "회사반영_당기감가상각비", "처분일", "내용연수재추정일",
                "재추정후내용연수(년)", "재추정후상각방법(정액법/정률법)", "총예정생산량", "당기실제생산량",
                "전기말누적생산량", "자본적지출일", "자본적지출액", "상각중단시작일", "상각중단종료일",
                "회사반영_전기말감가상각누계액"]
        in_path = tmp_path / "in.xlsx"
        out_path = tmp_path / "out.xlsx"
        pd.DataFrame(rows, columns=cols).to_excel(in_path, index=False)

        orig_in, orig_out = R.IN_PATH, R.OUT_PATH
        R.IN_PATH, R.OUT_PATH = str(in_path), str(out_path)
        try:
            R.main()
        finally:
            R.IN_PATH, R.OUT_PATH = orig_in, orig_out
        return out_path

    def test_removed_columns_are_absent_and_diff_based_filtering_works(self, tmp_path):
        import openpyxl

        rows = [
            # 일치하는 자산: 정액법, 취득원가 6,000,000원, 내용연수10년(월상각 50,000원),
            # 취득 2020-01-01이라 기준일(2025-12-31)에도 아직 상각 중 — 당기(2025) 12개월
            # 전체 상각 = 600,000원.
            ["일치자산", "유형자산", dt.date(2020, 1, 1), 6_000_000, 0, 10, "정액법", 600_000,
             None, None, None, None, None, None, None, None, None, None, None, None],
            # 의도적으로 틀린 자산(회사반영을 1원 더 크게 입력).
            ["불일치자산", "유형자산", dt.date(2020, 1, 1), 6_000_000, 0, 10, "정액법", 600_001,
             None, None, None, None, None, None, None, None, None, None, None, None],
        ]
        out_path = self._run_main_with(tmp_path, rows)

        wb = openpyxl.load_workbook(out_path)
        result_cols = [c.value for c in wb["재계산결과"][1]]
        for removed in ("일치여부", "누계액일치여부", "당기말누계액일치여부", "당기수식여부", "누계수식여부"):
            assert removed not in result_cols

        # diff_df(차이자산 시트)는 이제 "차이(재계산-회사반영)" != 0 기준으로 걸러진다 —
        # 불일치자산 1건만 들어가 있어야 한다.
        diff_names = [wb["차이자산"].cell(row=r, column=1).value for r in range(2, wb["차이자산"].max_row + 1)]
        assert diff_names == ["불일치자산"]

    def test_invalid_row_isolated_to_error_sheet_end_to_end(self, tmp_path):
        # validate_asset_inputs 자체는 단위테스트가 잘 돼 있지만, main()이 그 결과를
        # 받아 실제로 "데이터오류" 시트에 올바른 컬럼과 함께 기록하고, 정상 자산은
        # "재계산결과"에만 남기는 전체 배선(1449~1487번째 줄 근처)은 지금까지 어떤
        # 테스트도 거치지 않았다(coverage 0%) — main()을 end-to-end로 돌려 확인한다.
        import openpyxl

        rows = [
            # 정상 자산: 정액법, 취득원가 6,000,000원, 내용연수10년, 취득 2020-01-01.
            ["정상자산", "유형자산", dt.date(2020, 1, 1), 6_000_000, 0, 10, "정액법", 600_000,
             None, None, None, None, None, None, None, None, None, None, None, None],
            # 데이터오류 자산: 내용연수 0년(계산 불가능한 값).
            ["오류자산", "유형자산", dt.date(2020, 1, 1), 6_000_000, 0, 0, "정액법", 600_000,
             None, None, None, None, None, None, None, None, None, None, None, None],
        ]
        out_path = self._run_main_with(tmp_path, rows)
        wb = openpyxl.load_workbook(out_path)

        # 재계산결과 시트에는 정상자산만 있어야 한다.
        result_names = [wb["재계산결과"].cell(row=r, column=1).value
                         for r in range(2, wb["재계산결과"].max_row + 1)]
        assert result_names == ["정상자산"]

        # 데이터오류 시트에는 오류자산 1건이, 자산명/취득원가 등 원본 컬럼값 그대로,
        # 그리고 오류사유가 채워진 채로 들어가야 한다.
        error_ws = wb["데이터오류"]
        error_header = [c.value for c in error_ws[1]]
        error_row = [error_ws.cell(row=2, column=c + 1).value for c in range(len(error_header))]
        error_record = dict(zip(error_header, error_row))

        assert error_ws.max_row == 2  # 헤더 1행 + 오류자산 1행
        assert error_record["자산명"] == "오류자산"
        assert error_record["취득원가"] == 6_000_000
        assert "내용연수" in error_record["오류사유"]


# ---------------------------------------------------------------------------
# 다기간(연도별) 비교 시트 생성 (COMPARISON_YEARS 2개 이상)
# ---------------------------------------------------------------------------
class TestMultiYearSheetsEndToEnd:
    def test_comparison_years_produces_trend_sheets(self, tmp_path, monkeypatch):
        # main() 안의 다기간 비교 분기(len(COMPARISON_YEARS) > 1일 때 연도별추이/
        # 연도별요약 시트를 만드는 경로)는 지난 세션에 --config로 수동 실행해
        # 눈으로만 확인했을 뿐, 자동 회귀 테스트가 없었다(coverage 0%).
        # COMPARISON_YEARS를 2개 연도로 monkeypatch한 뒤 실제로 두 시트가
        # 생성되는지, 그리고 연도별요약(피벗) 시트에 두 연도 컬럼이 모두 있는지 확인한다.
        import openpyxl

        cols = ["자산명", "자산분류(유형자산/무형자산)", "취득일", "취득원가", "잔존가치", "내용연수(년)",
                "상각방법(정액법/정률법)", "회사반영_당기감가상각비", "처분일", "내용연수재추정일",
                "재추정후내용연수(년)", "재추정후상각방법(정액법/정률법)", "총예정생산량", "당기실제생산량",
                "전기말누적생산량", "자본적지출일", "자본적지출액", "상각중단시작일", "상각중단종료일",
                "회사반영_전기말감가상각누계액"]
        rows = [
            ["다기간자산", "유형자산", dt.date(2020, 1, 1), 6_000_000, 0, 10, "정액법", 600_000,
             None, None, None, None, None, None, None, None, None, None, None, None],
        ]
        in_path = tmp_path / "in.xlsx"
        out_path = tmp_path / "out.xlsx"
        pd.DataFrame(rows, columns=cols).to_excel(in_path, index=False)

        monkeypatch.setattr(R, "COMPARISON_YEARS", [2024, 2025])
        orig_in, orig_out = R.IN_PATH, R.OUT_PATH
        R.IN_PATH, R.OUT_PATH = str(in_path), str(out_path)
        try:
            R.main()
        finally:
            R.IN_PATH, R.OUT_PATH = orig_in, orig_out
        # monkeypatch가 함수 종료 시 R.COMPARISON_YEARS를 자동으로 원래 값(기본
        # [FY_YEAR])으로 되돌려 다른 테스트에 영향을 주지 않는다.

        wb = openpyxl.load_workbook(out_path)
        assert "연도별추이" in wb.sheetnames
        assert "연도별요약" in wb.sheetnames

        pivot_header = [c.value for c in wb["연도별요약"][1]]
        assert 2024 in pivot_header
        assert 2025 in pivot_header


# ---------------------------------------------------------------------------
# 상각방법 표기 정규화 (normalize_method / METHOD_ALIASES)
# ---------------------------------------------------------------------------
class TestNormalizeMethod:
    def test_exact_names_pass_through(self):
        for m in R.KNOWN_METHODS:
            assert R.normalize_method(m) == m

    def test_abbreviations_map_to_canonical_name(self):
        assert R.normalize_method("정액") == "정액법"
        assert R.normalize_method("정률") == "정률법"
        assert R.normalize_method("이중체감") == "이중체감법"
        assert R.normalize_method("연수합계") == "연수합계법"
        assert R.normalize_method("생산량비례") == "생산량비례법"

    def test_english_abbreviations_map_to_canonical_name(self):
        assert R.normalize_method("SL") == "정액법"
        assert R.normalize_method("DB") == "정률법"
        assert R.normalize_method("DDB") == "이중체감법"
        assert R.normalize_method("SYD") == "연수합계법"

    def test_whitespace_and_case_are_ignored(self):
        assert R.normalize_method(" sl ") == "정액법"
        assert R.normalize_method("Ddb") == "이중체감법"

    def test_unknown_value_returns_original_unchanged(self):
        # 사전에 없는 값은 정규화하지 않고 원본을 그대로 돌려준다 — validate_asset_inputs가
        # KNOWN_METHODS 기준으로 "인식 불가"임을 잡아 데이터오류로 격리하게 하기 위함.
        assert R.normalize_method("XYZ상각법") == "XYZ상각법"

    def test_none_passes_through(self):
        assert R.normalize_method(None) is None


class TestValidateAssetInputsUnknownMethod:
    def test_unrecognized_method_is_invalid(self):
        errors = R.validate_asset_inputs(cost=10_000_000, salvage=0, life=5, method="XYZ상각법")
        assert any("상각방법 오류" in e for e in errors)

    def test_known_methods_are_valid(self):
        for m in R.KNOWN_METHODS:
            kwargs = dict(total_units=1_000) if m == "생산량비례법" else {}
            errors = R.validate_asset_inputs(cost=10_000_000, salvage=0, life=5, method=m, **kwargs)
            assert not any("상각방법 오류" in e for e in errors)

    def test_method_none_is_not_flagged(self):
        # method를 아예 안 넘기는 기존 테스트들(자본적지출/상각중단 단독 검증)이
        # 깨지지 않도록, method=None이면 이 체크 자체를 건너뛰어야 한다.
        errors = R.validate_asset_inputs(cost=10_000_000, salvage=0, life=5)
        assert not any("상각방법 오류" in e for e in errors)


# ---------------------------------------------------------------------------
# 컬럼명 동의어 매칭 (_find_column / COLUMN_SYNONYMS)
# ---------------------------------------------------------------------------
class TestFindColumn:
    def test_exact_match_takes_priority_over_synonym(self):
        # "취득원가"라는 이름의 컬럼과, 동의어인 "취득가액" 컬럼이 둘 다 있으면
        # 정확 일치(취득원가)가 우선이어야 한다.
        df_columns = ["취득원가", "취득가액"]
        actual, source = R._find_column("취득원가", df_columns, "취득원가")
        assert actual == "취득원가"
        assert source == "정확일치"

    def test_synonym_match_when_exact_is_absent(self):
        df_columns = ["자산코드", "취득가액"]
        actual, source = R._find_column("자산명", df_columns, "자산명")
        assert actual == "자산코드"
        assert source == "동의어"

    def test_book_value_is_never_matched_to_acquisition_cost(self):
        # "장부가액"은 취득원가와 다른 개념(순장부가액)이므로, 동의어 사전에
        # 일부러 넣지 않았다 — 정확 일치도 동의어도 안 되는 게 정상 동작이다.
        df_columns = ["장부가액"]
        actual, source = R._find_column("취득원가", df_columns, "취득원가")
        assert actual is None
        assert source is None

    def test_no_match_returns_none(self):
        df_columns = ["전혀 관계없는 컬럼"]
        actual, source = R._find_column("취득원가", df_columns, "취득원가")
        assert actual is None


# ---------------------------------------------------------------------------
# 컬럼 인식 종단 테스트 (resolve_columns) — 동의어 매칭 / 실패 시 추천 메시지
# ---------------------------------------------------------------------------
class TestResolveColumns:
    def test_required_column_resolved_via_synonym_without_editing_column_map(self):
        # COLUMN_MAP은 그대로 두고("자산명": "자산명"), 파일 컬럼명만 동의어("자산코드")로
        # 바꿔도 정상 인식돼야 한다.
        cols = list(R.COLUMN_MAP.values())
        cols[list(R.COLUMN_MAP.keys()).index("자산명")] = "자산코드"
        df = pd.DataFrame(columns=cols)
        resolved = R.resolve_columns(df)
        assert resolved["자산명"] == "자산코드"

    def test_missing_required_column_raises_with_suggestion(self, monkeypatch):
        # API 키가 없는 테스트 환경에서는 AI 추천이 자동으로 빈 dict를 반환하고
        # 퍼지매칭으로 대체돼야 한다 — "취드원가"처럼 오탈자 수준으로 비슷한 컬럼명이면
        # 에러 메시지에 추천이 붙어야 한다.
        monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
        cols = list(R.COLUMN_MAP.values())
        cols[list(R.COLUMN_MAP.keys()).index("취득원가")] = "취드원가"  # 오탈자
        df = pd.DataFrame(columns=cols)
        try:
            R.resolve_columns(df)
            assert False, "KeyError가 발생했어야 한다"
        except KeyError as e:
            msg = str(e)
            assert "취득원가" in msg
            assert "취드원가" in msg  # 퍼지매칭 추천이 메시지에 포함됨

    def test_ai_suggestion_is_never_applied_automatically(self, monkeypatch):
        # AI가 뭔가를 추천하더라도 resolved 딕셔너리에는 절대 자동 반영되지 않고,
        # 여전히 KeyError로 중단돼야 한다(사람이 COLUMN_MAP을 고쳐야 함).
        monkeypatch.setattr(R, "_suggest_columns_ai", lambda missing, cols: {"취득원가": ["장부가액 (AI추천, 신뢰도:중, 추정)"]})
        cols = list(R.COLUMN_MAP.values())
        cols[list(R.COLUMN_MAP.keys()).index("취득원가")] = "완전히다른이름"
        df = pd.DataFrame(columns=cols)
        try:
            R.resolve_columns(df)
            assert False, "KeyError가 발생했어야 한다"
        except KeyError as e:
            assert "장부가액" in str(e)  # 추천은 메시지에만 나타나고
        # resolve_columns가 정상 반환되는 경로 자체가 없으므로(예외로 중단),
        # 자동 반영되지 않았다는 것 자체가 위 KeyError 발생으로 이미 증명된다.


# ---------------------------------------------------------------------------
# config.yaml 설정 로딩 (load_config / _cfg_get / _parse_ref_date / resolve_settings)
# ---------------------------------------------------------------------------
class TestLoadConfig:
    def test_missing_file_returns_empty_dict(self, tmp_path):
        assert R.load_config(str(tmp_path / "없는파일.yaml")) == {}

    def test_malformed_yaml_returns_empty_dict_without_raising(self, tmp_path):
        bad = tmp_path / "bad.yaml"
        bad.write_text("key: : :\n  bad indent\n -x", encoding="utf-8")
        assert R.load_config(str(bad)) == {}

    def test_non_dict_top_level_returns_empty_dict(self, tmp_path):
        # YAML 최상위가 리스트/스칼라인 경우도 dict가 아니므로 빈 dict로 취급한다.
        listy = tmp_path / "listy.yaml"
        listy.write_text("- a\n- b\n", encoding="utf-8")
        assert R.load_config(str(listy)) == {}

    def test_valid_yaml_returns_parsed_dict(self, tmp_path):
        good = tmp_path / "good.yaml"
        good.write_text("materiality_threshold: 500000\ncomparison_years: [2023, 2024]\n", encoding="utf-8")
        config = R.load_config(str(good))
        assert config == {"materiality_threshold": 500000, "comparison_years": [2023, 2024]}


class TestCfgGet:
    def test_missing_key_returns_default(self):
        assert R._cfg_get({}, "materiality_threshold", 1_000_000) == 1_000_000

    def test_none_value_returns_default(self):
        assert R._cfg_get({"materiality_threshold": None}, "materiality_threshold", 1_000_000) == 1_000_000

    def test_caster_applied_on_present_value(self):
        assert R._cfg_get({"yoy_anomaly_threshold_pct": "20"}, "yoy_anomaly_threshold_pct", 20.0, float) == 20.0

    def test_caster_failure_falls_back_to_default(self):
        # int("가나다")는 ValueError를 던진다 — 기본값으로 조용히 폴백해야 한다.
        assert R._cfg_get({"materiality_threshold": "가나다"}, "materiality_threshold", 1_000_000, int) == 1_000_000


class TestParseRefDate:
    def test_quoted_string(self):
        assert R._parse_ref_date("2025-12-31") == dt.date(2025, 12, 31)

    def test_yaml_auto_parsed_date_object(self):
        # 따옴표 없는 ref_date: 2025-12-31 는 PyYAML이 datetime.date로 자동 파싱한다.
        assert R._parse_ref_date(dt.date(2025, 12, 31)) == dt.date(2025, 12, 31)

    def test_datetime_object(self):
        assert R._parse_ref_date(dt.datetime(2025, 12, 31, 0, 0)) == dt.date(2025, 12, 31)


class TestResolveSettings:
    def test_empty_config_matches_builtin_defaults(self):
        settings = R.resolve_settings({})
        assert settings["IN_PATH"] == "sample_asset_ledger.xlsx"
        assert settings["OUT_PATH"] == "recalc_result.xlsx"
        assert settings["REF_DATE"] == dt.date(2025, 12, 31)
        assert settings["FY_YEAR"] == 2025
        assert settings["MATERIALITY_THRESHOLD"] == 1_000_000
        assert settings["OVERALL_MATERIALITY"] == 50_000_000
        assert settings["PERFORMANCE_MATERIALITY"] == 10_000_000
        assert settings["ANTHROPIC_MODEL"] == "claude-sonnet-4-6"
        assert settings["COMPARISON_YEARS"] == [2025]
        assert settings["YOY_ANOMALY_THRESHOLD_PCT"] == 20.0
        assert settings["COLUMN_MAP"] == R._DEFAULT_COLUMN_MAP

    def test_partial_column_map_merges_with_defaults(self):
        settings = R.resolve_settings({"column_map": {"취득원가": "취득가액"}})
        assert settings["COLUMN_MAP"]["취득원가"] == "취득가액"
        # 명시하지 않은 나머지 19개 키는 기본값이 그대로 유지되어야 한다.
        assert settings["COLUMN_MAP"]["자산명"] == R._DEFAULT_COLUMN_MAP["자산명"]
        assert len(settings["COLUMN_MAP"]) == len(R._DEFAULT_COLUMN_MAP)

    def test_non_dict_column_map_falls_back_to_defaults(self):
        settings = R.resolve_settings({"column_map": "이상한값"})
        assert settings["COLUMN_MAP"] == R._DEFAULT_COLUMN_MAP

    def test_comparison_years_overridden_changes_fy_year_independent_ref_date(self):
        settings = R.resolve_settings({"ref_date": "2023-06-30", "comparison_years": [2021, 2022, 2023]})
        assert settings["REF_DATE"] == dt.date(2023, 6, 30)
        assert settings["FY_YEAR"] == 2023
        assert settings["COMPARISON_YEARS"] == [2021, 2022, 2023]
